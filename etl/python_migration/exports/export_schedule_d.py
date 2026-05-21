#!/usr/bin/env python3
"""
Export TFRS Schedule D records to Excel files.

Built for LCFS issue #4400 — recover Schedule D content directly from the TFRS
database for any organization whose TFRS IDIR download is unavailable, so
analysts can review / re-upload it via the LCFS Portal.

For each compliance report selected (by TFRS compliance_report.id, or by org
name pattern) the script writes
<out_dir>/Schedule_D_<org>_<year>_CR<id>.xlsx with four sheets:

  - Summary: report metadata + one row per Schedule D sheet (feedstock /
             fuel type / fuel class)
  - Inputs:  GHGenius cell-level inputs (worksheet, cell, value, units, desc)
  - Outputs: computed CI components (Fuel Dispensing, LUC, etc.)
  - Outputs (pivot): outputs pivoted with one sheet per row and components
                     across columns

Usage:
  # by org name pattern (ILIKE)
  python export_schedule_d.py --org '%<organization name fragment>%'

  # by explicit report ids (use --cr-ids OR --org, at least one is required)
  python export_schedule_d.py --cr-ids 934 2100 1738 3433 3854

Env (or defaults shown):
  TFRS_HOST=localhost  TFRS_PORT=5435  TFRS_DB=tfrs
  TFRS_USER=tfrs       TFRS_PASSWORD=tfrs
"""

import argparse
import os
import re
import sys

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def connect():
    return psycopg2.connect(
        host=os.environ.get("TFRS_HOST", "localhost"),
        port=int(os.environ.get("TFRS_PORT", "5435")),
        dbname=os.environ.get("TFRS_DB", "tfrs"),
        user=os.environ.get("TFRS_USER", "tfrs"),
        password=os.environ.get("TFRS_PASSWORD", "tfrs"),
    )


def list_reports(cur, org_pattern, cr_ids):
    if cr_ids:
        cur.execute(
            """
            SELECT cr.id, o.name, cp.description, cr.schedule_d_id,
                   ws.analyst_status_id, ws.director_status_id
              FROM compliance_report cr
              JOIN organization o        ON o.id = cr.organization_id
              JOIN compliance_period cp  ON cp.id = cr.compliance_period_id
              LEFT JOIN compliance_report_workflow_state ws ON ws.id = cr.status_id
             WHERE cr.schedule_d_id IS NOT NULL
               AND cr.id = ANY(%s)
             ORDER BY cp.description, cr.id
            """,
            [list(cr_ids)],
        )
    else:
        cur.execute(
            """
            SELECT cr.id, o.name, cp.description, cr.schedule_d_id,
                   ws.analyst_status_id, ws.director_status_id
              FROM compliance_report cr
              JOIN organization o        ON o.id = cr.organization_id
              JOIN compliance_period cp  ON cp.id = cr.compliance_period_id
              LEFT JOIN compliance_report_workflow_state ws ON ws.id = cr.status_id
             WHERE cr.schedule_d_id IS NOT NULL
               AND o.name ILIKE %s
             ORDER BY cp.description, cr.id
            """,
            [org_pattern],
        )
    return cur.fetchall()


def fetch_schedule_d(cur, schedule_d_id):
    cur.execute(
        """
        SELECT s.id, s.feedstock, ft.name AS fuel_type, fc.fuel_class
          FROM compliance_report_schedule_d_sheet s
          JOIN approved_fuel_type ft ON s.fuel_type_id = ft.id
          JOIN fuel_class fc         ON s.fuel_class_id = fc.id
         WHERE s.schedule_id = %s
         ORDER BY s.id
        """,
        [schedule_d_id],
    )
    sheets = cur.fetchall()
    sheet_ids = [s[0] for s in sheets]

    inputs = []
    outputs = []
    if sheet_ids:
        cur.execute(
            """
            SELECT sheet_id, worksheet_name, cell, value, units, description
              FROM compliance_report_schedule_d_sheet_input
             WHERE sheet_id = ANY(%s)
             ORDER BY sheet_id, worksheet_name, cell
            """,
            [sheet_ids],
        )
        inputs = cur.fetchall()

        cur.execute(
            """
            SELECT sheet_id, description, intensity
              FROM compliance_report_schedule_d_sheet_output
             WHERE sheet_id = ANY(%s)
             ORDER BY sheet_id, description
            """,
            [sheet_ids],
        )
        outputs = cur.fetchall()

    return sheets, inputs, outputs


def autosize(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def write_workbook(path, org, year, cr_id, sheets, inputs, outputs, analyst, director):
    wb = Workbook()
    bold = Font(bold=True)

    # Summary tab
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Organization", org])
    ws.append(["Compliance period", year])
    ws.append(["TFRS compliance_report.id", cr_id])
    ws.append(["Analyst status", analyst or ""])
    ws.append(["Director status", director or ""])
    ws.append([])
    headers = ["Sheet ID", "Feedstock", "Fuel type", "Fuel class"]
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = bold
    for s in sheets:
        ws.append([s[0], s[1], s[2], s[3]])
    autosize(ws)

    # Inputs tab
    ws = wb.create_sheet("Inputs")
    headers = ["Sheet ID", "Worksheet", "Cell", "Value", "Units", "Description"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
    for row in inputs:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    autosize(ws)

    # Outputs tab — also pivot to wide form for readability
    ws = wb.create_sheet("Outputs")
    headers = ["Sheet ID", "Component", "Intensity"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
    for row in outputs:
        sheet_id, desc, intensity = row
        ws.append([sheet_id, desc, float(intensity) if intensity is not None else None])
    ws.freeze_panes = "A2"
    autosize(ws)

    # Pivot
    ws = wb.create_sheet("Outputs (pivot)")
    components = []
    seen = set()
    for _, desc, _ in outputs:
        if desc not in seen:
            seen.add(desc)
            components.append(desc)
    ws.append(["Sheet ID", "Feedstock", "Fuel type", "Fuel class"] + components)
    for c in ws[1]:
        c.font = bold
    sheet_meta = {s[0]: (s[1], s[2], s[3]) for s in sheets}
    by_sheet = {}
    for sid, desc, intensity in outputs:
        by_sheet.setdefault(sid, {})[desc] = (
            float(intensity) if intensity is not None else None
        )
    for sid in sorted(by_sheet.keys()):
        meta = sheet_meta.get(sid, ("", "", ""))
        row = [sid, meta[0], meta[1], meta[2]]
        for comp in components:
            row.append(by_sheet[sid].get(comp))
        ws.append(row)
    ws.freeze_panes = "E2"
    autosize(ws)

    wb.save(path)


def safe_filename(s):
    return re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--org", default=None,
                   help="organization.name ILIKE pattern (e.g. '%%fuels%%')")
    p.add_argument("--cr-ids", nargs="+", type=int, default=None,
                   help="specific TFRS compliance_report.id values to export")
    p.add_argument("--out-dir", default="./schedule_d_exports",
                   help="output directory (default: ./schedule_d_exports)")
    args = p.parse_args()

    if not args.org and not args.cr_ids:
        p.error("must supply --org or --cr-ids")

    os.makedirs(args.out_dir, exist_ok=True)

    with connect() as conn:
        cur = conn.cursor()
        reports = list_reports(cur, args.org, args.cr_ids)
        if not reports:
            print("No Schedule D reports found for that filter.", file=sys.stderr)
            sys.exit(2)

        for cr_id, org, year, schedule_d_id, analyst, director in reports:
            sheets, inputs, outputs = fetch_schedule_d(cur, schedule_d_id)
            fname = f"Schedule_D_{safe_filename(org)}_{year}_CR{cr_id}.xlsx"
            path = os.path.join(args.out_dir, fname)
            write_workbook(
                path, org, year, cr_id, sheets, inputs, outputs, analyst, director
            )
            print(
                f"wrote {path}  ({len(sheets)} sheets, {len(inputs)} inputs, "
                f"{len(outputs)} outputs)"
            )


if __name__ == "__main__":
    main()
