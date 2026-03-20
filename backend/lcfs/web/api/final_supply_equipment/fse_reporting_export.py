import io
import datetime

from fastapi import Depends
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from starlette.responses import StreamingResponse

from lcfs.db.models import Organization, UserProfile
from lcfs.utils.constants import FILE_MEDIA_TYPE
from lcfs.web.api.compliance_report.repo import ComplianceReportRepository
from lcfs.web.api.final_supply_equipment.repo import FinalSupplyEquipmentRepository
from lcfs.web.core.decorators import service_handler
from lcfs.web.exception.exceptions import DataNotFoundException

FSE_UPDATE_EXPORT_FILENAME = "FSE_BulkUpdate"
FSE_UPDATE_SHEETNAME = "FSE"

HEADERS = [
    "Site name",
    "Registration #",
    "Serial #",
    "Dates of supply from",
    "Dates of supply to",
    "kWh usage",
    "Compliance notes",
]

COLUMN_WIDTHS = [25, 18, 18, 22, 22, 14, 40]


class FSEReportingExporter:
    def __init__(
        self,
        repo: FinalSupplyEquipmentRepository = Depends(FinalSupplyEquipmentRepository),
        compliance_report_repo: ComplianceReportRepository = Depends(),
    ) -> None:
        self.repo = repo
        self.compliance_report_repo = compliance_report_repo

    @service_handler
    async def export_update_template(
        self,
        compliance_report_id: int,
        user: UserProfile,
        organization: Organization,
    ) -> StreamingResponse:
        """
        Generates a pre-populated Excel template for bulk-updating FSE reporting data
        (supply dates, kWh usage, compliance notes) keyed on FSE Registration Number.
        Column A (Registration Number) is read-only; columns B–E are editable.
        """
        compliance_report = (
            await self.compliance_report_repo.get_compliance_report_by_id(
                compliance_report_id
            )
        )
        if not compliance_report:
            raise DataNotFoundException("Compliance report not found.")

        compliance_period = compliance_report.compliance_period
        group_uuid = compliance_report.compliance_report_group_uuid

        fse_rows = await self._load_fse_data(
            organization.organization_id, group_uuid
        )

        wb = self._build_workbook(fse_rows, compliance_period)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        compliance_period_desc = compliance_period.description
        filename = (
            f"{FSE_UPDATE_EXPORT_FILENAME}_{organization.name}"
            f"-{compliance_period_desc}.xlsx"
        )
        headers_dict = {"Content-Disposition": f'attachment; filename="{filename}"'}

        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type=FILE_MEDIA_TYPE["XLSX"].value,
            headers=headers_dict,
        )

    async def _load_fse_data(
        self, organization_id: int, compliance_report_group_uuid: str | None
    ) -> list:
        records = await self.repo.get_fse_for_bulk_update_template(
            organization_id=organization_id,
            compliance_report_group_uuid=compliance_report_group_uuid,
        )
        rows = []
        for record in records:
            # Show only registration number when:
            # - the row is explicitly marked inactive
            record_group = getattr(record, "compliance_report_group_uuid", None)
            is_inactive = (
                not getattr(record, "is_active", True)
                or record_group != compliance_report_group_uuid
            )
            serial_number = getattr(record, "serial_number", None) or ""
            if is_inactive:
                rows.append(
                    [
                        getattr(record, "site_name", None) or "",
                        record.registration_number or "",
                        serial_number,
                        None,
                        None,
                        None,
                        None,
                    ]
                )
                continue
            supply_from = record.supply_from_date
            supply_to = record.supply_to_date
            if isinstance(supply_from, datetime.datetime):
                supply_from = supply_from.date()
            if isinstance(supply_to, datetime.datetime):
                supply_to = supply_to.date()
            rows.append(
                [
                    getattr(record, "site_name", None) or "",
                    record.registration_number or "",
                    serial_number,
                    supply_from,
                    supply_to,
                    record.kwh_usage if record.kwh_usage is not None else 0,
                    record.compliance_notes,
                ]
            )
        return rows

    def _build_workbook(self, rows: list, compliance_period) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = FSE_UPDATE_SHEETNAME

        effective_date = compliance_period.effective_date
        expiration_date = compliance_period.expiration_date

        # Header row – bold only, no background (consistent with SpreadsheetBuilder)
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")

        for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Column layout:
        # Col A (1): Site name        – locked for existing rows, unlocked for new rows
        # Col B (2): Registration #   – locked for existing rows, unlocked for new rows
        # Col C (3): Serial #         – locked (read-only) always
        # Col D (4): Supply from      – editable, date
        # Col E (5): Supply to        – editable, date
        # Col F (6): kWh usage        – editable, integer
        # Col G (7): Compliance notes – editable, text

        # Pre-populated data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in (1, 2, 3):
                    # Site name, Registration # & Serial # – locked (read-only)
                    cell.protection = Protection(locked=True)
                    cell.alignment = Alignment(horizontal="left")
                elif col_idx in (4, 5):
                    cell.protection = Protection(locked=False)
                    cell.number_format = "yyyy-mm-dd"
                    cell.alignment = Alignment(horizontal="left")
                elif col_idx == 6:
                    cell.protection = Protection(locked=False)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.protection = Protection(locked=False)
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)

        # Empty rows for new entries – unlock editable columns; Serial # stays locked
        first_empty_row = len(rows) + 2
        for row_idx in range(first_empty_row, first_empty_row + 500):
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx == 3:
                    # Serial # – always locked
                    cell.protection = Protection(locked=True)
                else:
                    cell.protection = Protection(locked=False)
                if col_idx in (4, 5):
                    cell.number_format = "yyyy-mm-dd"
                elif col_idx == 6:
                    cell.number_format = "#,##0"

        # Data validators (allow blank – do not block partial updates)
        date_validator = DataValidation(
            type="date",
            operator="between",
            formula1=(
                f"DATE({effective_date.year},{effective_date.month},{effective_date.day})"
            ),
            formula2=(
                f"DATE({expiration_date.year},{expiration_date.month},{expiration_date.day})"
            ),
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Date",
            error=(
                f"Please enter a date within the {compliance_period.description} "
                "calendar year."
            ),
        )
        date_validator.add("D2:D10000")
        date_validator.add("E2:E10000")
        ws.add_data_validation(date_validator)

        kwh_validator = DataValidation(
            type="whole",
            allow_blank=True,
            showErrorMessage=True,
            error="Please enter a valid integer.",
        )
        kwh_validator.add("F2:F10000")
        ws.add_data_validation(kwh_validator)

        # Protect the sheet – only rows with locked=False cells will be editable
        ws.protection.sheet = True
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = False

        return wb
