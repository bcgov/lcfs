import asyncio
import concurrent.futures
import datetime
import io
import json
import re
import structlog
import uuid
from fastapi import Depends, HTTPException, status, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from typing import List

from lcfs.db.dependencies import db_url, set_user_context
from lcfs.db.models import UserProfile
from lcfs.services.clamav.client import ClamAVService
from lcfs.services.redis.dependency import get_redis_client
from lcfs.settings import settings
from lcfs.utils.constants import POSTAL_REGEX
from lcfs.web.api.compliance_report.repo import ComplianceReportRepository
from lcfs.web.api.final_supply_equipment.repo import FinalSupplyEquipmentRepository
from lcfs.web.api.final_supply_equipment.schema import (
    FinalSupplyEquipmentCreateSchema,
    PortsEnum,
)
from lcfs.web.api.final_supply_equipment.services import FinalSupplyEquipmentServices
from lcfs.web.api.fuel_supply.repo import FuelSupplyRepository
from lcfs.web.core.decorators import service_handler
from lcfs.web.exception.exceptions import DataNotFoundException

logger = structlog.get_logger(__name__)


class FinalSupplyEquipmentImporter:
    def __init__(
        self,
        repo: FinalSupplyEquipmentRepository = Depends(),
        fse_service: FinalSupplyEquipmentServices = Depends(),
        compliance_report_repo: ComplianceReportRepository = Depends(),
        clamav_service: ClamAVService = Depends(),
        redis_client: Redis = Depends(get_redis_client),
        executor: concurrent.futures.ThreadPoolExecutor = Depends(),
    ) -> None:
        self.repo = repo
        self.fse_service = fse_service
        self.compliance_report_repo = compliance_report_repo
        self.clamav_service = clamav_service
        self.redis_client = redis_client
        self.executor = executor

    @service_handler
    async def import_data(
        self,
        compliance_report_id: int,
        user: UserProfile,
        org_code: str,
        file: UploadFile,
        overwrite: bool,
    ) -> str:
        """
        Initiates the import job in a separate thread executor.
        Returns a job_id that can be used to track progress via get_status.
        """
        compliance_report = (
            await self.compliance_report_repo.get_compliance_report_by_id(
                compliance_report_id
            )
        )
        if not compliance_report:
            raise DataNotFoundException("Compliance report not found.")

        job_id = str(uuid.uuid4())

        # Initialize job status in Redis
        await _update_progress(
            self.redis_client, job_id, 0, "Starting import job...", 0, 0, []
        )

        # Read file into memory once to enable scanning and openpyxl parsing
        file_contents = await file.read()

        # Create an in-memory buffer and attach it to a new UploadFile
        buffer = io.BytesIO(file_contents)
        buffer.seek(0)
        copied_file = UploadFile(filename=file.filename, file=buffer)

        loop = asyncio.get_running_loop()
        loop.run_in_executor(
            self.executor,
            lambda: asyncio.run(
                import_async(
                    compliance_report_id,
                    user,
                    org_code,
                    copied_file,
                    job_id,
                    overwrite,
                )
            ),
        )

        return job_id

    async def get_status(self, job_id: str) -> dict:
        """
        Retrieves and returns the job's progress and status from Redis.
        """
        progress_data_str = await self.redis_client.get(f"jobs/{job_id}")
        if not progress_data_str:
            return {"progress": 0, "status": "No job found with this ID."}

        try:
            progress_data = json.loads(progress_data_str)
            return {
                "progress": progress_data.get("progress", 0),
                "status": progress_data.get("status", "No status available."),
                "created": progress_data.get("created", 0),
                "rejected": progress_data.get("rejected", 0),
                "errors": progress_data.get("errors", []),
            }
        except json.JSONDecodeError:
            return {"progress": 0, "status": "Invalid status data found."}


async def import_async(
    compliance_report_id: int,
    user: UserProfile,
    org_code: str,
    file: UploadFile,
    job_id: str,
    overwrite: bool,
):
    """
    Performs the actual import in an async context.
    This is offloaded to a thread pool to avoid blocking.
    """
    logger.debug("Importing FSE Data...")

    engine = create_async_engine(db_url, future=True)
    try:
        async with AsyncSession(engine) as session:
            async with session.begin():
                await set_user_context(session, user.keycloak_username)

                fs_repo = FuelSupplyRepository(session)
                fse_repo = FinalSupplyEquipmentRepository(session)
                cr_repo = ComplianceReportRepository(session, fs_repo)
                fse_service = FinalSupplyEquipmentServices(fse_repo, cr_repo)
                clamav_service = ClamAVService()
                redis_client = Redis(
                    host=settings.redis_host,
                    port=settings.redis_port,
                    password=settings.redis_pass,
                    db=settings.redis_base or 0,
                    decode_responses=True,
                    max_connections=10,
                    socket_timeout=5,
                    socket_connect_timeout=5,
                )

                await _update_progress(
                    redis_client, job_id, 5, "Initializing services..."
                )

                if overwrite:
                    await _update_progress(
                        redis_client, job_id, 10, "Deleting old data..."
                    )
                    await fse_service.delete_all(compliance_report_id)
                    await fse_repo.reset_seq_by_org(org_code)

                # Optional: Scan the file with ClamAV if enabled
                if settings.clamav_enabled:
                    await _update_progress(
                        redis_client, job_id, 15, "Scanning file with ClamAV..."
                    )
                    clamav_service.scan_file(file)

                await _update_progress(
                    redis_client, job_id, 20, "Loading Excel sheet..."
                )

                try:
                    sheet = _load_sheet(file)
                    row_count = sheet.max_row

                    created = 0
                    rejected = 0
                    errors = []

                    await _update_progress(
                        redis_client,
                        job_id,
                        20,
                        "Beginning data import...",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                    )

                    valid_intended_user_types = await fse_repo.get_intended_user_types()
                    valid_use_types = await fse_repo.get_intended_use_types()
                    valid_use_type_names = {obj.type for obj in valid_use_types}
                    valid_user_type_names = {
                        obj.type_name for obj in valid_intended_user_types
                    }

                    # Iterate through all data rows, skipping the header
                    for row_idx, row in enumerate(
                        sheet.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        # Only update progress in Redis every 10 rows for efficiency
                        if (row_idx + 1) % 10 == 0:
                            current_progress = 20 + ((row_idx / row_count) * 80)
                            status_msg = (
                                f"Importing row {row_idx - 1} of {row_count - 1}..."
                            )
                            await _update_progress(
                                redis_client,
                                job_id,
                                current_progress,
                                status_msg,
                                created=created,
                                rejected=rejected,
                                errors=errors,
                            )

                        # Check if the entire row is empty
                        if all(cell is None for cell in row):
                            continue

                        # Validate row
                        error = _validate_row(
                            row, row_idx, valid_use_type_names, valid_user_type_names
                        )
                        if error:
                            errors.append(error)
                            rejected += 1
                            continue

                        # Parse row data and insert into DB
                        try:
                            fse_data = _parse_row(row, compliance_report_id)
                            await fse_service.create_final_supply_equipment(
                                fse_data, org_code
                            )
                            created += 1
                        except Exception as ex:
                            logger.error(str(ex))
                            errors.append(f"Row {row_idx}: {ex}")
                            rejected += 1

                    # Final update at 100%
                    await _update_progress(
                        redis_client,
                        job_id,
                        100,
                        "Import process completed.",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                    )
                    logger.debug(
                        f"Completed importing FSE data, {created} rows created"
                    )

                    return {
                        "success": True,
                        "created": created,
                        "errors": errors,
                        "rejected": rejected,
                    }

                except DataNotFoundException as dnfe:
                    await _update_progress(
                        redis_client, job_id, 100, "Data not found error.", 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND, detail=str(dnfe)
                    )
                except Exception as e:
                    await _update_progress(
                        redis_client, job_id, 100, "Import process failed.", 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Could not import FSE data: {str(e)}",
                    )
    finally:
        await engine.dispose()


def _load_sheet(file: UploadFile) -> Worksheet:
    """
    Loads and returns the 'FSE' worksheet from the provided Excel file.
    Raises an exception if the sheet does not exist.
    """
    workbook = load_workbook(data_only=True, filename=file.file)
    if "FSE" not in workbook.sheetnames:
        raise Exception("Uploaded Excel does not contain a 'FSE' sheet.")
    return workbook["FSE"]


def _validate_row(
    row: tuple,
    row_idx: int,
    valid_use_types: set[str],
    valid_user_types: set[str],
) -> str | None:
    """
    Validates a single row of data and returns an error string if invalid.
    Returns None if the row is valid.
    """
    (
        org_name,
        supply_from_date,
        supply_to_date,
        kwh_usage,
        serial_number,
        manufacturer,
        model,
        level_of_equipment,
        ports,
        intended_use_types,
        intended_user_types,
        street_address,
        city,
        postal_code,
        latitude,
        longitude,
        notes,
    ) = row

    missing_fields = []
    if supply_from_date is None:
        missing_fields.append("Supply from date")
    if supply_to_date is None:
        missing_fields.append("Supply to date")
    if serial_number is None:
        missing_fields.append("Serial #")
    if manufacturer is None:
        missing_fields.append("Manufacturer")
    if level_of_equipment is None:
        missing_fields.append("Level of equipment")
    if street_address is None:
        missing_fields.append("Street address")
    if city is None:
        missing_fields.append("City")
    if postal_code is None:
        missing_fields.append("Postal code")
    if latitude is None:
        missing_fields.append("Latitude")
    if longitude is None:
        missing_fields.append("Longitude")
    if not intended_use_types or len(intended_use_types) < 4:
        missing_fields.append("Intended use")
    if not intended_user_types or len(intended_user_types) < 4:
        missing_fields.append("Intended users")

    if missing_fields:
        return f"Row {row_idx}: Missing required fields: {', '.join(missing_fields)}"

    # Validate postal code
    postal_code_pattern = re.compile(POSTAL_REGEX)
    if not postal_code_pattern.match(postal_code):
        return f"Row {row_idx}: Invalid postal code"

    # Validate intended uses
    invalid_uses = [use for use in intended_use_types if use not in valid_use_types]
    if invalid_uses:
        return f"Row {row_idx}: Invalid intended use(s): {', '.join(invalid_uses)}"

    # Validate intended users
    invalid_users = [user for user in intended_user_types if user not in valid_user_types]
    if invalid_users:
        return f"Row {row_idx}: Invalid intended user(s): {', '.join(invalid_users)}"

    return None


def _parse_row(
    row: tuple, compliance_report_id: int
) -> FinalSupplyEquipmentCreateSchema:
    """
    Parses a valid row into a FinalSupplyEquipmentCreateSchema object.
    """
    (
        org_name,
        supply_from_date,
        supply_to_date,
        kwh_usage,
        serial_number,
        manufacturer,
        model,
        level_of_equipment,
        ports,
        intended_use_types,
        intended_user_types,
        street_address,
        city,
        postal_code,
        latitude,
        longitude,
        notes,
    ) = row

    supply_from_date = _parse_date(supply_from_date)
    supply_to_date = _parse_date(supply_to_date)
    kwh_usage = int(kwh_usage) if kwh_usage else 0
    latitude = float(latitude) if latitude else 0.0
    longitude = float(longitude) if longitude else 0.0

    return FinalSupplyEquipmentCreateSchema(
        compliance_report_id=compliance_report_id,
        organization_name=org_name or "",
        supply_from_date=supply_from_date,
        supply_to_date=supply_to_date,
        kwh_usage=kwh_usage,
        serial_nbr=str(serial_number) or "",
        manufacturer=str(manufacturer) or "",
        model=str(model) or "",
        level_of_equipment=level_of_equipment or "",
        ports=PortsEnum(ports) if ports else None,
        intended_use_types=intended_use_types,
        intended_user_types=intended_user_types,
        street_address=str(street_address) or "",
        city=str(city) or "",
        postal_code=postal_code or "",
        latitude=latitude,
        longitude=longitude,
        notes=notes or "",
    )


def _parse_date(date_value) -> datetime.datetime:
    """
    Safely parses various date formats into a datetime object.
    """
    if isinstance(date_value, (datetime.date, datetime.datetime)):
        return date_value
    return datetime.datetime.strptime(str(date_value), "%Y-%m-%d")


async def _update_progress(
    redis_client: Redis,
    job_id: str,
    progress: float,
    status_msg: str,
    created: int = 0,
    rejected: int = 0,
    errors: List[str] = None,
):
    """
    Persists the job status and progress in Redis.
    """
    if errors is None:
        errors = []
    data = {
        "progress": progress,
        "status": status_msg,
        "created": created,
        "rejected": rejected,
        "errors": errors,
    }
    await redis_client.set(f"jobs/{job_id}", json.dumps(data), ex=60)
