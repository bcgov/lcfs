import asyncio
import concurrent.futures
import datetime
import io
import json
import structlog
import uuid
from typing import List

from fastapi import Depends, HTTPException, status, UploadFile
from openpyxl import load_workbook
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine

from lcfs.db.dependencies import db_url, set_user_context
from lcfs.db.models import UserProfile
from lcfs.db.models.compliance.ComplianceReportChargingEquipment import (
    ComplianceReportChargingEquipment,
)
from lcfs.services.clamav.client import ClamAVService
from lcfs.services.redis.dependency import get_redis_client
from lcfs.settings import settings
from lcfs.utils.constants import ALLOWED_MIME_TYPES, ALLOWED_FILE_TYPES, MAX_FILE_SIZE_BYTES, MAX_FILE_SIZE_MB
from lcfs.web.api.compliance_report.repo import ComplianceReportRepository
from lcfs.web.api.final_supply_equipment.repo import FinalSupplyEquipmentRepository
from lcfs.web.api.fuel_supply.repo import FuelSupplyRepository
from lcfs.web.api.organizations.repo import OrganizationsRepository
from lcfs.web.core.decorators import service_handler
from lcfs.web.exception.exceptions import DataNotFoundException

logger = structlog.get_logger(__name__)

FSE_UPDATE_SHEETNAME = "FSE"


class FSEReportingImporter:
    def __init__(
        self,
        repo: FinalSupplyEquipmentRepository = Depends(),
        compliance_report_repo: ComplianceReportRepository = Depends(),
        clamav_service: ClamAVService = Depends(),
        redis_client: Redis = Depends(get_redis_client),
        executor: concurrent.futures.ThreadPoolExecutor = Depends(),
    ) -> None:
        self.repo = repo
        self.compliance_report_repo = compliance_report_repo
        self.clamav_service = clamav_service
        self.redis_client = redis_client
        self.executor = executor

    @service_handler
    async def import_data(
        self,
        compliance_report_id: int,
        user: UserProfile,
        file: UploadFile,
    ) -> str:
        """
        Initiates the bulk-update import job in a thread executor.
        Returns a job_id for progress polling via get_status().
        """
        compliance_report = (
            await self.compliance_report_repo.get_compliance_report_by_id(
                compliance_report_id
            )
        )
        if not compliance_report:
            raise DataNotFoundException("Compliance report not found.")

        job_id = str(uuid.uuid4())
        await _update_progress(
            self.redis_client, job_id, 0, "Starting import job...", 0, 0, 0, []
        )

        file_contents = await file.read()

        if file.content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"File type '{file.content_type or 'unknown'}' is not allowed. "
                    f"Please upload files of the following types: {ALLOWED_FILE_TYPES}"
                ),
            )

        if len(file_contents) > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"File size exceeds the maximum limit of {MAX_FILE_SIZE_MB} MB.",
            )

        buffer = io.BytesIO(file_contents)
        buffer.seek(0)
        copied_file = UploadFile(filename=file.filename, file=buffer)

        loop = asyncio.get_running_loop()
        loop.run_in_executor(
            self.executor,
            lambda: asyncio.run(
                _import_async(
                    compliance_report_id,
                    user,
                    compliance_report.compliance_report_group_uuid,
                    compliance_report.organization_id,
                    copied_file,
                    job_id,
                )
            ),
        )

        return job_id

    async def get_status(self, job_id: str) -> dict:
        progress_data_str = await self.redis_client.get(f"jobs/{job_id}")
        if not progress_data_str:
            return {"progress": 0, "status": "No job found with this ID."}

        try:
            data = json.loads(progress_data_str)
            return {
                "progress": data.get("progress", 0),
                "status": data.get("status", "No status available."),
                "created": data.get("created", 0),
                "rejected": data.get("rejected", 0),
                "skipped": data.get("skipped", 0),
                "errors": data.get("errors", []),
            }
        except json.JSONDecodeError:
            return {"progress": 0, "status": "Invalid status data found."}


async def _import_async(
    compliance_report_id: int,
    user: UserProfile,
    compliance_report_group_uuid: str,
    report_organization_id: int,
    file: UploadFile,
    job_id: str,
):
    """
    Performs the actual bulk-update import in an async context offloaded to a thread pool.
    """
    logger.debug("Importing FSE reporting bulk update...")

    engine = create_async_engine(db_url, future=True)
    try:
        async with AsyncSession(engine) as session:
            async with session.begin():
                await set_user_context(session, user.keycloak_username)

                fs_repo = FuelSupplyRepository(session)
                fse_repo = FinalSupplyEquipmentRepository(session)
                cr_repo = ComplianceReportRepository(session, fs_repo)
                redis_client = Redis(
                    host=settings.redis_host,
                    port=settings.redis_port,
                    password=settings.redis_pass,
                    db=settings.redis_base or 0,
                    decode_responses=True,
                    max_connections=10,
                    socket_timeout=5,
                    socket_connect_timeout=5,
                )

                await _update_progress(
                    redis_client, job_id, 5, "Initializing services..."
                )

                if settings.clamav_enabled:
                    await _update_progress(
                        redis_client, job_id, 10, "Scanning file with ClamAV..."
                    )
                    ClamAVService().scan_file(file)

                await _update_progress(
                    redis_client, job_id, 20, "Loading Excel sheet..."
                )

                try:
                    sheet = _load_sheet(file)
                    row_count = sheet.max_row

                    updated = 0
                    skipped = 0
                    rejected = 0
                    errors: List[str] = []

                    await _update_progress(
                        redis_client,
                        job_id,
                        20,
                        "Beginning data import...",
                        updated=updated,
                        skipped=skipped,
                        rejected=rejected,
                        errors=errors,
                    )

                    for row_idx, row in enumerate(
                        sheet.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        if (row_idx + 1) % 10 == 0:
                            current_progress = 20 + ((row_idx / row_count) * 80)
                            await _update_progress(
                                redis_client,
                                job_id,
                                current_progress,
                                f"Processing row {row_idx - 1} of {row_count - 1}...",
                                updated=updated,
                                skipped=skipped,
                                rejected=rejected,
                                errors=errors,
                            )

                        if all(cell is None for cell in row):
                            continue

                        expanded = list(row) + [None] * max(0, 5 - len(list(row)))
                        registration_number = expanded[0]
                        supply_from_raw = expanded[1]
                        supply_to_raw = expanded[2]
                        kwh_usage_raw = expanded[3]
                        compliance_notes = expanded[4]

                        # Registration number is required
                        if not registration_number:
                            errors.append(
                                f"Row {row_idx}: FSE Registration Number is required."
                            )
                            rejected += 1
                            continue

                        registration_number = str(registration_number).strip()

                        # Parse dates only when provided; blank dates are silently skipped
                        supply_from_date = None
                        supply_to_date = None
                        dates_provided = (
                            supply_from_raw is not None or supply_to_raw is not None
                        )
                        if dates_provided:
                            if supply_from_raw is not None:
                                try:
                                    supply_from_date = _parse_date(supply_from_raw)
                                except (ValueError, TypeError):
                                    errors.append(
                                        f"Row {row_idx}: Invalid 'Dates of supply from' "
                                        f"value: '{supply_from_raw}'."
                                    )
                                    rejected += 1
                                    continue

                            if supply_to_raw is not None:
                                try:
                                    supply_to_date = _parse_date(supply_to_raw)
                                except (ValueError, TypeError):
                                    errors.append(
                                        f"Row {row_idx}: Invalid 'Dates of supply to' "
                                        f"value: '{supply_to_raw}'."
                                    )
                                    rejected += 1
                                    continue

                            if (
                                supply_from_date is not None
                                and supply_to_date is not None
                                and supply_from_date > supply_to_date
                            ):
                                errors.append(
                                    f"Row {row_idx}: 'Dates of supply from' must not be "
                                    "after 'Dates of supply to'."
                                )
                                rejected += 1
                                continue

                        # Blank kWh skips the entire row — no other fields are updated
                        if kwh_usage_raw is None:
                            skipped += 1
                            continue

                        # Validate kWh usage
                        try:
                            kwh_usage = float(kwh_usage_raw)
                            if kwh_usage < 0:
                                raise ValueError("kWh must be non-negative")
                        except (ValueError, TypeError):
                            errors.append(
                                f"Row {row_idx}: Invalid kWh Usage value: "
                                f"'{kwh_usage_raw}'. Must be a numeric value."
                            )
                            rejected += 1
                            continue

                        # Look up the ChargingEquipment by registration number
                        equipment = (
                            await fse_repo.get_charging_equipment_by_registration_number(
                                registration_number=registration_number,
                                organization_id=report_organization_id,
                            )
                        )

                        if equipment is None:
                            errors.append(
                                f"Row {row_idx}: FSE Registration Number "
                                f"'{registration_number}' not found for this organization."
                            )
                            rejected += 1
                            continue

                        charging_equipment_id = equipment.charging_equipment_id
                        charging_equipment_version = equipment.charging_equipment_version

                        # Find existing ComplianceReportChargingEquipment record
                        existing_record = (
                            await fse_repo.get_fse_reporting_record_for_group(
                                charging_equipment_id=charging_equipment_id,
                                charging_equipment_version=charging_equipment_version,
                                compliance_report_group_uuid=compliance_report_group_uuid,
                            )
                        )

                        notes_value = (
                            str(compliance_notes).strip() if compliance_notes else None
                        )

                        if existing_record is None:
                            # Create new record only when dates are supplied
                            if supply_from_date is None or supply_to_date is None:
                                errors.append(
                                    f"Row {row_idx}: Registration number "
                                    f"'{registration_number}' has no existing reporting "
                                    "record. Provide both supply dates to create one."
                                )
                                rejected += 1
                                continue
                            new_record = ComplianceReportChargingEquipment(
                                charging_equipment_id=charging_equipment_id,
                                charging_equipment_version=charging_equipment_version,
                                compliance_report_id=compliance_report_id,
                                compliance_report_group_uuid=compliance_report_group_uuid,
                                organization_id=report_organization_id,
                                supply_from_date=supply_from_date,
                                supply_to_date=supply_to_date,
                                kwh_usage=kwh_usage,
                                compliance_notes=notes_value,
                                is_active=True,
                            )
                            session.add(new_record)
                            await session.flush()
                        else:
                            # Update existing record; blank fields are skipped.
                            # Always activate the row — the mere presence of the
                            # registration number in the upload checks it in the report.
                            await fse_repo.bulk_update_fse_reporting_record(
                                charging_equipment_compliance_id=(
                                    existing_record.charging_equipment_compliance_id
                                ),
                                supply_from_date=supply_from_date,
                                supply_to_date=supply_to_date,
                                kwh_usage=kwh_usage,
                                compliance_notes=notes_value,
                                activate=True,
                            )

                        updated += 1

                    await _update_progress(
                        redis_client,
                        job_id,
                        100,
                        "Import process completed.",
                        updated=updated,
                        skipped=skipped,
                        rejected=rejected,
                        errors=errors,
                    )
                    logger.debug(
                        f"FSE bulk update completed: {updated} updated, "
                        f"{skipped} skipped, {rejected} rejected"
                    )

                except DataNotFoundException as dnfe:
                    await _update_progress(
                        redis_client, job_id, 100, "Data not found error.", 0, 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND, detail=str(dnfe)
                    )
                except Exception as e:
                    await _update_progress(
                        redis_client, job_id, 100, "Import process failed.", 0, 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Could not process FSE bulk update: {str(e)}",
                    )
    finally:
        await engine.dispose()


def _load_sheet(file: UploadFile):
    """Loads the FSE_Update worksheet from the uploaded Excel file."""
    workbook = load_workbook(data_only=True, filename=file.file)
    if FSE_UPDATE_SHEETNAME not in workbook.sheetnames:
        raise Exception(
            f"Uploaded Excel does not contain an '{FSE_UPDATE_SHEETNAME}' sheet. "
            "Please use the template downloaded from this page."
        )
    return workbook[FSE_UPDATE_SHEETNAME]


def _parse_date(value) -> datetime.date:
    """Parses various date formats into a date object."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if value is None:
        raise ValueError("Date value is None")
    # Try common string formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date format: {value}")


async def _update_progress(
    redis_client: Redis,
    job_id: str,
    progress: float,
    status_msg: str,
    updated: int = 0,
    skipped: int = 0,
    rejected: int = 0,
    errors: List[str] | None = None,
):
    if errors is None:
        errors = []
    data = {
        "progress": progress,
        "status": status_msg,
        "created": updated,   # reuse 'created' key so frontend ImportDialog works
        "updated": updated,
        "skipped": skipped,
        "rejected": rejected,
        "errors": errors,
    }
    await redis_client.set(f"jobs/{job_id}", json.dumps(data), ex=60)
