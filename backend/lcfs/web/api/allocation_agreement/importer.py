import asyncio
import concurrent.futures
import io
import json
import re
import structlog
import uuid
from fastapi import Depends, HTTPException, status, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from typing import List

from lcfs.db.base import current_user_var
from lcfs.db.dependencies import db_url, set_user_context
from lcfs.db.models import UserProfile
from lcfs.services.clamav.client import ClamAVService
from lcfs.services.redis.dependency import get_redis_client
from lcfs.settings import settings
from lcfs.utils.constants import POSTAL_REGEX, ALLOWED_MIME_TYPES, ALLOWED_FILE_TYPES, MAX_FILE_SIZE_BYTES, MAX_FILE_SIZE_MB
from lcfs.web.api.compliance_report.services import ComplianceReportServices
from lcfs.web.api.allocation_agreement.repo import AllocationAgreementRepository
from lcfs.web.api.allocation_agreement.schema import (
    AllocationAgreementCreateSchema,
    ImportResultSchema,
    ImportSummarySchema,
    RejectedRowSchema,
)
from lcfs.web.api.allocation_agreement.services import AllocationAgreementServices
from lcfs.web.core.decorators import service_handler
from lcfs.web.exception.exceptions import DataNotFoundException
from lcfs.web.api.fuel_code.repo import FuelCodeRepository
from lcfs.web.api.fuel_supply.services import FuelSupplyServices
from lcfs.web.api.compliance_report.repo import ComplianceReportRepository
from lcfs.web.api.fuel_supply.repo import FuelSupplyRepository

logger = structlog.get_logger(__name__)


class AllocationAgreementImporter:

    def __init__(
        self,
        repo: AllocationAgreementRepository = Depends(),
        fuel_code_repo: FuelCodeRepository = Depends(),
        compliance_report_services: ComplianceReportServices = Depends(),
        clamav_service: ClamAVService = Depends(),
        redis_client: Redis = Depends(get_redis_client),
        executor: concurrent.futures.ThreadPoolExecutor = Depends(),
    ) -> None:
        self.repo = repo
        self.fuel_code_repo = fuel_code_repo
        self.compliance_report_services = compliance_report_services
        self.clamav_service = clamav_service
        self.redis_client = redis_client
        self.executor = executor

    @service_handler
    async def import_data(
        self,
        compliance_report_id: int,
        user: UserProfile,
        file: UploadFile,
        overwrite: bool,
    ) -> str:
        """
        Initiates the import job in a separate thread executor.
        Returns a job_id that can be used to track progress via get_status.
        """
        job_id = str(uuid.uuid4())

        # Initialize job status in Redis
        await _update_progress(
            self.redis_client, job_id, 0, "Starting import job...", 0, 0, []
        )

        # Read file into memory once to enable scanning and openpyxl parsing
        file_contents = await file.read()

        # Validate MIME type
        if file.content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"File type '{file.content_type or 'unknown'}' is not allowed. Please upload files of the following types: {ALLOWED_FILE_TYPES}",
            )

        # Validate file size
        file_size = len(file_contents)
        if file_size > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"File size exceeds the maximum limit of {MAX_FILE_SIZE_MB} MB.",
            )

        # Create an in-memory buffer and attach it to a new UploadFile
        buffer = io.BytesIO(file_contents)
        buffer.seek(0)
        copied_file = UploadFile(filename=file.filename, file=buffer)

        loop = asyncio.get_running_loop()
        loop.run_in_executor(
            self.executor,
            lambda: asyncio.run(
                import_async(
                    compliance_report_id,
                    user,
                    copied_file,
                    job_id,
                    overwrite,
                )
            ),
        )

        return job_id

    async def get_import_result(self, job_id: str) -> ImportResultSchema:
        """
        Get complete import results including rejected rows for frontend processing
        """
        redis_client = await self.redis_client.get_redis()
        progress_data_str = await redis_client.get(f"jobs/{job_id}")
        
        if progress_data_str is None:
            raise HTTPException(
                status_code=404,
                detail=f"Job with ID {job_id} not found or has expired",
            )
        
        progress_data = json.loads(progress_data_str)
        
        return ImportResultSchema(
            job_id=job_id,
            status=progress_data.get("status", "unknown"),
            summary=ImportSummarySchema(
                total=progress_data.get("created", 0) + progress_data.get("rejected", 0),
                imported=progress_data.get("created", 0),
                rejected=progress_data.get("rejected", 0)
            ),
            imported_rows=progress_data.get("imported_row_ids", []),
            rejected_rows=[
                RejectedRowSchema(**row) for row in progress_data.get("rejected_rows", [])
            ]
        )

    async def get_status(self, job_id: str) -> dict:
        """
        Retrieves and returns the job's progress and status from Redis.
        """
        progress_data_str = await self.redis_client.get(f"jobs/{job_id}")
        if not progress_data_str:
            return {"progress": 0, "status": "No job found with this ID."}

        try:
            progress_data = json.loads(progress_data_str)
            return {
                "progress": progress_data.get("progress", 0),
                "status": progress_data.get("status", "No status available."),
                "created": progress_data.get("created", 0),
                "rejected": progress_data.get("rejected", 0),
                "errors": progress_data.get("errors", []),
                "rejected_rows": progress_data.get("rejected_rows", []),
            }
        except json.JSONDecodeError:
            return {"progress": 0, "status": "Invalid status data found."}


async def import_async(
    compliance_report_id: int,
    user: UserProfile,
    file: UploadFile,
    job_id: str,
    overwrite: bool,
):
    """
    Performs the actual allocation agreement import in an async context.
    This is offloaded to a thread pool to avoid blocking.
    """
    logger.debug("Importing Allocation Agreement Data...")

    engine = create_async_engine(db_url, future=True)
    try:
        async with AsyncSession(engine) as session:
            async with session.begin():
                await set_user_context(session, user.keycloak_username)

                current_user_var.set(user)

                fuel_code_repo = FuelCodeRepository(session)
                compliance_report_repo = ComplianceReportRepository(session)
                aa_repo = AllocationAgreementRepository(
                    db=session, fuel_repo=fuel_code_repo
                )

                aa_service = AllocationAgreementServices(
                    repo=aa_repo,
                    fuel_repo=fuel_code_repo,
                    fuel_supply_service=FuelSupplyServices(
                        repo=FuelSupplyRepository(session),
                        fuel_repo=fuel_code_repo,
                        compliance_report_repo=compliance_report_repo,
                    ),
                    compliance_report_repo=compliance_report_repo,
                )

                redis_client = Redis(
                    host=settings.redis_host,
                    port=settings.redis_port,
                    password=settings.redis_pass,
                    db=settings.redis_base or 0,
                    decode_responses=True,
                    max_connections=10,
                    socket_timeout=5,
                    socket_connect_timeout=5,
                )
                clamav_service = ClamAVService()

                await _update_progress(
                    redis_client, job_id, 5, "Initializing services..."
                )

                if overwrite:
                    await _update_progress(
                        redis_client, job_id, 10, "Deleting old data..."
                    )
                    await aa_service.delete_all(
                        compliance_report_id, user.keycloak_username
                    )

                # Optional: Scan the file with ClamAV if enabled
                if settings.clamav_enabled:
                    await _update_progress(
                        redis_client, job_id, 15, "Scanning file with ClamAV..."
                    )
                    clamav_service.scan_file(file)

                await _update_progress(
                    redis_client, job_id, 20, "Loading Excel sheet..."
                )

                try:
                    table_options = await aa_repo.get_table_options(
                        # Use the actual compliance period from DB or a param
                        (
                            await compliance_report_repo.get_compliance_report_by_id(
                                compliance_report_id
                            )
                        ).compliance_period.description
                    )

                    # Build sets for validation
                    valid_fuel_types = {
                        obj["fuel_type"] for obj in table_options.get("fuel_types", [])
                    }
                    valid_fuel_categories = {
                        obj.category for obj in table_options.get("fuel_categories", [])
                    }
                    valid_provisions = {
                        obj.name
                        for obj in table_options.get("provisions_of_the_act", [])
                    }

                    sheet = _load_sheet(file)
                    row_count = sheet.max_row

                    created = 0
                    rejected = 0
                    errors = []
                    rejected_rows = []
                    imported_row_ids = []

                    await _update_progress(
                        redis_client,
                        job_id,
                        20,
                        "Beginning data import...",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                        rejected_rows=rejected_rows,
                        imported_row_ids=imported_row_ids,
                    )

                    # Iterate through data rows, skipping the header.
                    for row_idx, row in enumerate(
                        sheet.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        # Only update progress in Redis every 10 rows for efficiency
                        if (row_idx + 1) % 10 == 0:
                            current_progress = 20 + ((row_idx / row_count) * 80)
                            status_msg = (
                                f"Importing row {row_idx - 1} of {row_count - 1}..."
                            )
                            await _update_progress(
                                redis_client,
                                job_id,
                                current_progress,
                                status_msg,
                                created=created,
                                rejected=rejected,
                                errors=errors,
                                rejected_rows=rejected_rows,
                                imported_row_ids=imported_row_ids,
                            )

                        # Check if the entire row is empty
                        if all(cell is None for cell in row):
                            continue

                        # Validate row with detailed error information
                        validation_errors = _validate_row_detailed(
                            row,
                            row_idx,
                            valid_fuel_types,
                            valid_fuel_categories,
                            valid_provisions,
                        )
                        
                        if validation_errors:
                            # Row has validation errors - preserve for frontend editing
                            row_data = _row_to_dict(row)
                            rejected_rows.append({
                                "row_number": row_idx,
                                "data": row_data,
                                "errors": validation_errors
                            })
                            # Also add simple error message for legacy compatibility
                            error_messages = [error["message"] for error in validation_errors]
                            errors.append(f"Row {row_idx}: {'; '.join(error_messages)}")
                            rejected += 1
                            continue

                        # Parse row data and insert into DB
                        try:
                            aa_data = _parse_row(row, compliance_report_id)
                            result = await aa_service.create_allocation_agreement(aa_data)
                            imported_row_ids.append(result.allocation_agreement_id)
                            created += 1
                        except Exception as ex:
                            logger.error(str(ex))
                            # Database error - also preserve row data
                            row_data = _row_to_dict(row)
                            rejected_rows.append({
                                "row_number": row_idx,
                                "data": row_data,
                                "errors": [{"field": "general", "message": str(ex)}]
                            })
                            errors.append(f"Row {row_idx}: {ex}")
                            rejected += 1

                    # Final update at 100%
                    await _update_progress(
                        redis_client,
                        job_id,
                        100,
                        "Import process completed.",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                        rejected_rows=rejected_rows,
                        imported_row_ids=imported_row_ids,
                    )
                    logger.debug(
                        f"Completed importing Allocation Agreement data, {created} rows created"
                    )
                    return {
                        "success": True,
                        "created": created,
                        "errors": errors,
                        "rejected": rejected,
                    }

                except DataNotFoundException as dnfe:
                    await _update_progress(
                        redis_client, job_id, 100, "Data not found error.", 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND, detail=str(dnfe)
                    )
                except Exception as e:
                    await _update_progress(
                        redis_client, job_id, 100, "Import process failed.", 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Could not import allocation agreement data: {str(e)}",
                    )
    finally:
        await engine.dispose()


def _load_sheet(file: UploadFile) -> Worksheet:
    """
    Loads and returns the 'Allocation Agreement' worksheet from the provided Excel file.
    Raises an exception if the sheet does not exist.
    """
    workbook = load_workbook(data_only=True, filename=file.file)
    sheet_name = "Allocation Agreements"
    if sheet_name not in workbook.sheetnames:
        raise Exception(f"Uploaded Excel does not contain a '{sheet_name}' sheet.")
    return workbook[sheet_name]


def _validate_row(
    row: tuple,
    row_idx: int,
    valid_fuel_types: set,
    valid_fuel_categories: set,
    valid_provisions: set,
) -> str | None:
    """
    Validates a single row for allocation agreement import.
    Expected columns:
      0: Responsibility
      1: Legal name of transaction partner
      2: Address for service
      3: Email
      4: Phone
      5: Fuel type
      6: Fuel type other
      7: Fuel category
      8: Determining Carbon Intensity
      9: Fuel code
      10: Quantity
      11: Units
    """
    (
        responsibility,
        transaction_partner,
        postal_address,
        email,
        phone,
        fuel_type,
        fuel_type_other,
        fuel_category,
        provision,
        fuel_code,
        quantity,
    ) = row

    missing_fields = []
    if responsibility is None:
        missing_fields.append("Responsibility")
    if transaction_partner is None:
        missing_fields.append("Legal name of transaction partner")
    if postal_address is None:
        missing_fields.append("Address for service")
    if email is None:
        missing_fields.append("Email")
    if phone is None:
        missing_fields.append("Phone")
    if fuel_type is None:
        missing_fields.append("Fuel type")
    if fuel_category is None:
        missing_fields.append("Fuel category")
    if provision is None:
        missing_fields.append("Determining Carbon Intensity")
    if quantity is None:
        missing_fields.append("Quantity")

    if missing_fields:
        return f"Row {row_idx}: Missing required fields: {', '.join(missing_fields)}"

    # Validate email address.
    email_pattern = re.compile(r"^[\w\.-]+@[\w\.-]+\.\w+$")
    if not email_pattern.match(email):
        return f"Row {row_idx}: Invalid email address"

    # Validate quantity (must be integer > 0).
    try:
        if int(quantity) <= 0:
            return f"Row {row_idx}: Quantity must be greater than 0"
    except Exception:
        return f"Row {row_idx}: Quantity must be a number"

    # Validate lookups for fuel type, fuel category, and provision.
    if fuel_type not in valid_fuel_types:
        return f"Row {row_idx}: Invalid fuel type: {fuel_type}"

    if fuel_category not in valid_fuel_categories:
        return f"Row {row_idx}: Invalid fuel category: {fuel_category}"

    if provision not in valid_provisions:
        return f"Row {row_idx}: Invalid determining carbon intensity: {provision}"

    return None


def _validate_row_detailed(
    row: tuple,
    row_idx: int,
    valid_fuel_types: set,
    valid_fuel_categories: set,
    valid_provisions: set,
) -> List[dict]:
    """
    Validates a single row and returns detailed validation errors.
    Returns a list of validation error dictionaries.
    """
    (
        responsibility,
        transaction_partner,
        postal_address,
        email,
        phone,
        fuel_type,
        fuel_type_other,
        fuel_category,
        provision,
        fuel_code,
        quantity,
    ) = row

    errors = []

    # Check required fields
    field_mappings = {
        "allocationTransactionType": responsibility,
        "transactionPartner": transaction_partner,
        "postalAddress": postal_address,
        "transactionPartnerEmail": email,
        "transactionPartnerPhone": phone,
        "fuelType": fuel_type,
        "fuelCategory": fuel_category,
        "provisionOfTheAct": provision,
        "quantity": quantity,
    }

    for field_name, field_value in field_mappings.items():
        if field_value is None or field_value == "":
            errors.append({
                "field": field_name,
                "message": f"{field_name} is required"
            })

    # Skip further validation if any required fields are missing
    if errors:
        return errors

    # Validate email address
    email_pattern = re.compile(r"^[\w\.-]+@[\w\.-]+\.\w+$")
    if not email_pattern.match(email):
        errors.append({
            "field": "transactionPartnerEmail",
            "message": "Invalid email address format"
        })

    # Validate quantity (must be integer > 0)
    try:
        if int(quantity) <= 0:
            errors.append({
                "field": "quantity",
                "message": "Quantity must be greater than 0"
            })
    except (ValueError, TypeError):
        errors.append({
            "field": "quantity",
            "message": "Quantity must be a valid number"
        })

    # Validate dropdown values
    if fuel_type not in valid_fuel_types:
        errors.append({
            "field": "fuelType",
            "message": f"Invalid fuel type: {fuel_type}. Must be one of: {', '.join(list(valid_fuel_types)[:3])}..."
        })

    if fuel_category not in valid_fuel_categories:
        errors.append({
            "field": "fuelCategory",
            "message": f"Invalid fuel category: {fuel_category}. Must be one of: {', '.join(list(valid_fuel_categories)[:3])}..."
        })

    if provision not in valid_provisions:
        errors.append({
            "field": "provisionOfTheAct",
            "message": f"Invalid determining carbon intensity: {provision}. Must be one of: {', '.join(list(valid_provisions)[:3])}..."
        })

    return errors


def _row_to_dict(row: tuple) -> dict:
    """
    Converts a row tuple to a dictionary with field names.
    """
    (
        responsibility,
        transaction_partner,
        postal_address,
        email,
        phone,
        fuel_type,
        fuel_type_other,
        fuel_category,
        provision,
        fuel_code,
        quantity,
    ) = row

    return {
        "allocationTransactionType": responsibility,
        "transactionPartner": transaction_partner,
        "postalAddress": postal_address,
        "transactionPartnerEmail": email,
        "transactionPartnerPhone": phone,
        "fuelType": fuel_type,
        "fuelTypeOther": fuel_type_other,
        "fuelCategory": fuel_category,
        "provisionOfTheAct": provision,
        "fuelCode": fuel_code,
        "quantity": quantity,
    }


def _parse_row(
    row: tuple, compliance_report_id: int
) -> AllocationAgreementCreateSchema:
    """
    Parses a valid row into an AllocationAgreementCreateSchema object.
    """
    (
        responsibility,
        transaction_partner,
        postal_address,
        email,
        phone,
        fuel_type,
        fuel_type_other,
        fuel_category,
        provision,
        fuel_code,
        quantity_raw,
    ) = row

    quantity = int(quantity_raw) if quantity_raw is not None else 0

    return AllocationAgreementCreateSchema(
        compliance_report_id=compliance_report_id,
        allocation_transaction_type=responsibility,
        transaction_partner=transaction_partner,
        postal_address=postal_address,
        transaction_partner_email=email,
        transaction_partner_phone=phone,
        fuel_type=fuel_type,
        fuel_type_other=fuel_type_other or None,
        fuel_category=fuel_category,
        provision_of_the_act=provision,
        fuel_code=fuel_code or None,
        quantity=quantity,
    )


async def _update_progress(
    redis_client: Redis,
    job_id: str,
    progress: float,
    status_msg: str,
    created: int = 0,
    rejected: int = 0,
    errors: List[str] = None,
    rejected_rows: List[dict] = None,
    imported_row_ids: List[int] = None,
):
    """
    Persists the job status and progress in Redis.
    """
    if errors is None:
        errors = []
    if rejected_rows is None:
        rejected_rows = []
    if imported_row_ids is None:
        imported_row_ids = []
    data = {
        "progress": progress,
        "status": status_msg,
        "created": created,
        "rejected": rejected,
        "errors": errors,
        "rejected_rows": rejected_rows,
        "imported_row_ids": imported_row_ids,
    }
    await redis_client.set(f"jobs/{job_id}", json.dumps(data), ex=3600)  # Extended TTL for import results
