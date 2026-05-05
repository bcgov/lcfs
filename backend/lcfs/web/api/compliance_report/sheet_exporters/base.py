import decimal
import re
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Any, List, Union

from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from lcfs.web.api.compliance_report.schema import (
    SHOW_COL_STRIPES,
    SHOW_ROW_STRIPES,
    TABLE_STYLE,
    ComplianceReportSummarySchema,
)


class SheetExporter(ABC):
    sheet_name: str
    min_compliance_year: int | None = None

    @abstractmethod
    async def export_to_workbook(
        self,
        wb: Workbook,
        report,
        is_government: bool = True,
        summary: ComplianceReportSummarySchema | None = None,
    ) -> None:
        pass

    def supports(self, compliance_year: int) -> bool:
        return (
            self.min_compliance_year is None
            or compliance_year >= self.min_compliance_year
        )


class SheetExporterSupport:
    def _format_cell(self, cell, value) -> None:
        if isinstance(value, int):
            cell.number_format = "#,##0"
        elif isinstance(value, float) or isinstance(value, decimal.Decimal):
            cell.number_format = "#,##0.00##############"

    def _add_table(self, ws, title: str, cols: int, rows: int) -> None:
        if rows <= 0:
            return
        table_name = re.sub(r"[^A-Za-z0-9_]", "", title.replace(" ", "")) + "Tbl"
        table_ref = f"A1:{get_column_letter(cols)}{rows+1}"

        tab = Table(displayName=table_name, ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE,
            showRowStripes=SHOW_ROW_STRIPES,
            showColumnStripes=SHOW_COL_STRIPES,
        )
        ws.add_table(tab)

    def _auto_size_columns(self, ws) -> None:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            header_cell = col[0]
            header_length = len(str(header_cell.value)) if header_cell.value else 0
            content_width = max(max_length, header_length)
            min_width = 12
            header_text = str(header_cell.value).lower() if header_cell.value else ""

            if any(
                keyword in header_text
                for keyword in ["name", "description", "address", "organization"]
            ):
                min_width = 25
            elif any(
                keyword in header_text
                for keyword in ["fuel type", "fuel code", "manufacturer", "model"]
            ):
                min_width = 18
            elif any(keyword in header_text for keyword in ["email", "phone"]):
                min_width = 20
            elif any(
                keyword in header_text for keyword in ["date", "serial", "registration"]
            ):
                min_width = 15
            elif any(
                keyword in header_text
                for keyword in [
                    "quantity",
                    "units",
                    "compliance",
                    "value",
                    "ci",
                    "density",
                    "eer",
                    "energy",
                ]
            ):
                min_width = 14
            elif any(keyword in header_text for keyword in ["line"]):
                min_width = 8

            padding = max(4, int(content_width * 0.1))
            calculated_width = content_width + padding
            final_width = max(min_width, calculated_width)
            final_width = min(final_width, 50)
            ws.column_dimensions[get_column_letter(col[0].column)].width = final_width

    def _add_sheet(
        self,
        wb: Workbook,
        title: str,
        data: List[List[Union[str, int, float, datetime]]],
        locked_columns: set[int] | None = None,
    ) -> None:
        if not data or len(data) <= 1:
            return

        headers = data[0]
        rows = data[1:]

        has_total_row = False
        if rows and len(rows) >= 2 and len(rows[-1]) > 1 and rows[-1][1] == "Total":
            has_total_row = True
            data_rows = rows[:-2]
            empty_row = rows[-2]
            total_row = rows[-1]
        else:
            data_rows = rows

        if len(data_rows) == 0:
            return

        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)

        for row_idx, row in enumerate(data_rows, start=2):
            ws.append(row)
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                self._format_cell(cell, val)

        self._add_table(ws, title, len(headers), len(data_rows))

        if has_total_row:
            ws.append(empty_row)
            ws.append(total_row)
            total_row_idx = len(data_rows) + 3
            for col_idx, val in enumerate(total_row, start=1):
                cell = ws.cell(row=total_row_idx, column=col_idx)
                cell.font = Font(bold=True)
                self._format_cell(cell, val)

        if locked_columns:
            for ws_row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in ws_row:
                    cell.protection = Protection(locked=(cell.column in locked_columns))
            ws.protection.sheet = True

        self._auto_size_columns(ws)

    def _add_centered_title(self, ws, title: str, columns: int) -> None:
        row_num = ws.max_row + 1
        ws.append([title] + [""] * (columns - 1))
        ws.merge_cells(
            start_row=row_num, start_column=1, end_row=row_num, end_column=columns
        )
        cell = ws.cell(row=row_num, column=1)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)

    def _format_date(self, val) -> Union[datetime, str]:
        if isinstance(val, datetime):
            return val
        if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
            try:
                return val
            except Exception:
                pass
        return str(val) if val else None

    def _extract_display_value(self, value: Any, *attrs: str) -> Any:
        if value is None:
            return None
        if attrs:
            for attr in attrs:
                if hasattr(value, attr):
                    attr_value = getattr(value, attr)
                    if attr_value is not None:
                        return attr_value
        if hasattr(value, "value"):
            enum_value = getattr(value, "value")
            if enum_value is not None:
                return enum_value
        if isinstance(value, (str, int, float, decimal.Decimal, datetime)):
            return value
        return str(value)


class TabularSheetExporter(SheetExporter, SheetExporterSupport, ABC):
    annual_columns = []
    quarterly_columns = []
    locked_columns: set[int] | None = None

    async def export_to_workbook(
        self,
        wb: Workbook,
        report,
        is_government: bool = True,
        summary: ComplianceReportSummarySchema | None = None,
    ) -> None:
        data = await self.load_data(report, is_government=is_government)
        if data:
            self._add_sheet(
                wb,
                self.sheet_name,
                data,
                locked_columns=self.locked_columns,
            )

    def get_columns(self, is_quarterly: bool):
        return self.quarterly_columns if is_quarterly else self.annual_columns

    @abstractmethod
    async def load_data(self, report, is_government: bool = True) -> List[List[Any]]:
        pass
