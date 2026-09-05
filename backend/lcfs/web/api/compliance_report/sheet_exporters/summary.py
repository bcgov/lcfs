from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from lcfs.web.api.compliance_report.schema import (
    LOW_CARBON_SUMMARY_TITLE,
    PENALTY_SUMMARY_TITLE,
    RENEWABLE_REQUIREMENT_TITLE,
    SUMMARY_SHEET,
    TABLE_STYLE,
    ComplianceReportSummarySchema,
)

from .base import SheetExporter, SheetExporterSupport


class SummarySheetExporter(SheetExporter, SheetExporterSupport):
    sheet_name = SUMMARY_SHEET

    async def export_to_workbook(
        self,
        wb: Workbook,
        report,
        is_government: bool = True,
        summary: ComplianceReportSummarySchema | None = None,
    ) -> None:
        if summary is None:
            return

        ws = wb.create_sheet(title=SUMMARY_SHEET)
        bold = Font(bold=True)

        def append_and_bold(row):
            ws.append(row)
            for i, _ in enumerate(row):
                cell = ws.cell(row=ws.max_row, column=i + 1)
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")

        ws.append(["", "", "", "", ""])
        self._add_centered_title(ws, RENEWABLE_REQUIREMENT_TITLE, 5)

        header_row = ws.max_row + 1
        append_and_bold(["Line", "Description", "Gasoline", "Diesel", "Jet"])

        for line in summary.renewable_fuel_target_summary:
            row = [
                line.line,
                line.description,
                line.gasoline,
                line.diesel,
                line.jet_fuel,
            ]
            ws.append(row)
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                if line.line == 11 and col_idx > 2:
                    cell.number_format = '"$"#,##0.00'
                else:
                    self._format_cell(cell, val)

        end_row = ws.max_row
        if end_row > header_row:
            tab = Table(
                displayName="RenewableTbl",
                ref=f"A{header_row}:{get_column_letter(5)}{end_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(
                name=TABLE_STYLE, showRowStripes=False, showColumnStripes=False
            )
            ws.add_table(tab)

        ws.append(["", "", ""])
        self._add_centered_title(ws, LOW_CARBON_SUMMARY_TITLE, 3)

        header_row = ws.max_row + 1
        append_and_bold(["Line", "Description", "Value"])

        for line in summary.low_carbon_fuel_target_summary:
            row = [line.line, line.description, line.value]
            ws.append(row)
            cell = ws.cell(row=ws.max_row, column=len(row))
            cell.number_format = "#,##0"
            if line.line == 21:
                cell.number_format = '"$"#,##0.00'

        end_row = ws.max_row
        if end_row > header_row:
            tab = Table(
                displayName="LowCarbonTbl",
                ref=f"A{header_row}:{get_column_letter(3)}{end_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(
                name=TABLE_STYLE, showRowStripes=False, showColumnStripes=False
            )
            ws.add_table(tab)

        penalty_lines = [
            line
            for line in summary.non_compliance_penalty_summary
            if line.line not in (11, 21) or (line.total_value or 0) > 0
        ]
        has_payable_penalty = any(
            line.line in (11, 21) and (line.total_value or 0) > 0
            for line in penalty_lines
        )
        has_penalty_status = any(
            line.line in (11, 21)
            and (line.invoice_sent is not None or line.payment_received is not None)
            for line in penalty_lines
        )
        include_penalty_status = (
            is_government and has_payable_penalty and has_penalty_status
        )

        penalty_column_count = 5 if include_penalty_status else 3
        ws.append([""] * penalty_column_count)
        self._add_centered_title(ws, PENALTY_SUMMARY_TITLE, penalty_column_count)

        header_row = ws.max_row + 1
        penalty_headers = ["Line", "Description", "Total Value"]
        if include_penalty_status:
            penalty_headers.extend(["Invoice sent", "Payment received"])
        append_and_bold(penalty_headers)

        for line in penalty_lines:
            row = [
                "",
                line.description,
                line.total_value,
            ]
            if include_penalty_status:
                row.extend(
                    [
                        (
                            "Yes"
                            if line.invoice_sent
                            else ("No" if line.invoice_sent is not None else "")
                        ),
                        (
                            "Yes"
                            if line.payment_received
                            else ("No" if line.payment_received is not None else "")
                        ),
                    ]
                )
            ws.append(row)
            cell = ws.cell(row=ws.max_row, column=3)
            cell.number_format = '"$"#,##0.00'

        end_row = ws.max_row
        if end_row > header_row:
            tab = Table(
                displayName="PenaltyTbl",
                ref=f"A{header_row}:{get_column_letter(penalty_column_count)}{end_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(
                name=TABLE_STYLE, showRowStripes=False, showColumnStripes=False
            )
            ws.add_table(tab)

        self._auto_size_columns(ws)
