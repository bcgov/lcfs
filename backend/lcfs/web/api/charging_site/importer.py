import asyncio
import concurrent.futures
import io
import json
import re
import structlog
import uuid
from fastapi import Depends, HTTPException, status, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from typing import List

from lcfs.db.dependencies import db_url, set_user_context
from lcfs.db.models import UserProfile
from lcfs.services.clamav.client import ClamAVService
from lcfs.services.redis.dependency import get_redis_client
from lcfs.settings import settings
from lcfs.utils.constants import (
    POSTAL_REGEX,
    ALLOWED_MIME_TYPES,
    ALLOWED_FILE_TYPES,
    MAX_FILE_SIZE_BYTES,
    MAX_FILE_SIZE_MB,
)
from lcfs.web.api.charging_site.repo import ChargingSiteRepository
from lcfs.web.api.charging_site.schema import ChargingSiteCreateSchema
from lcfs.web.api.charging_site.services import ChargingSiteService
from lcfs.web.core.decorators import service_handler
from lcfs.web.exception.exceptions import DataNotFoundException
from lcfs.web.api.organizations.repo import OrganizationsRepository

logger = structlog.get_logger(__name__)


class ChargingSiteImporter:
    def __init__(
        self,
        repo: ChargingSiteRepository = Depends(),
        cs_service: ChargingSiteService = Depends(),
        clamav_service: ClamAVService = Depends(),
        redis_client: Redis = Depends(get_redis_client),
        executor: concurrent.futures.ThreadPoolExecutor = Depends(),
    ) -> None:
        self.repo = repo
        self.cs_service = cs_service
        self.clamav_service = clamav_service
        self.redis_client = redis_client
        self.executor = executor

    @service_handler
    async def import_data(
        self,
        organization_id: int,
        user: UserProfile,
        org_code: str,
        file: UploadFile,
        overwrite: bool,
        site_ids: List[int] = None,
    ) -> str:
        """
        Initiates the import job in a separate thread executor.
        Returns a job_id that can be used to track progress via get_status.
        """
        job_id = str(uuid.uuid4())

        # Initialize job status in Redis
        await _update_progress(
            self.redis_client, job_id, 0, "Starting import job...", 0, 0, []
        )

        # Read file into memory once to enable scanning and openpyxl parsing
        file_contents = await file.read()

        # Validate MIME type
        if file.content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"File type '{file.content_type or 'unknown'}' is not allowed. Please upload files of the following types: {ALLOWED_FILE_TYPES}",
            )

        # Validate file size
        file_size = len(file_contents)
        if file_size > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"File size exceeds the maximum limit of {MAX_FILE_SIZE_MB} MB.",
            )

        # Create an in-memory buffer and attach it to a new UploadFile
        buffer = io.BytesIO(file_contents)
        buffer.seek(0)
        copied_file = UploadFile(filename=file.filename, file=buffer)

        # Start the import task without blocking
        asyncio.create_task(
            import_async(
                organization_id,
                user,
                org_code,
                copied_file,
                job_id,
                overwrite,
                site_ids,
            )
        )

        return job_id

    async def get_status(self, job_id: str) -> dict:
        """
        Retrieves and returns the job's progress and status from Redis.
        """
        logger.debug(f"Getting status for job_id: {job_id}")
        progress_data_str = await self.redis_client.get(f"jobs/{job_id}")
        logger.debug(f"Redis data for job {job_id}: {progress_data_str}")

        if not progress_data_str:
            logger.warning(f"No job found with ID: {job_id}")
            return {"progress": 0, "status": "No job found with this ID."}

        try:
            progress_data = json.loads(progress_data_str)
            logger.debug(f"Parsed progress data: {progress_data}")
            return {
                "progress": progress_data.get("progress", 0),
                "status": progress_data.get("status", "No status available."),
                "created": progress_data.get("created", 0),
                "rejected": progress_data.get("rejected", 0),
                "errors": progress_data.get("errors", []),
            }
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error for job {job_id}: {e}")
            return {"progress": 0, "status": "Invalid status data found."}


async def import_async(
    organization_id: int,
    user: UserProfile,
    org_code: str,
    file: UploadFile,
    job_id: str,
    overwrite: bool,
    site_ids: List[int] = None,
):
    """
    Performs the actual import in an async context.
    """
    logger.debug(f"Starting import_async for job_id: {job_id}")

    engine = create_async_engine(db_url, future=True)
    try:
        async with AsyncSession(engine) as session:
            async with session.begin():
                await set_user_context(session, user.keycloak_username)

                cs_repo = ChargingSiteRepository(session)
                org_repo = OrganizationsRepository(session)
                cs_service = ChargingSiteService(repo=cs_repo)
                clamav_service = ClamAVService()
                # Use a new Redis client for this async context
                redis_client = Redis(
                    host=settings.redis_host,
                    port=settings.redis_port,
                    password=settings.redis_pass,
                    db=settings.redis_base or 0,
                    decode_responses=True,
                )

                logger.debug(f"About to update progress for job {job_id}")
                await _update_progress(
                    redis_client, job_id, 5, "Initializing services..."
                )
                logger.debug(f"Progress updated for job {job_id}")

                if overwrite:
                    await _update_progress(
                        redis_client, job_id, 10, "Deleting old data..."
                    )
                    if site_ids:
                        await cs_service.delete_charging_sites_by_ids(site_ids)

                # Optional: Scan the file with ClamAV if enabled
                if settings.clamav_enabled:
                    await _update_progress(
                        redis_client, job_id, 15, "Scanning file with ClamAV..."
                    )
                    clamav_service.scan_file(file)

                await _update_progress(
                    redis_client, job_id, 20, "Loading Excel sheet..."
                )

                try:
                    sheet = _load_sheet(file)
                    row_count = sheet.max_row

                    created = 0
                    rejected = 0
                    errors = []

                    await _update_progress(
                        redis_client,
                        job_id,
                        20,
                        "Beginning data import...",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                    )

                    valid_intended_user_types = await cs_repo.get_intended_user_types()
                    valid_user_type_names = {
                        obj.type_name for obj in valid_intended_user_types
                    }

                    # Iterate through all data rows, skipping the header
                    for row_idx, row in enumerate(
                        sheet.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        # Only update progress in Redis every 10 rows for efficiency
                        if (row_idx + 1) % 10 == 0:
                            current_progress = 20 + ((row_idx / row_count) * 80)
                            status_msg = (
                                f"Importing row {row_idx - 1} of {row_count - 1}..."
                            )
                            await _update_progress(
                                redis_client,
                                job_id,
                                current_progress,
                                status_msg,
                                created=created,
                                rejected=rejected,
                                errors=errors,
                            )

                        # Check if the entire row is empty
                        if all(cell is None for cell in row):
                            continue

                        row = list(row)
                        row[8] = [row[8]] if row[8] is not None else []

                        # Validate row
                        error = _validate_row(row, row_idx, valid_user_type_names)
                        if error:
                            errors.append(error)
                            rejected += 1
                            continue

                        # Parse row data and insert into DB
                        try:
                            cs_data = _parse_row(row, organization_id)
                            await cs_service.create_charging_site(
                                cs_data, organization_id
                            )
                            created += 1
                        except Exception as ex:
                            logger.error(str(ex))
                            errors.append(f"Row {row_idx}: {ex}")
                            rejected += 1

                    # Final update at 100%
                    await _update_progress(
                        redis_client,
                        job_id,
                        100,
                        "Import process completed.",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                    )
                    logger.debug(
                        f"Completed importing charging site data, {created} rows created"
                    )

                    return {
                        "success": True,
                        "created": created,
                        "errors": errors,
                        "rejected": rejected,
                    }
                except DataNotFoundException as dnfe:
                    await _update_progress(
                        redis_client, job_id, 100, "Data not found error.", 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND, detail=str(dnfe)
                    )
                except Exception as e:
                    await _update_progress(
                        redis_client, job_id, 100, "Import process failed.", 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Could not import charging site data: {str(e)}",
                    )
                finally:
                    await redis_client.aclose()
    finally:
        await engine.dispose()


def _load_sheet(file: UploadFile) -> Worksheet:
    """
    Loads and returns the 'ChargingSites' worksheet from the provided Excel file.
    Raises an exception if the sheet does not exist.
    """
    workbook = load_workbook(data_only=True, filename=file.file)
    if "ChargingSites" not in workbook.sheetnames:
        raise Exception("Uploaded Excel does not contain a 'ChargingSites' sheet.")
    return workbook["ChargingSites"]


def _validate_row(
    row: tuple,
    row_idx: int,
    valid_user_types: set[str],
) -> str | None:
    """
    Validates a single row of data and returns an error string if invalid.
    Returns None if the row is valid.
    """
    (
        org_name,
        site_code,
        site_name,
        street_address,
        city,
        postal_code,
        latitude,
        longitude,
        intended_user_types,
        status,
        notes,
    ) = row

    missing_fields = []
    if site_name is None:
        missing_fields.append("Site Name")
    if street_address is None:
        missing_fields.append("Street Address")
    if city is None:
        missing_fields.append("City")
    if postal_code is None:
        missing_fields.append("Postal Code")
    if latitude is None:
        missing_fields.append("Latitude")
    if longitude is None:
        missing_fields.append("Longitude")
    if not intended_user_types:
        missing_fields.append("Intended Users")

    if missing_fields:
        return f"Row {row_idx}: Missing required fields: {', '.join(missing_fields)}"

    # Validate postal code
    postal_code_pattern = re.compile(POSTAL_REGEX)
    if not postal_code_pattern.match(postal_code):
        return f"Row {row_idx}: Invalid postal code"

    # Validate intended users
    invalid_users = [
        user for user in intended_user_types if user not in valid_user_types
    ]
    if invalid_users:
        return f"Row {row_idx}: Invalid intended user(s): {', '.join(invalid_users)}"

    return None


def _parse_row(row: tuple, organization_id: int) -> ChargingSiteCreateSchema:
    """
    Parses a valid row into a ChargingSiteCreateSchema object.
    """
    (
        org_name,
        site_code,
        site_name,
        street_address,
        city,
        postal_code,
        latitude,
        longitude,
        intended_user_types,
        status,
        notes,
    ) = row

    latitude = float(latitude) if latitude else 0.0
    longitude = float(longitude) if longitude else 0.0

    return ChargingSiteCreateSchema(
        organization_id=organization_id,
        site_code=str(site_code) or "",
        site_name=str(site_name) or "",
        street_address=str(street_address) or "",
        city=str(city) or "",
        postal_code=postal_code or "",
        latitude=latitude,
        longitude=longitude,
        intended_users=intended_user_types,
        current_status=status or "Draft",
        notes=notes or "",
    )


async def _update_progress(
    redis_client: Redis,
    job_id: str,
    progress: float,
    status_msg: str,
    created: int = 0,
    rejected: int = 0,
    errors: List[str] = None,
):
    """
    Persists the job status and progress in Redis.
    """
    if errors is None:
        errors = []
    data = {
        "progress": progress,
        "status": status_msg,
        "created": created,
        "rejected": rejected,
        "errors": errors,
    }
    logger.debug(f"Updating progress for job {job_id}: {data}")
    try:
        result = await redis_client.set(
            f"jobs/{job_id}", json.dumps(data), ex=300
        )  # 5 minutes
        logger.debug(f"Redis set result for job {job_id}: {result}")
    except Exception as e:
        logger.error(f"Failed to update progress for job {job_id}: {e}")
