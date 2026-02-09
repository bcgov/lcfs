import asyncio
import concurrent.futures
import io
import json
import structlog
import uuid
from typing import Iterable, List

from fastapi import Depends, HTTPException, status, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine

from lcfs.db.dependencies import db_url, set_user_context
from lcfs.db.models import UserProfile
from lcfs.services.clamav.client import ClamAVService
from lcfs.services.redis.dependency import get_redis_client
from lcfs.settings import settings
from lcfs.utils.constants import (
    ALLOWED_MIME_TYPES,
    ALLOWED_FILE_TYPES,
    MAX_FILE_SIZE_BYTES,
    MAX_FILE_SIZE_MB,
)
from lcfs.web.api.charging_equipment.repo import ChargingEquipmentRepository
from lcfs.web.api.charging_equipment.schema import ChargingEquipmentCreateSchema
from lcfs.web.api.charging_equipment.services import ChargingEquipmentServices
from lcfs.web.core.decorators import service_handler


logger = structlog.get_logger(__name__)


def _parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        stripped = str(value).strip()
        if stripped == "":
            return None
        return float(stripped)
    except (ValueError, TypeError):
        return None


class ChargingEquipmentImporter:
    def __init__(
        self,
        repo: ChargingEquipmentRepository = Depends(),
        ce_service: ChargingEquipmentServices = Depends(),
        clamav_service: ClamAVService = Depends(),
        redis_client: Redis = Depends(get_redis_client),
        executor: concurrent.futures.ThreadPoolExecutor = Depends(),
    ) -> None:
        self.repo = repo
        self.ce_service = ce_service
        self.clamav_service = clamav_service
        self.redis_client = redis_client
        self.executor = executor

    @service_handler
    async def import_data(
        self,
        organization_id: int,
        user: UserProfile,
        org_code: str,
        file: UploadFile,
        overwrite: bool,
    ) -> str:
        """
        Initiates the import job in a separate thread executor.
        Returns a job_id that can be used to track progress via get_status.
        """
        job_id = str(uuid.uuid4())

        # Initialize job status in Redis
        await _update_progress(
            self.redis_client, job_id, 0, "Starting import job...", 0, 0, []
        )

        # Read file into memory once to enable scanning and openpyxl parsing
        file_contents = await file.read()

        # Validate MIME type
        if file.content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"File type '{file.content_type or 'unknown'}' is not allowed. Please upload files of the following types: {ALLOWED_FILE_TYPES}",
            )

        # Validate file size
        file_size = len(file_contents)
        if file_size > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"File size exceeds the maximum limit of {MAX_FILE_SIZE_MB} MB.",
            )

        # Create an in-memory buffer and attach it to a new UploadFile
        buffer = io.BytesIO(file_contents)
        buffer.seek(0)
        copied_file = UploadFile(filename=file.filename, file=buffer)

        loop = asyncio.get_running_loop()
        loop.run_in_executor(
            self.executor,
            lambda: asyncio.run(
                import_async(
                    organization_id,
                    user,
                    org_code,
                    copied_file,
                    job_id,
                    overwrite,
                )
            ),
        )

        return job_id

    async def get_status(self, job_id: str) -> dict:
        """
        Retrieves and returns the job's progress and status from Redis.
        """
        progress_data_str = await self.redis_client.get(f"jobs/{job_id}")
        if not progress_data_str:
            return {"progress": 0, "status": "No job found with this ID."}

        try:
            progress_data = json.loads(progress_data_str)
            return {
                "progress": progress_data.get("progress", 0),
                "status": progress_data.get("status", "No status available."),
                "created": progress_data.get("created", 0),
                "rejected": progress_data.get("rejected", 0),
                "errors": progress_data.get("errors", []),
            }
        except json.JSONDecodeError:
            return {"progress": 0, "status": "Invalid status data found."}


async def import_async(
    organization_id: int,
    user: UserProfile,
    org_code: str,
    file: UploadFile,
    job_id: str,
    overwrite: bool,
):
    """
    Performs the actual import in an async context.
    This is offloaded to a thread pool to avoid blocking.
    """
    logger.debug("Importing Charging Equipment Data...")

    engine = create_async_engine(db_url, future=True)
    try:
        async with AsyncSession(engine) as session:
            async with session.begin():
                await set_user_context(session, user.keycloak_username)

                ce_repo = ChargingEquipmentRepository(session)
                ce_service = ChargingEquipmentServices(repo=ce_repo)
                clamav_service = ClamAVService()
                redis_client = Redis(
                    host=settings.redis_host,
                    port=settings.redis_port,
                    password=settings.redis_pass,
                    db=settings.redis_base or 0,
                    decode_responses=True,
                    max_connections=10,
                    socket_timeout=5,
                    socket_connect_timeout=5,
                )

                await _update_progress(
                    redis_client, job_id, 5, "Initializing services..."
                )

                if overwrite:
                    await _update_progress(
                        redis_client, job_id, 10, "Deleting old data..."
                    )
                    await ce_service.delete_all_for_organization(organization_id)

                # Optional: Scan the file with ClamAV if enabled
                if settings.clamav_enabled:
                    await _update_progress(
                        redis_client, job_id, 15, "Scanning file with ClamAV..."
                    )
                    clamav_service.scan_file(file)

                await _update_progress(
                    redis_client, job_id, 20, "Loading Excel sheet..."
                )

                try:
                    sheet = _load_sheet(file)
                    row_count = sheet.max_row

                    created = 0
                    rejected = 0
                    errors: List[str] = []
                    success_rows: List[dict] = []

                    await _update_progress(
                        redis_client,
                        job_id,
                        20,
                        "Beginning data import...",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                        successes=success_rows,
                    )

                    # Preload lookups
                    charging_sites = await ce_repo.get_charging_sites_by_organization(
                        organization_id
                    )
                    site_lookup_by_name = {
                        (s.site_name or "").strip().lower(): {
                            "id": s.charging_site_id,
                            "latitude": s.latitude,
                            "longitude": s.longitude,
                        }
                        for s in charging_sites
                        if s.site_name
                    }
                    site_lookup_by_code = {
                        (s.site_code or "").strip().upper(): {
                            "id": s.charging_site_id,
                            "latitude": s.latitude,
                            "longitude": s.longitude,
                        }
                        for s in charging_sites
                        if s.site_code
                    }
                    levels = await ce_repo.get_levels_of_equipment()
                    level_name_to_id = {l.name: l.level_of_equipment_id for l in levels}
                    end_use_types = await ce_repo.get_end_use_types()
                    end_use_name_to_id = {
                        e.type: e.end_use_type_id for e in end_use_types
                    }
                    end_user_types = await ce_repo.get_end_user_types()
                    end_user_name_to_id = {
                        u.type_name: u.end_user_type_id for u in end_user_types
                    }
                    existing_serials = await ce_repo.get_serial_numbers_for_organization(
                        organization_id
                    )
                    duplicate_tracker = _DuplicateSerialTracker(existing_serials)

                    # Iterate through all data rows, skipping the header
                    for row_idx, row in enumerate(
                        sheet.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        # Only update progress in Redis every 10 rows for efficiency
                        if (row_idx + 1) % 10 == 0:
                            current_progress = 20 + ((row_idx / row_count) * 80)
                            status_msg = (
                                f"Importing row {row_idx - 1} of {row_count - 1}..."
                            )
                            await _update_progress(
                                redis_client,
                                job_id,
                                current_progress,
                                status_msg,
                                created=created,
                                rejected=rejected,
                                errors=errors,
                                successes=success_rows,
                            )

                        # Check if the entire row is empty
                        if all(cell is None for cell in row):
                            continue

                        row_values = list(row)
                        expanded = row_values + [None] * max(0, 11 - len(row_values))

                        site_name = expanded[0]
                        serial_number = expanded[1]
                        manufacturer = expanded[2]
                        model = expanded[3]
                        level_name = expanded[4]
                        ports = expanded[5]
                        intended_uses_str = expanded[6]
                        intended_users_str = expanded[7]
                        notes = expanded[8]
                        latitude_value = expanded[9]
                        longitude_value = expanded[10]

                        # Validate required fields
                        missing_fields = []
                        if not site_name:
                            missing_fields.append("Charging Site")
                        if not serial_number:
                            missing_fields.append("Serial Number")
                        if not manufacturer:
                            missing_fields.append("Manufacturer")
                        if not level_name:
                            missing_fields.append("Level of Equipment")
                        if missing_fields:
                            errors.append(
                                f"Row {row_idx}: Missing required fields: {', '.join(missing_fields)}"
                            )
                            rejected += 1
                            continue

                        # Lookups
                        site_info = None
                        if site_name:
                            normalized_site = str(site_name).strip()
                            site_info = site_lookup_by_name.get(
                                normalized_site.lower()
                            )
                            if not site_info:
                                site_info = site_lookup_by_code.get(
                                    normalized_site.upper()
                                )
                        if not site_info:
                            reference = site_name or ""
                            errors.append(
                                f"Row {row_idx}: Charging Site '{reference}' not found for your organization"
                            )
                            rejected += 1
                            continue
                        charging_site_id = site_info["id"]
                        latitude = site_info.get("latitude")
                        longitude = site_info.get("longitude")

                        user_latitude = _parse_float(latitude_value)
                        user_longitude = _parse_float(longitude_value)
                        if user_latitude is not None:
                            latitude = user_latitude
                        if user_longitude is not None:
                            longitude = user_longitude

                        level_id = level_name_to_id.get(str(level_name))
                        if not level_id:
                            errors.append(
                                f"Row {row_idx}: Level of Equipment '{level_name}' not found"
                            )
                            rejected += 1
                            continue

                        intended_use_ids: List[int] = []
                        if intended_uses_str:
                            for name in str(intended_uses_str).split(","):
                                clean = name.strip()
                                if not clean:
                                    continue
                                end_use_id = end_use_name_to_id.get(clean)
                                if end_use_id:
                                    intended_use_ids.append(end_use_id)
                                else:
                                    errors.append(
                                        f"Row {row_idx}: Intended Use '{clean}' not found; skipping this value"
                                    )

                        intended_user_ids: List[int] = []
                        if intended_users_str:
                            for name in str(intended_users_str).split(","):
                                clean = name.strip()
                                if not clean:
                                    continue
                                end_user_id = end_user_name_to_id.get(clean)
                                if end_user_id:
                                    intended_user_ids.append(end_user_id)
                                else:
                                    errors.append(
                                        f"Row {row_idx}: Intended User '{clean}' not found; skipping this value"
                                    )

                        if duplicate_tracker.is_duplicate(serial_number):
                            rejected += 1
                            continue

                        try:
                            ce_data = ChargingEquipmentCreateSchema(
                                charging_site_id=charging_site_id,
                                serial_number=str(serial_number),
                                manufacturer=str(manufacturer),
                                model=str(model) if model else None,
                                level_of_equipment_id=level_id,
                                ports=str(ports) if ports else None,
                                latitude=latitude,
                                longitude=longitude,
                                notes=str(notes) if notes else None,
                                intended_use_ids=intended_use_ids,
                                intended_user_ids=intended_user_ids,
                            )
                            created_equipment = await ce_service.create_charging_equipment(
                                user, ce_data
                            )
                            success_rows.append(
                                {
                                    "row": row_idx,
                                    "chargingEquipmentId": created_equipment.charging_equipment_id,
                                }
                            )
                            created += 1
                        except Exception as ex:
                            logger.error(str(ex))
                            errors.append(f"Row {row_idx}: {ex}")
                            rejected += 1

                    duplicate_summary = duplicate_tracker.summary_message()
                    if duplicate_summary:
                        errors.append(duplicate_summary)

                    # Final update at 100%
                    await _update_progress(
                        redis_client,
                        job_id,
                        100,
                        "Import process completed.",
                        created=created,
                        rejected=rejected,
                        errors=errors,
                        successes=success_rows,
                    )
                    logger.debug(
                        f"Completed importing charging equipment data, {created} rows created"
                    )

                    return {
                        "success": True,
                        "created": created,
                        "errors": errors,
                        "rejected": rejected,
                    }

                except Exception as e:
                    await _update_progress(
                        redis_client, job_id, 100, "Import process failed.", 0, 0
                    )
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Could not import charging equipment data: {str(e)}",
                    )
    finally:
        await engine.dispose()


def _load_sheet(file: UploadFile) -> Worksheet:
    """
    Loads and returns the first worksheet from the provided Excel file.
    """
    workbook = load_workbook(data_only=True, filename=file.file)
    # Use first sheet if a specific name is not enforced
    sheet_name = workbook.sheetnames[0]
    return workbook[sheet_name]


async def _update_progress(
    redis_client: Redis,
    job_id: str,
    progress: float,
    status_msg: str,
    created: int = 0,
    rejected: int = 0,
    errors: List[str] | None = None,
    successes: List[dict] | None = None,
):
    """
    Persists the job status and progress in Redis.
    """
    if errors is None:
        errors = []
    data = {
        "progress": progress,
        "status": status_msg,
        "created": created,
        "rejected": rejected,
        "errors": errors,
        "successes": successes or [],
    }
    await redis_client.set(f"jobs/{job_id}", json.dumps(data), ex=60)


class _DuplicateSerialTracker:
    """
    Tracks duplicate serial numbers within a single upload while
    considering existing records for an organization.
    """

    def __init__(self, existing_serials: Iterable[str] | None = None) -> None:
        normalized_existing: set[str] = set()
        for serial in existing_serials or []:
            normalized = _normalize_serial(serial)
            if normalized:
                normalized_existing.add(normalized)
        self._existing_serials = normalized_existing
        self._current_upload_serials: set[str] = set()
        self._duplicate_count = 0

    def is_duplicate(self, serial_number) -> bool:
        normalized = _normalize_serial(serial_number)
        if not normalized:
            return False
        if (
            normalized in self._existing_serials
            or normalized in self._current_upload_serials
        ):
            self._duplicate_count += 1
            return True
        self._current_upload_serials.add(normalized)
        return False

    def summary_message(self) -> str | None:
        if not self._duplicate_count:
            return None
        record_label = "record" if self._duplicate_count == 1 else "records"
        verb = "was" if self._duplicate_count == 1 else "were"
        return (
            f"{self._duplicate_count} {record_label} "
            f"with duplicate serial numbers {verb} not uploaded."
        )


def _normalize_serial(serial_number) -> str:
    """
    Normalizes serial numbers for duplicate comparisons.
    """
    if serial_number is None:
        return ""
    return str(serial_number).strip().upper()
