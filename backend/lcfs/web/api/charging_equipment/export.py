import io
from typing import List, Tuple

from fastapi import Depends
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from starlette.responses import StreamingResponse

from lcfs.db.models import Organization, UserProfile
from lcfs.utils.constants import FILE_MEDIA_TYPE
from lcfs.utils.spreadsheet_builder import SpreadsheetBuilder, SpreadsheetColumn
from lcfs.web.api.charging_equipment.repo import ChargingEquipmentRepository
from lcfs.web.core.decorators import service_handler


CE_EXPORT_FILENAME = "ChargingEquipment"
CE_EXPORT_SHEETNAME = "ChargingEquipment"
VALIDATION_SHEETNAME = "VALUES"
CE_EXPORT_COLUMNS = [
    SpreadsheetColumn("Charging Site", "text"),
    SpreadsheetColumn("Serial Number", "text"),
    SpreadsheetColumn("Manufacturer", "text"),
    SpreadsheetColumn("Model", "text"),
    SpreadsheetColumn("Level of Equipment", "text"),
    SpreadsheetColumn("Ports", "text"),
    SpreadsheetColumn("Intended Uses", "text"),
    SpreadsheetColumn("Intended Users", "text"),
    SpreadsheetColumn("Notes", "text"),
    SpreadsheetColumn("Latitude", "text"),
    SpreadsheetColumn("Longitude", "text"),
]


class ChargingEquipmentExporter:
    def __init__(
        self,
        repo: ChargingEquipmentRepository = Depends(ChargingEquipmentRepository),
    ) -> None:
        self.repo = repo

    @service_handler
    async def export(
        self,
        organization_id: int,
        user: UserProfile,
        organization: Organization,
        include_data: bool = True,
    ) -> StreamingResponse:
        """
        Prepares a list of charging equipment in a downloadable spreadsheet.
        """
        export_format = "xlsx"

        builder = SpreadsheetBuilder(file_format=export_format)

        validators, charging_site_count = await self._create_validators(
            organization, builder
        )

        data = []
        if include_data:
            data = await self.load_charging_equipment_data(organization_id)
        column_formulas = None
        formula_end_row = None
        format_max_row = None
        if not include_data and charging_site_count:
            column_formulas = self._build_charging_site_formulas(charging_site_count)
            formula_end_row = 10000
            format_max_row = 10000

        builder.add_sheet(
            sheet_name=CE_EXPORT_SHEETNAME,
            columns=CE_EXPORT_COLUMNS,
            rows=data,
            styles={"bold_headers": True},
            validators=validators,
            column_formulas=column_formulas,
            formula_end_row=formula_end_row,
            format_max_row=format_max_row,
        )
        file_content = builder.build_spreadsheet()

        filename = f"{CE_EXPORT_FILENAME}_{organization.name}.{export_format}"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

        return StreamingResponse(
            io.BytesIO(file_content),
            media_type=FILE_MEDIA_TYPE[export_format.upper()].value,
            headers=headers,
        )

    async def _create_validators(
        self, organization: Organization, builder
    ) -> Tuple[List[DataValidation], int]:
        validators: List[DataValidation] = []

        # Options
        charging_sites = await self.repo.get_charging_sites_by_organization(
            organization.organization_id
        )
        levels = await self.repo.get_levels_of_equipment()
        end_use_types = await self.repo.get_end_use_types()
        end_user_types = await self.repo.get_end_user_types()

        ports_options = ["Single port", "Dual port"]

        site_names = [s.site_name or "" for s in charging_sites]
        level_names = [level.name for level in levels]
        end_use_names = [e.type for e in end_use_types]
        end_user_names = [u.type_name for u in end_user_types]
        site_latitudes = [s.latitude for s in charging_sites]
        site_longitudes = [s.longitude for s in charging_sites]

        # Charging Site validator (Column A)
        charging_site_end_row = len(site_names) + 1
        site_validator = DataValidation(
            type="list",
            formula1=f"=VALUES!$A$2:$A${charging_site_end_row}",
            showErrorMessage=True,
            error="Please select a valid charging site",
        )
        site_validator.add("A2:A10000")
        validators.append(site_validator)

        # Level of Equipment validator (Column E)
        level_validator = DataValidation(
            type="list",
            formula1="=VALUES!$B$2:$B$1000",
            showErrorMessage=True,
            error="Please select a valid level of equipment",
        )
        level_validator.add("E2:E10000")
        validators.append(level_validator)

        # Ports validator (Column F)
        ports_validator = DataValidation(
            type="list",
            formula1="=VALUES!$C$2:$C$1000",
            showErrorMessage=True,
            error="Please select either 'Single port' or 'Dual port'",
        )
        ports_validator.add("F2:F10000")
        validators.append(ports_validator)

        # Intended Uses validator (Column G) - Allow comma-separated values
        intended_uses_validator = DataValidation(
            type="list",
            formula1="=VALUES!$D$2:$D$1000",
            showErrorMessage=False,  # Allow comma-separated
            showInputMessage=True,
            promptTitle="Intended Uses",
            prompt="Select from dropdown or type comma-separated values",
            allow_blank=False,
        )
        intended_uses_validator.add("G2:G10000")
        validators.append(intended_uses_validator)

        # Intended Users validator (Column H) - Allow comma-separated values
        intended_users_validator = DataValidation(
            type="list",
            formula1="=VALUES!$E$2:$E$1000",
            showErrorMessage=False,  # Allow comma-separated
            showInputMessage=True,
            promptTitle="Intended Users",
            prompt="Select from dropdown or type comma-separated values (e.g., Public, Fleet, Multi-unit residential building)",
            allow_blank=False,
        )
        intended_users_validator.add("H2:H10000")
        validators.append(intended_users_validator)

        # Build VALUES sheet data (A: Charging Sites, B: Levels, C: Ports, D: Intended Uses, E: Intended Users, F: Latitude, G: Longitude)
        data = [
            site_names,
            level_names,
            ports_options,
            end_use_names,
            end_user_names,
            site_latitudes,
            site_longitudes,
        ]

        # Determine the maximum length among all columns
        max_length = max((len(options) for options in data), default=0)

        # Build rows ensuring each row has an entry for each column
        rows = [
            [col[i] if i < len(col) else None for col in data]
            for i in range(max_length)
        ]

        builder.add_sheet(
            sheet_name=VALIDATION_SHEETNAME,
            columns=[
                SpreadsheetColumn("Charging Sites", "text"),
                SpreadsheetColumn("Levels", "text"),
                SpreadsheetColumn("Ports", "text"),
                SpreadsheetColumn("Intended Uses", "text"),
                SpreadsheetColumn("Intended Users", "text"),
                SpreadsheetColumn("Latitude", "text"),
                SpreadsheetColumn("Longitude", "text"),
            ],
            rows=rows,
            styles={"bold_headers": True},
            validators=[],
            position=1,
        )

        return validators, len(site_names)

    def _build_charging_site_formulas(self, charging_site_count: int) -> dict[int, str]:
        charging_site_column = self._get_column_letter("Charging Site")
        latitude_column_index = self._get_column_index("Latitude")
        longitude_column_index = self._get_column_index("Longitude")
        lookup_end = charging_site_count + 1
        lookup_range = f"VALUES!$A$2:$G${lookup_end}"

        latitude_formula = (
            f'=IF(${charging_site_column}{{row}}="","",'
            f"IFERROR(IF(VLOOKUP(${charging_site_column}{{row}},{lookup_range},6,FALSE)=0,"
            f'"",VLOOKUP(${charging_site_column}{{row}},{lookup_range},6,FALSE)),""))'
        )
        longitude_formula = (
            f'=IF(${charging_site_column}{{row}}="","",'
            f"IFERROR(IF(VLOOKUP(${charging_site_column}{{row}},{lookup_range},7,FALSE)=0,"
            f'"",VLOOKUP(${charging_site_column}{{row}},{lookup_range},7,FALSE)),""))'
        )

        return {
            latitude_column_index: latitude_formula,
            longitude_column_index: longitude_formula,
        }

    def _get_column_index(self, label: str) -> int:
        for idx, column in enumerate(CE_EXPORT_COLUMNS, start=1):
            if column.label == label:
                return idx
        raise ValueError(f"Missing column label: {label}")

    def _get_column_letter(self, label: str) -> str:
        return get_column_letter(self._get_column_index(label))

    async def load_charging_equipment_data(self, organization_id: int):
        results = await self.repo.get_all_equipment_by_organization_id(organization_id)
        data = []
        for eq in results:
            data.append(
                [
                    eq.charging_site.site_name if eq.charging_site else "",
                    eq.serial_number or "",
                    eq.manufacturer or "",
                    eq.model or "",
                    eq.level_of_equipment.name if eq.level_of_equipment else "",
                    (eq.ports.value if getattr(eq, "ports", None) else ""),
                    ", ".join(use.type for use in getattr(eq, "intended_uses", [])),
                    ", ".join(
                        user.type_name for user in getattr(eq, "intended_users", [])
                    ),
                    eq.notes or "",
                    eq.latitude if getattr(eq, "latitude", None) is not None else "",
                    eq.longitude if getattr(eq, "longitude", None) is not None else "",
                ]
            )
        return data
