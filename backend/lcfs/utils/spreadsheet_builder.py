import pandas as pd
import structlog
import xlwt
from io import BytesIO
from openpyxl import styles
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from typing import Literal, TypedDict, List, Dict, Any, Optional

logger = structlog.get_logger(__name__)
MAX_XLS_WIDTH = 65535


class SpreadsheetColumn:
    def __init__(
        self,
        label: str,
        column_type: Literal["float", "int", "text", "date", "decimal6"],
    ):
        self.label = label
        self.column_type = column_type


class RawSpreadsheet:
    def __init__(self, label: str, data: List[List[Any]], styles: List[List[Any]]):
        self.label = label
        self.data = data
        self.styles = styles


class SheetData(TypedDict):
    sheet_name: str
    columns: List[SpreadsheetColumn]
    rows: List[Any]
    styles: Dict[str, Any]
    validators: List[DataValidation]
    column_formulas: Dict[int, str]
    formula_start_row: int
    formula_end_row: int
    format_max_row: int


class SpreadsheetBuilder:
    """
    A class to build spreadsheets in xlsx, xls, or csv format.
    Allows adding multiple sheets with custom styling and exporting them as a byte stream.
    """

    def __init__(self, file_format: str = "xls"):
        if file_format not in ["xlsx", "xls", "csv"]:
            raise ValueError(f"Unsupported file format: {file_format}")

        self.file_format = file_format
        self.sheets_data: List[SheetData] = []
        self.raw_sheets: List[RawSpreadsheet] = []

    def add_sheet(
        self,
        sheet_name: str,
        columns: List[SpreadsheetColumn],
        rows: list,
        styles: Optional[Dict] = None,
        validators: Optional[List[DataValidation]] = None,
        position: int = 0,
        column_formulas: Optional[Dict[int, str]] = None,
        formula_start_row: int = 2,
        formula_end_row: Optional[int] = None,
        format_max_row: Optional[int] = None,
    ):
        # Avoid mutable default arguments by setting validators to an empty list if None.
        validators = validators or []
        self.sheets_data.insert(
            position,
            {
                "sheet_name": sheet_name,
                "columns": columns,
                "rows": rows,
                "styles": styles or {},
                "validators": validators,
                "column_formulas": column_formulas or {},
                "formula_start_row": formula_start_row,
                "formula_end_row": formula_end_row or 0,
                "format_max_row": format_max_row or 0,
            },
        )
        return self

    def add_raw_sheet(self, data: RawSpreadsheet):
        self.raw_sheets.append(data)

    def build_spreadsheet(self) -> bytes:
        try:
            output = BytesIO()
            # Map file formats to the corresponding writer methods
            writer_func = {
                "xlsx": self._write_xlsx,
                "xls": self._write_xls,
                "csv": self._write_csv,
            }[self.file_format]
            writer_func(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error("Failed to build spreadsheet", error=str(e), exc_info=e)
            raise

    def _set_cell_format(self, cell, cell_type: Optional[str]) -> None:
        """
        Helper method to set formatting on a cell based on cell_type.
        """
        # Default alignment is left.
        cell.alignment = styles.Alignment(horizontal="left")
        if cell_type == "int":
            cell.number_format = "#,##0"
            cell.alignment = styles.Alignment(horizontal="right")
        elif cell_type == "float":
            cell.number_format = "#,##0.00"
            cell.alignment = styles.Alignment(horizontal="right")
        elif cell_type == "date":
            cell.number_format = "yyyy-mm-dd"
        elif cell_type == "decimal6":
            cell.number_format = "#,##0.00####"
            cell.alignment = styles.Alignment(horizontal="right")

    def _write_xlsx(self, output):
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Process standard sheets.
            for sheet in self.sheets_data:
                self._write_sheet_to_excel(writer, sheet)
                self._auto_adjust_column_width_xlsx(writer, sheet["sheet_name"])

            # Process raw sheets.
            for sheet in self.raw_sheets:
                header, *data = sheet.data
                df = pd.DataFrame(data, columns=header)
                df.to_excel(writer, index=False, sheet_name=sheet.label)
                worksheet: Worksheet = writer.sheets[sheet.label]

                # Apply formatting based on the provided styles.
                for row_idx, row in enumerate(
                    worksheet.iter_rows(max_row=len(sheet.data))
                ):
                    for col_idx, cell in enumerate(row):
                        cell_style = sheet.styles[row_idx][col_idx]
                        cell.font = cell_style.get("font")
                        cell.border = cell_style.get("border")
                        self._set_cell_format(cell, cell_style.get("type"))

                self._auto_adjust_column_width_xlsx(writer, sheet.label)

    def _write_sheet_to_excel(self, writer, sheet: SheetData):
        columns = [column.label for column in sheet["columns"]]
        df = pd.DataFrame(sheet["rows"], columns=columns)
        df.to_excel(writer, index=False, sheet_name=sheet["sheet_name"])
        worksheet: Worksheet = writer.sheets[sheet["sheet_name"]]

        # Add data validations.
        for validator in sheet["validators"]:
            worksheet.add_data_validation(validator)

        column_formulas = sheet.get("column_formulas") or {}
        if column_formulas:
            formula_start_row = sheet.get("formula_start_row") or 2
            formula_end_row = sheet.get("formula_end_row") or 0
            if not formula_end_row:
                formula_end_row = max(formula_start_row, len(sheet["rows"]) + 1)
            self._apply_column_formulas(
                worksheet, column_formulas, formula_start_row, formula_end_row
            )

        # Apply number formatting based on column type.
        format_max_row = sheet.get("format_max_row") or 2000
        for col_idx, column in enumerate(sheet["columns"]):
            for row in worksheet.iter_rows(
                min_row=2,
                max_row=format_max_row,
                min_col=col_idx + 1,
                max_col=col_idx + 1,
            ):
                for cell in row:
                    self._set_cell_format(cell, column.column_type)

        self._apply_excel_styling(writer, sheet)

    def _apply_column_formulas(
        self,
        worksheet: Worksheet,
        column_formulas: Dict[int, str],
        start_row: int,
        end_row: int,
    ) -> None:
        for row in range(start_row, end_row + 1):
            for col_idx, template in column_formulas.items():
                worksheet.cell(row=row, column=col_idx, value=template.format(row=row))

    def _apply_excel_styling(self, writer, sheet):
        worksheet = writer.sheets[sheet["sheet_name"]]
        # Reset header style: no border and non-bold font.
        for cell in worksheet[1]:
            cell.border = styles.Border(
                left=styles.Side(style=None),
                right=styles.Side(style=None),
                top=styles.Side(style=None),
                bottom=styles.Side(style=None),
            )
            cell.font = styles.Font(bold=False)

        # Bold headers if specified.
        if sheet["styles"].get("bold_headers"):
            for cell in worksheet[1]:
                cell.font = styles.Font(bold=True)

    def _write_xls(self, output):
        book = xlwt.Workbook()
        for sheet_data in self.sheets_data:
            self._write_sheet_to_xls(book, sheet_data)
        book.save(output)

    def _write_sheet_to_xls(self, book: xlwt.Workbook, sheet_data: SheetData):
        sheet: xlwt.Worksheet = book.add_sheet(sheet_data["sheet_name"])

        # Pre-define styles for headers and cells.
        bold_style = xlwt.XFStyle()
        bold_font = xlwt.Font()
        bold_font.bold = True
        bold_style.font = bold_font

        default_style = xlwt.XFStyle()

        left_aligned_num_style = xlwt.XFStyle()
        left_aligned_num_style.num_format_str = "#,##0"
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_LEFT
        left_aligned_num_style.alignment = alignment

        plain_style = xlwt.XFStyle()
        header_style = (
            bold_style if sheet_data["styles"].get("bold_headers") else default_style
        )

        # Write headers.
        for col_index, column in enumerate(sheet_data["columns"]):
            sheet.write(0, col_index, column.label, header_style)

        self._auto_adjust_column_width_xls(sheet, sheet_data)

        # Write rows with appropriate styles.
        for row_index, row in enumerate(sheet_data["rows"], start=1):
            for col_index, value in enumerate(row):
                cell_style = (
                    left_aligned_num_style
                    if isinstance(value, (int, float))
                    else plain_style
                )
                sheet.write(row_index, col_index, value, cell_style)

    def _auto_adjust_column_width_xlsx(self, writer: pd.ExcelWriter, sheet_name: str):
        worksheet = writer.sheets[sheet_name]
        for column in worksheet.columns:
            # Compute the maximum length among all cells in the column.
            max_length = max(
                (
                    (
                        0
                        if cell.data_type == "f"
                        else (len(str(cell.value)) if cell.value is not None else 0)
                    )
                    for cell in column
                ),
                default=0,
            )

            # Get header length for comparison
            header_cell = column[0]
            header_length = len(str(header_cell.value)) if header_cell.value else 0

            # Use the maximum of content length and header length
            content_width = max(max_length, header_length)

            # Set minimum width based on content type and typical requirements
            # Default minimum width for readability
            min_width = 12

            # Adjust minimum width based on header content to handle different column types
            header_text = str(header_cell.value).lower() if header_cell.value else ""

            if any(
                keyword in header_text
                for keyword in [
                    "name",
                    "description",
                    "address",
                    "organization",
                    "notes",
                ]
            ):
                min_width = 25  # Wide columns for text content
            elif any(
                keyword in header_text
                for keyword in ["fuel type", "fuel code", "manufacturer", "model"]
            ):
                min_width = 18  # Medium-wide for categorical data
            elif any(keyword in header_text for keyword in ["email", "phone"]):
                min_width = 20  # Contact information
            elif any(
                keyword in header_text for keyword in ["date", "serial", "registration"]
            ):
                min_width = 15  # Date and ID columns
            elif any(
                keyword in header_text
                for keyword in [
                    "quantity",
                    "units",
                    "compliance",
                    "value",
                    "ci",
                    "density",
                    "eer",
                    "energy",
                ]
            ):
                min_width = 14  # Numeric columns
            elif any(keyword in header_text for keyword in ["line"]):
                min_width = 8  # Short line number columns

            # Calculate final width with padding
            # Add extra padding for better readability (minimum 4 characters)
            padding = max(4, int(content_width * 0.1))  # 10% padding, minimum 4 chars
            calculated_width = content_width + padding

            # Apply minimum and maximum constraints
            final_width = max(min_width, calculated_width)
            # Set reasonable maximum to prevent extremely wide columns
            final_width = min(final_width, 50)

            column_letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[column_letter].width = final_width

    def _auto_adjust_column_width_xls(self, sheet, sheet_data):
        # Calculate the maximum width for each column using header and row values.
        for col_idx, column in enumerate(sheet_data["columns"]):
            header_length = len(str(column.label))
            cell_lengths = [
                len(str(row[col_idx]))
                for row in sheet_data["rows"]
                if len(row) > col_idx
            ]
            max_length = (
                max([header_length] + cell_lengths) if cell_lengths else header_length
            )
            sheet.col(col_idx).width = min(256 * (max_length + 2), MAX_XLS_WIDTH)

    def _write_csv(self, output):
        if self.sheets_data:
            columns = [col.label for col in self.sheets_data[0]["columns"]]
            df = pd.DataFrame(self.sheets_data[0]["rows"], columns=columns)
            df.to_csv(output, index=False)
