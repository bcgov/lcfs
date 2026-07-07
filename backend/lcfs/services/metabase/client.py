import base64
import io
import json
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

import requests
import structlog
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:  # pragma: no cover - chart fallback is table-only without Pillow.
    Image = ImageDraw = ImageFont = None

from lcfs.settings import settings

logger = structlog.get_logger(__name__)

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]\*:/\\?]")
WORKBOOK_REPORT_KEYWORDS = (
    "annual",
    "quarterly",
    "quarter",
    "monthly market",
    "monthly report",
    "monthly data",
    "month report",
    "month data",
)


@dataclass
class MetabaseTable:
    name: str
    columns: List[str]
    rows: List[List[Any]]
    visualization_settings: Dict[str, Any] = field(default_factory=dict)
    display: Optional[str] = None


@dataclass
class MetabaseDashboardReport:
    dashboard_name: str
    dashboard_description: Optional[str]
    dashboard_url: str
    tables: List[MetabaseTable]


class MetabaseClient:
    """Small client for the Metabase dashboard/card APIs used by reports."""

    def __init__(self):
        self.base_url = settings.metabase_base_url.rstrip("/")
        self.timeout = settings.metabase_request_timeout_seconds
        self.session = requests.Session()

    def fetch_credit_market_report(self) -> MetabaseDashboardReport:
        self._validate_configuration()
        self._authenticate()

        dashboard_id = settings.metabase_credit_market_dashboard_id
        dashboard = self._get_json(f"/api/dashboard/{dashboard_id}")
        cards = self._extract_cards(dashboard)
        tables = [
            table
            for table in (self._fetch_table(dashboard_id, card) for card in cards)
            if table
        ]
        return MetabaseDashboardReport(
            dashboard_name=dashboard.get("name") or "Credit Market Report",
            dashboard_description=dashboard.get("description"),
            dashboard_url=f"{self.base_url}/dashboard/{dashboard_id}",
            tables=tables,
        )

    def _validate_configuration(self) -> None:
        missing = []
        if not self.base_url:
            missing.append("metabase_base_url")
        if not settings.metabase_credit_market_dashboard_id:
            missing.append("metabase_credit_market_dashboard_id")
        if (
            not settings.metabase_api_key
            and not (settings.metabase_username and settings.metabase_password)
        ):
            missing.append(
                "metabase_api_key or metabase_username/metabase_password"
            )

        if missing:
            raise ValueError(f"Missing Metabase configuration: {', '.join(missing)}")

    def _authenticate(self) -> None:
        if settings.metabase_api_key:
            self.session.headers.update({"X-API-Key": settings.metabase_api_key})
            return

        response = self.session.post(
            f"{self.base_url}/api/session",
            json={
                "username": settings.metabase_username,
                "password": settings.metabase_password,
            },
            timeout=self.timeout,
        )
        response.raise_for_status()
        session_id = response.json().get("id")
        if not session_id:
            raise ValueError("Metabase session response did not include a session id")
        self.session.headers.update({"X-Metabase-Session": session_id})

    def _get_json(self, path: str) -> Any:
        response = self.session.get(f"{self.base_url}{path}", timeout=self.timeout)
        response.raise_for_status()
        return response.json()

    def _post_json(self, path: str, payload: Optional[Dict[str, Any]] = None) -> Any:
        url = f"{self.base_url}{path}"
        for attempt in range(1, settings.metabase_query_poll_attempts + 1):
            response = self.session.post(
                url,
                json=payload or {},
                timeout=self.timeout,
            )
            result = self._response_json(response)
            if isinstance(result, dict) and result.get("data"):
                return result

            if response.status_code != requests.codes.accepted:
                response.raise_for_status()
                return result

            logger.info(
                "Metabase query accepted but still running; retrying",
                path=path,
                attempt=attempt,
                max_attempts=settings.metabase_query_poll_attempts,
                retry_delay_seconds=settings.metabase_query_poll_interval_seconds,
            )
            if attempt < settings.metabase_query_poll_attempts:
                time.sleep(settings.metabase_query_poll_interval_seconds)

        response.raise_for_status()
        raise TimeoutError(
            "Metabase query did not complete after "
            f"{settings.metabase_query_poll_attempts} attempts: {path}"
        )

    def _response_json(self, response: requests.Response) -> Any:
        try:
            return response.json()
        except ValueError:
            return None

    def _extract_cards(self, dashboard: Dict[str, Any]) -> List[Dict[str, Any]]:
        cards = dashboard.get("dashcards") or dashboard.get("ordered_cards") or []
        return [card for card in cards if self._card_id(card)]

    def _fetch_table(
        self, dashboard_id: int, dashboard_card: Dict[str, Any]
    ) -> Optional[MetabaseTable]:
        card_id = self._card_id(dashboard_card)
        if card_id is None:
            return None

        dashboard_card_id = dashboard_card.get("id")
        if dashboard_card_id:
            try:
                result = self._post_json(
                    f"/api/dashboard/{dashboard_id}/dashcard/{dashboard_card_id}/card/{card_id}/query"
                )
            except requests.HTTPError as exc:
                status_code = (
                    exc.response.status_code if exc.response is not None else None
                )
                if status_code not in (400, 404, 405):
                    raise
                result = self._post_json(f"/api/card/{card_id}/query")
        else:
            result = self._post_json(f"/api/card/{card_id}/query")

        data = result.get("data") or {}
        rows = data.get("rows") or []
        cols = data.get("cols") or []
        if not cols and not rows:
            return None

        columns = [
            col.get("display_name") or col.get("name") or f"Column {index + 1}"
            for index, col in enumerate(cols)
        ]
        card = dashboard_card.get("card") or {}
        name = (
            dashboard_card.get("visualization_settings", {}).get("card.title")
            or card.get("name")
            or dashboard_card.get("name")
            or f"Card {card_id}"
        )

        return MetabaseTable(
            name=name,
            columns=columns,
            rows=rows,
            visualization_settings=card.get("visualization_settings") or {},
            display=card.get("display"),
        )

    def _card_id(self, dashboard_card: Dict[str, Any]) -> Optional[int]:
        if dashboard_card.get("card_id"):
            return dashboard_card["card_id"]
        card = dashboard_card.get("card") or {}
        return card.get("id")


class CreditMarketReportBuilder:
    """Builds the email payload pieces for the monthly credit market report."""

    def build_workbook(self, report: MetabaseDashboardReport) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        workbook_tables = self._workbook_tables(report)

        for index, table in enumerate(workbook_tables, start=1):
            workbook_table = self._workbook_table(table)
            sheet = workbook.create_sheet(self._sheet_title(workbook_table.name, index))
            self._write_table_sheet(sheet, workbook_table)

        if not workbook_tables:
            sheet = workbook.create_sheet("No data")
            sheet["A1"] = (
                "No annual, monthly, or quarterly report data was returned from Metabase."
            )

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def build_attachment(self, report: MetabaseDashboardReport) -> Dict[str, Any]:
        workbook = self.build_workbook(report)
        generated_date = datetime.now().strftime("%Y-%m-%d")
        return {
            "filename": f"credit-market-report-{generated_date}.xlsx",
            "contentType": EXCEL_CONTENT_TYPE,
            "encoding": "base64",
            "content": base64.b64encode(workbook).decode("ascii"),
        }

    def build_email_context(self, report: MetabaseDashboardReport) -> Dict[str, Any]:
        now = datetime.now()
        sections = self._dashboard_sections(report)
        return {
            "environment": settings.environment.lower(),
            "subject": f"{report.dashboard_name} - Monthly Credit Market Report",
            "dashboard_name": report.dashboard_name,
            "dashboard_description": report.dashboard_description,
            "generated_at": f"{now.strftime('%B')} {now.day}, {now.year}",
            "metric_cards": sections["metric_cards"],
            "summary_cards": sections["summary_cards"],
            "chart_sections": sections["chart_sections"],
            "attached_table_names": [
                table.name for table in self._workbook_tables(report)
            ],
        }

    def _write_table_sheet(self, sheet, table: MetabaseTable) -> None:
        header_fill = PatternFill("solid", fgColor="003366")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(bottom=Side(style="thin", color="B7C9D6"))

        for column_index, column_name in enumerate(table.columns, start=1):
            cell = sheet.cell(
                row=1,
                column=column_index,
                value=self._display_column_name(column_name),
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for row_index, row in enumerate(table.rows, start=2):
            for column_index, value in enumerate(row, start=1):
                column_name = (
                    table.columns[column_index - 1]
                    if len(table.columns) >= column_index
                    else ""
                )
                cell = sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=self._excel_value(
                        value,
                        table_name=table.name,
                        column_name=column_name,
                    ),
                )
                number_format = self._excel_number_format(
                    table_name=table.name,
                    column_name=column_name,
                )
                if number_format:
                    cell.number_format = number_format

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        self._size_columns(
            sheet,
            [self._display_column_name(column) for column in table.columns],
            table.rows,
        )

    def _size_columns(self, sheet, columns: List[str], rows: List[List[Any]]) -> None:
        for index, header in enumerate(columns, start=1):
            values = [header] + [
                (
                    ""
                    if len(row) < index or row[index - 1] is None
                    else str(row[index - 1])
                )
                for row in rows[:100]
            ]
            width = min(max(len(value) for value in values) + 2, 60)
            sheet.column_dimensions[get_column_letter(index)].width = width

    def _sheet_title(self, name: str, index: int) -> str:
        title = INVALID_SHEET_TITLE_CHARS.sub(" ", name).strip() or f"Table {index}"
        return title[:31]

    def _excel_value(
        self, value: Any, table_name: str = "", column_name: str = ""
    ) -> Any:
        if isinstance(value, (dict, list)):
            return json.dumps(value, default=str)
        month = self._parse_month(value)
        lower_table = table_name.lower()
        lower_column = column_name.lower()
        if month and "calculated effective date" in lower_column:
            if "annual" in lower_table and "year" in lower_column:
                return month.year
            if "quarter" in lower_table and "quarter" in lower_column:
                quarter = ((month.month - 1) // 3) + 1
                return f"Q{quarter}, {month.year}"
        if month:
            return self._month_label(month, value)
        if self._is_currency_workbook_column(table_name, column_name):
            numeric = self._numeric_value(value)
            if numeric is not None:
                return float(numeric.quantize(Decimal("0.01")))
        if self._is_volume_workbook_column(column_name):
            numeric = self._numeric_value(value)
            if numeric is not None:
                return (
                    int(numeric) if numeric == numeric.to_integral() else float(numeric)
                )
        return value

    def _excel_number_format(self, table_name: str, column_name: str) -> Optional[str]:
        if self._is_currency_workbook_column(table_name, column_name):
            return "$#,##0.00"
        if self._is_volume_workbook_column(column_name):
            return "#,##0"
        return None

    def _is_currency_workbook_column(self, table_name: str, column_name: str) -> bool:
        lower_table = table_name.lower()
        lower_column = re.sub(r"\s+", " ", str(column_name)).strip().lower()
        if "price" in lower_column:
            return True
        if "weighted average" in lower_column:
            return True
        return "annual" in lower_table and "sum of transfer value" in lower_column

    def _is_volume_workbook_column(self, column_name: str) -> bool:
        lower_column = re.sub(r"\s+", " ", str(column_name)).strip().lower()
        return (
            lower_column
            in {
                "sum of quantity",
                "volume (credits)",
                "credit volume",
                "category a1 - credit volume",
                "category a - credit volume",
                "category b - credit volume",
                "category c - credit volume",
            }
            and "transfer" not in lower_column
        )

    def _workbook_tables(self, report: MetabaseDashboardReport) -> List[MetabaseTable]:
        tables = [
            table
            for table in report.tables
            if any(
                keyword in table.name.lower() for keyword in WORKBOOK_REPORT_KEYWORDS
            )
        ]
        return sorted(tables, key=self._workbook_table_order)

    def _workbook_table_order(self, table: MetabaseTable) -> int:
        lower_name = table.name.lower()
        if "monthly" in lower_name or "month" in lower_name:
            return 0
        if "quarter" in lower_name:
            return 1
        if "annual" in lower_name:
            return 2
        return 3

    def _workbook_table(self, table: MetabaseTable) -> MetabaseTable:
        order = self._workbook_column_order(table.columns)
        if self._is_monthly_report_table(table):
            order = self._monthly_workbook_column_order(table.columns, order)

        if order == list(range(len(table.columns))):
            return table

        return MetabaseTable(
            name=table.name,
            columns=[
                self._display_column_name(table.columns[index]) for index in order
            ],
            rows=[
                [row[index] if len(row) > index else None for index in order]
                for row in table.rows
            ],
            visualization_settings=table.visualization_settings,
            display=table.display,
        )

    def _workbook_column_order(self, columns: List[str]) -> List[int]:
        indexed_columns = list(enumerate(columns))
        transfer_indexes = [
            index
            for index, column in indexed_columns
            if self._is_transfer_count_column(column)
        ]
        volume_indexes = [
            index
            for index, column in indexed_columns
            if self._is_credit_volume_column(column)
        ]
        prioritized = []
        if transfer_indexes:
            prioritized.append(transfer_indexes[0])
        if volume_indexes:
            prioritized.append(volume_indexes[0])

        if not prioritized:
            return [index for index, _ in indexed_columns]

        leading = [
            index
            for index, _ in indexed_columns
            if index < prioritized[0] and index not in prioritized
        ]
        remaining = [
            index
            for index, _ in indexed_columns
            if index not in set(leading + prioritized)
        ]
        return leading + prioritized + remaining

    def _monthly_workbook_column_order(
        self, columns: List[str], order: List[int]
    ) -> List[int]:
        a1_indexes = [
            index
            for index in order
            for column in [columns[index]]
            if self._is_category_a1_column(column)
        ]
        non_a1_indexes = [
            index
            for index in order
            for column in [columns[index]]
            if not self._is_category_a1_column(column)
        ]
        front = non_a1_indexes[:6]
        back = non_a1_indexes[6:]
        return front + a1_indexes + back

    def _is_transfer_count_column(self, column: str) -> bool:
        normalized = re.sub(r"[^a-z0-9]+", " ", str(column).lower()).strip()
        if "category" in normalized:
            return False
        return normalized in {
            "count",
            "transfers",
            "transfers number",
            "transfers #",
            "number of transfers",
        }

    def _is_credit_volume_column(self, column: str) -> bool:
        normalized = re.sub(r"[^a-z0-9]+", " ", str(column).lower()).strip()
        if "category" in normalized:
            return False
        return normalized in {
            "sum of quantity",
            "volume credits",
            "volumes credits",
            "credit volume",
        }

    def _is_category_a1_column(self, column: str) -> bool:
        normalized = column.lower().replace(" ", "")
        return "categorya1" in normalized

    def _dashboard_sections(
        self, report: MetabaseDashboardReport
    ) -> Dict[str, List[Dict[str, Any]]]:
        metric_cards = self._monthly_metric_cards(report)
        summary_cards = []
        chart_sections = []

        for table in report.tables:
            table_context = self._table_context(table)
            lower_name = table.name.lower()
            if metric_cards and self._is_monthly_report_table(table):
                continue
            if "monthly price" in lower_name or "all time" in lower_name:
                summary_cards.append(table_context)
            elif not metric_cards and self._is_metric_card(table):
                metric_cards.append(self._metric_context(table))
            elif "trend" in lower_name or "volume" in lower_name:
                chart_sections.append(table_context)

        if not metric_cards:
            metric_cards = [
                self._metric_context(table)
                for table in report.tables
                if self._is_compact_table(table)
            ][:3]
            if not metric_cards:
                logger.warning(
                    "Credit market report did not find expected metric card data",
                    dashboard_name=report.dashboard_name,
                )

        shown_names = {
            section["name"] for section in metric_cards + summary_cards + chart_sections
        }
        remaining_tables = [
            self._table_context(table)
            for table in report.tables
            if table.name not in shown_names and table.rows
        ]
        chart_sections.extend(remaining_tables[:2])
        if not summary_cards:
            logger.warning(
                "Credit market report did not find expected summary table data",
                dashboard_name=report.dashboard_name,
            )
        if not chart_sections:
            logger.warning(
                "Credit market report did not find expected trend or volume table data",
                dashboard_name=report.dashboard_name,
            )

        return {
            "metric_cards": metric_cards[:3],
            "summary_cards": summary_cards[:2],
            "chart_sections": chart_sections[:2],
        }

    def _monthly_metric_cards(
        self, report: MetabaseDashboardReport
    ) -> List[Dict[str, Any]]:
        monthly_table = next(
            (
                table
                for table in report.tables
                if self._is_monthly_report_table(table) and table.rows
            ),
            None,
        )
        if not monthly_table:
            logger.warning(
                "Credit market report did not find monthly market report data",
                dashboard_name=report.dashboard_name,
            )
            return []

        month_index = self._column_index(monthly_table, ["month"])
        if month_index is None:
            return []

        current_row = self._latest_month_row(monthly_table, month_index)
        if not current_row or len(current_row) <= month_index:
            return []

        current_month = self._parse_month(current_row[month_index])
        previous_year_row = self._same_month_previous_year_row(
            monthly_table, current_month, month_index
        )
        current_label = self._month_label(current_month, current_row[month_index])
        previous_label = self._previous_year_label(current_month)

        metric_specs = [
            {
                "name": "Transfers (Number)",
                "columns": ["transfers (#)", "transfers", "sum of quantity"],
                "currency": False,
            },
            {
                "name": "Total Volume (Credits)",
                "columns": ["volume (credits)", "credit volume", "count"],
                "currency": False,
            },
            {
                "name": "Category A1 Transfers (weighted average)",
                "columns": [
                    "category a1 - average price",
                    "category a - average price",
                    "weighted average price",
                    "weighted average",
                ],
                "currency": True,
            },
        ]

        cards = []
        for spec in metric_specs:
            column_index = self._column_index(monthly_table, spec["columns"])
            if column_index is None or len(current_row) <= column_index:
                continue

            current_value = current_row[column_index]
            previous_value = (
                previous_year_row[column_index]
                if previous_year_row and len(previous_year_row) > column_index
                else None
            )
            current_number = self._numeric_value(current_value)
            previous_number = self._numeric_value(previous_value)
            comparison = ""
            direction = "neutral"
            if current_number is not None and previous_number not in (None, Decimal(0)):
                change = ((current_number - previous_number) / previous_number) * 100
                direction = "up" if change >= 0 else "down"
                comparison = (
                    f"{self._format_percentage(abs(change))} vs. {previous_label}: "
                    f"{self._format_metric_value(previous_value, spec['currency'])}"
                )

            cards.append(
                {
                    "name": spec["name"],
                    "primary": self._format_metric_value(
                        current_value, spec["currency"]
                    ),
                    "subtitle": current_label,
                    "comparison": comparison,
                    "direction": direction,
                }
            )

        return cards

    def _metric_context(self, table: MetabaseTable) -> Dict[str, Any]:
        values = self._flatten_values(table.rows)
        primary = self._format_value(values[0]) if values else "No data"
        subtitle = self._format_value(values[1]) if len(values) > 1 else ""
        comparison = self._format_value(values[2]) if len(values) > 2 else ""
        direction = "neutral"
        if "-" in comparison or "down" in comparison.lower():
            direction = "down"
        elif comparison:
            direction = "up"

        return {
            "name": table.name,
            "primary": primary,
            "subtitle": subtitle,
            "comparison": comparison,
            "direction": direction,
        }

    def _table_context(self, table: MetabaseTable) -> Dict[str, Any]:
        rows = self._table_rows(table)
        return {
            "name": table.name,
            "columns": (
                ["Metric", "Value"]
                if self._is_single_row_summary(table)
                else [self._display_column_name(column) for column in table.columns]
            ),
            "rows": rows,
            "row_count": len(table.rows),
            "image_data_uri": self._chart_image_data_uri(table),
        }

    def _is_metric_card(self, table: MetabaseTable) -> bool:
        lower_name = table.name.lower()
        if any(
            keyword in lower_name
            for keyword in ("transfers (number)", "total volume", "category a1")
        ):
            return True
        return table.display in {"scalar", "smartscalar"} or self._is_compact_table(
            table
        )

    def _is_compact_table(self, table: MetabaseTable) -> bool:
        return len(table.rows) <= 2 and len(table.columns) <= 4

    def _is_monthly_report_table(self, table: MetabaseTable) -> bool:
        lower_name = table.name.lower()
        return "monthly" in lower_name and (
            "report" in lower_name or "market" in lower_name
        )

    def _is_single_row_summary(self, table: MetabaseTable) -> bool:
        return len(table.rows) == 1 and len(table.columns) > 1

    def _table_rows(self, table: MetabaseTable) -> List[List[str]]:
        if self._is_single_row_summary(table):
            row = table.rows[0]
            return [
                [
                    self._display_column_name(column),
                    (
                        self._format_summary_value(column, row[index])
                        if len(row) > index
                        else ""
                    ),
                ]
                for index, column in enumerate(table.columns)
            ]
        return [[self._format_value(value) for value in row] for row in table.rows[:8]]

    def _format_summary_value(self, column: str, value: Any) -> str:
        lower_column = column.lower()
        if (
            "price" in lower_column
            or "ca$" in lower_column
            or "cad" in lower_column
            or "transfer value" in lower_column
        ):
            return self._format_metric_value(value, currency=True)
        return self._format_value(value)

    def _display_column_name(self, column: str) -> str:
        normalized = re.sub(r"\s+", " ", str(column)).strip().lower()
        if normalized == "sum of quantity":
            return "Volume (Credits)"
        if normalized == "count":
            return "Transfers"
        return str(column)

    def _flatten_values(self, rows: List[List[Any]]) -> List[Any]:
        return [value for row in rows for value in row if value not in (None, "")]

    def _format_value(self, value: Any) -> str:
        month = self._parse_month(value)
        if month:
            return self._month_label(month, value)
        if isinstance(value, float):
            return f"{value:,.2f}"
        if isinstance(value, int):
            return f"{value:,}"
        return "" if value is None else str(value)

    def _column_index(
        self, table: MetabaseTable, column_names: List[str]
    ) -> Optional[int]:
        normalized_targets = [name.lower() for name in column_names]
        normalized_columns = [column.lower() for column in table.columns]
        for target in normalized_targets:
            for index, normalized_column in enumerate(normalized_columns):
                if target in normalized_column:
                    return index
        return None

    def _parse_month(self, value: Any) -> Optional[datetime]:
        if isinstance(value, datetime):
            return value
        text = str(value).strip()
        iso_text = text.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(iso_text)
        except ValueError:
            pass
        for date_format in ("%B, %Y", "%B %Y", "%b, %Y", "%b %Y", "%Y-%m", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, date_format)
            except ValueError:
                continue
        return None

    def _same_month_previous_year_row(
        self,
        table: MetabaseTable,
        current_month: Optional[datetime],
        month_index: int,
    ) -> Optional[List[Any]]:
        if not current_month:
            return None
        for row in table.rows[1:]:
            if len(row) <= month_index:
                continue
            row_month = self._parse_month(row[month_index])
            if (
                row_month
                and row_month.year == current_month.year - 1
                and row_month.month == current_month.month
            ):
                return row
        return None

    def _latest_month_row(
        self, table: MetabaseTable, month_index: int
    ) -> Optional[List[Any]]:
        dated_rows = []
        for row in table.rows:
            if len(row) <= month_index:
                continue
            row_month = self._parse_month(row[month_index])
            if row_month:
                dated_rows.append(
                    ((row_month.year, row_month.month, row_month.day), row)
                )
        if not dated_rows:
            return table.rows[0] if table.rows else None
        return max(dated_rows, key=lambda item: item[0])[1]

    def _month_label(self, month: Optional[datetime], fallback: Any) -> str:
        if not month:
            return self._format_value(fallback)
        return f"{month.strftime('%b')} {month.year}"

    def _previous_year_label(self, month: Optional[datetime]) -> str:
        if not month:
            return "previous year"
        return f"{month.strftime('%b')} {month.year - 1}"

    def _numeric_value(self, value: Any) -> Optional[Decimal]:
        if value is None:
            return None
        text = str(value).strip().replace(",", "")
        multiplier = Decimal(1)
        if text.lower().endswith("k"):
            multiplier = Decimal(1000)
            text = text[:-1]
        text = re.sub(r"[^0-9.\-]", "", text)
        if text in {"", "-", "."}:
            return None
        try:
            return Decimal(text) * multiplier
        except InvalidOperation:
            return None

    def _format_metric_value(self, value: Any, currency: bool) -> str:
        numeric_value = self._numeric_value(value)
        if numeric_value is None:
            return self._format_value(value)
        if currency:
            return f"CA${numeric_value:,.2f}"
        if numeric_value == numeric_value.to_integral():
            return f"{int(numeric_value):,}"
        return f"{numeric_value:,.2f}"

    def _format_percentage(self, value: Decimal) -> str:
        text = f"{value:.2f}".rstrip("0").rstrip(".")
        return f"{text}%"

    def _chart_image_data_uri(self, table: MetabaseTable) -> Optional[str]:
        lower_name = table.name.lower()
        if "trend" not in lower_name and "volume" not in lower_name:
            return None
        image = ChartImageRenderer().render(
            table, chart_type="bar" if "volume" in lower_name else "line"
        )
        if not image:
            return None
        encoded = base64.b64encode(image).decode("ascii")
        return f"data:image/png;base64,{encoded}"


class ChartImageRenderer:
    width = 980
    height = 320
    margin_left = 56
    margin_right = 28
    margin_top = 52
    margin_bottom = 36
    bg = (255, 255, 255)
    axis = (213, 218, 226)
    grid = (232, 235, 240)
    line_colors = [(166, 133, 203), (247, 159, 92), (247, 204, 67)]
    bar_color = (93, 154, 222)

    def render(self, table: MetabaseTable, chart_type: str) -> Optional[bytes]:
        if Image is None:
            logger.warning("Pillow is unavailable; rendering chart table fallback")
            return None
        series = self._series(table)
        if not series:
            return None
        return (
            self._render_bar_chart(table, series[0])
            if chart_type == "bar"
            else self._render_line_chart(table, series[:3])
        )

    def _render_line_chart(
        self, table: MetabaseTable, series: List[Dict[str, Any]]
    ) -> bytes:
        image = Image.new("RGB", (self.width, self.height), self.bg)
        draw = ImageDraw.Draw(image)
        font = self._font(12)
        small_font = self._font(10)
        self._draw_grid(draw)
        self._draw_legend(draw, series, chart_type="line", font=small_font)

        values_for_scale = [value for item in series for value in item["values"]]
        min_value = min(values_for_scale)
        max_value = max(values_for_scale)
        if max_value == min_value:
            max_value += 1

        for index, item in enumerate(series):
            values = item["values"]
            color = self.line_colors[index % len(self.line_colors)]
            points = self._scaled_points(values, min_value, max_value)
            for start, end in zip(points, points[1:]):
                draw.line([start, end], fill=color, width=3)
            for point_index, point in enumerate(points):
                draw.ellipse(
                    (point[0] - 4, point[1] - 4, point[0] + 4, point[1] + 4),
                    fill=color,
                    outline=(255, 255, 255),
                )
                if self._should_label_point(point_index, len(points)):
                    self._draw_centered_text(
                        draw,
                        self._format_chart_number(
                            values[point_index], currency=item["currency"]
                        ),
                        point[0],
                        max(self.margin_top - 2, point[1] - 18),
                        font,
                        color,
                    )

        self._draw_x_labels(draw, self._labels(table), small_font)
        return self._image_png(image)

    def _render_bar_chart(self, table: MetabaseTable, item: Dict[str, Any]) -> bytes:
        values = item["values"]
        max_value = max(values) if values else 0
        image = Image.new("RGB", (self.width, self.height), self.bg)
        draw = ImageDraw.Draw(image)
        font = self._font(11)
        small_font = self._font(10)
        self._draw_grid(draw)
        self._draw_legend(draw, [item], chart_type="bar", font=small_font)
        if max_value <= 0:
            return self._image_png(image)

        chart_width = self.width - self.margin_left - self.margin_right
        usable_height = self.height - self.margin_top - self.margin_bottom
        step = chart_width / max(len(values), 1)
        bar_width = max(2, int(step * 0.58))
        base_y = self.height - self.margin_bottom
        for index, value in enumerate(values):
            x = int(self.margin_left + index * step + (step - bar_width) / 2)
            bar_height = int((value / max_value) * usable_height)
            top_y = base_y - bar_height
            draw.rectangle(
                (x, top_y, x + bar_width, base_y),
                fill=self.bar_color,
                outline=self.bar_color,
            )
            if index in self._bar_label_indexes(values):
                self._draw_centered_text(
                    draw,
                    self._format_chart_number(value, currency=item["currency"]),
                    x + int(bar_width / 2),
                    max(self.margin_top - 2, top_y - 16),
                    font,
                    (76, 87, 115),
                )

        self._draw_x_labels(draw, self._labels(table), small_font)
        return self._image_png(image)

    def _draw_grid(self, draw) -> None:
        left = self.margin_left
        right = self.width - self.margin_right
        top = self.margin_top
        bottom = self.height - self.margin_bottom
        for index in range(6):
            y = top + int(((bottom - top) / 5) * index)
            draw.line((left, y, right, y), fill=self.grid, width=1)
        draw.line((left, bottom, right, bottom), fill=self.axis, width=1)
        draw.line((left, top, left, bottom), fill=self.axis, width=1)

    def _draw_legend(
        self, draw, series: List[Dict[str, Any]], chart_type: str, font
    ) -> None:
        x = self.margin_left
        y = 16
        for index, item in enumerate(series):
            color = (
                self.bar_color
                if chart_type == "bar"
                else self.line_colors[index % len(self.line_colors)]
            )
            if chart_type == "bar":
                draw.rectangle((x, y + 3, x + 12, y + 13), fill=color, outline=color)
            else:
                draw.line((x, y + 8, x + 16, y + 8), fill=color, width=3)
                draw.ellipse((x + 6, y + 4, x + 12, y + 10), fill=color)
            label = self._legend_label(item["name"])
            draw.text((x + 22, y), label, fill=(76, 87, 115), font=font)
            x += 22 + self._estimate_text_width(label)

    def _series(self, table: MetabaseTable) -> List[Dict[str, Any]]:
        numeric_columns = []
        for column_index, column in enumerate(table.columns):
            column_name = column.lower()
            if any(
                keyword in column_name
                for keyword in ("date", "month", "year", "period")
            ):
                continue
            values = []
            for row in table.rows:
                if len(row) <= column_index:
                    continue
                parsed = self._number(row[column_index])
                if parsed is not None:
                    values.append(parsed)
            if len(values) >= 2:
                numeric_columns.append(
                    {
                        "name": self._display_column_name(column),
                        "values": values,
                        "currency": self._is_currency_column(column),
                    }
                )
        return numeric_columns

    def _labels(self, table: MetabaseTable) -> List[str]:
        label_index = 0
        for index, column in enumerate(table.columns):
            if any(
                keyword in column.lower()
                for keyword in ("date", "month", "year", "period")
            ):
                label_index = index
                break
        return [
            self._format_chart_label(row[label_index])
            for row in table.rows
            if len(row) > label_index
        ]

    def _format_chart_label(self, value: Any) -> str:
        if isinstance(value, datetime):
            return f"{value.strftime('%b')} {value.year}"
        text = str(value).strip()
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
            return f"{parsed.strftime('%b')} {parsed.year}"
        except ValueError:
            pass
        for date_format in ("%B, %Y", "%B %Y", "%b, %Y", "%Y-%m", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(text, date_format)
                return f"{parsed.strftime('%b')} {parsed.year}"
            except ValueError:
                continue
        return text

    def _number(self, value: Any) -> Optional[float]:
        if isinstance(value, (int, float)):
            return float(value)
        if value is None:
            return None
        text = str(value).strip().replace(",", "")
        multiplier = 1
        if text.lower().endswith("k"):
            multiplier = 1000
            text = text[:-1]
        text = re.sub(r"[^0-9.\-]", "", text)
        if text in {"", "-", "."}:
            return None
        try:
            return float(text) * multiplier
        except ValueError:
            return None

    def _is_currency_column(self, column: str) -> bool:
        lower_column = column.lower()
        return any(
            keyword in lower_column
            for keyword in ("price", "ca$", "cad", "transfer value", "weighted average")
        )

    def _display_column_name(self, column: str) -> str:
        normalized = re.sub(r"\s+", " ", str(column)).strip().lower()
        if normalized == "sum of quantity":
            return "Volume (Credits)"
        if normalized == "count":
            return "Transfers"
        return str(column)

    def _legend_label(self, name: str) -> str:
        label = re.sub(r"\s+", " ", str(name)).strip()
        if len(label) > 28:
            label = f"{label[:25]}..."
        return label

    def _estimate_text_width(self, text: str) -> int:
        return max(24, len(text) * 7)

    def _scaled_points(
        self, values: List[float], min_value: float, max_value: float
    ) -> List[tuple]:
        chart_width = self.width - self.margin_left - self.margin_right
        usable_height = self.height - self.margin_top - self.margin_bottom
        points = []
        for index, value in enumerate(values):
            x = self.margin_left + int((index / max(len(values) - 1, 1)) * chart_width)
            y = self.margin_top + int(
                (1 - ((value - min_value) / (max_value - min_value))) * usable_height
            )
            points.append((x, y))
        return points

    def _draw_x_labels(self, draw, labels: List[str], font) -> None:
        if not labels:
            return
        chart_width = self.width - self.margin_left - self.margin_right
        y = self.height - self.margin_bottom + 8
        for index in self._x_label_indexes(len(labels)):
            x = self.margin_left + int((index / max(len(labels) - 1, 1)) * chart_width)
            self._draw_centered_text(draw, labels[index], x, y, font, (76, 87, 115))

    def _x_label_indexes(self, count: int) -> List[int]:
        if count <= 6:
            return list(range(count))
        indexes = {0, count - 1}
        for step in range(1, 5):
            indexes.add(round((count - 1) * step / 5))
        return sorted(indexes)

    def _should_label_point(self, index: int, count: int) -> bool:
        if count <= 12:
            return True
        return index in self._x_label_indexes(count)

    def _bar_label_indexes(self, values: List[float]) -> set:
        if len(values) <= 24:
            return set(range(len(values)))
        top_indexes = sorted(
            range(len(values)), key=lambda index: values[index], reverse=True
        )[:12]
        return set(top_indexes + self._x_label_indexes(len(values)))

    def _format_chart_number(self, value: float, currency: bool) -> str:
        prefix = "CA$" if currency else ""
        absolute_value = abs(value)
        if absolute_value >= 1_000_000:
            return f"{prefix}{value / 1_000_000:.1f}M"
        if absolute_value >= 1_000:
            return f"{prefix}{value / 1_000:.1f}k"
        if currency:
            return f"{prefix}{value:,.0f}"
        if value == int(value):
            return f"{int(value):,}"
        return f"{value:,.1f}"

    def _draw_centered_text(self, draw, text: str, x: int, y: int, font, fill) -> None:
        bbox = draw.textbbox((0, 0), text, font=font)
        width = bbox[2] - bbox[0]
        draw.text((x - int(width / 2), y), text, fill=fill, font=font)

    def _font(self, size: int):
        try:
            return ImageFont.truetype("DejaVuSans.ttf", size)
        except Exception:
            return ImageFont.load_default()

    def _image_png(self, image) -> bytes:
        output = io.BytesIO()
        image.save(output, format="PNG")
        return output.getvalue()
