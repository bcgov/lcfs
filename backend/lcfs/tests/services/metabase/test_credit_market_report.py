import base64
import io
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import load_workbook

from lcfs.services.jobs.jobs import send_monthly_credit_market_report
from lcfs.services.metabase.client import (
    CreditMarketReportBuilder,
    MetabaseClient,
    MetabaseDashboardReport,
    MetabaseTable,
)


@pytest.fixture
def mock_app():
    app = MagicMock()
    app.state.db_session_factory = MagicMock()
    return app


def test_builds_workbook_with_only_report_data_tabs():
    report = MetabaseDashboardReport(
        dashboard_name="Credit Market",
        dashboard_description=None,
        dashboard_url="https://metabase.example/dashboard/1",
        tables=[
            MetabaseTable(
                name="Annual Report Data",
                columns=[
                    "Year",
                    "Volume (Credits)",
                    "Transfers (#)",
                    "Weighted average",
                    "Sum of Transfer Value (CA$)",
                ],
                rows=[[2026, 100, 2, "152.98", "129299893.93"]],
            ),
            MetabaseTable(
                name="Quarterly Market Report",
                columns=[
                    "Quarter",
                    "Volume (Credits)",
                    "Transfers (#)",
                    "Weighted average price",
                ],
                rows=[["Q1, 2026", 75, 3, "160.28"]],
            ),
            MetabaseTable(
                name="Monthly Market Report",
                columns=[
                    "Month",
                    "Sum of Quantity",
                    "Count",
                    "Weighted average price",
                    "Min of Price Per Unit (CA$)",
                    "Max of Price Per Unit (CA$)",
                    "Category A - Transfers",
                    "Category A1 - Transfers",
                    "Category A1 - Credit Volume",
                    "Category A1 - Average Price",
                ],
                rows=[
                    [
                        "June, 2026",
                        75491,
                        6,
                        "$143.32",
                        "$108.50",
                        "$169.92",
                        6,
                        2,
                        7444,
                        "$120.30",
                    ]
                ],
            ),
            MetabaseTable(
                name="Monthly Price (CA$ per credit)",
                columns=["Metric", "Value"],
                rows=[["Weighted average price", "CA$143.32"]],
            ),
        ],
    )

    attachment = CreditMarketReportBuilder().build_attachment(report)
    workbook_bytes = base64.b64decode(attachment["content"])
    workbook = load_workbook(io.BytesIO(workbook_bytes))

    assert workbook.sheetnames == [
        "Monthly Market Report",
        "Quarterly Market Report",
        "Annual Report Data",
    ]
    assert workbook["Monthly Market Report"]["B1"].value == "Transfers"
    assert workbook["Monthly Market Report"]["C1"].value == "Volume (Credits)"
    assert workbook["Monthly Market Report"]["B2"].value == 6
    assert workbook["Monthly Market Report"]["C2"].value == 75491
    assert workbook["Monthly Market Report"]["D2"].value == 143.32
    assert workbook["Monthly Market Report"]["D2"].number_format == "$#,##0.00"
    assert workbook["Monthly Market Report"]["E2"].value == 108.5
    assert workbook["Monthly Market Report"]["E2"].number_format == "$#,##0.00"
    assert workbook["Monthly Market Report"]["F2"].value == 169.92
    assert workbook["Monthly Market Report"]["F2"].number_format == "$#,##0.00"
    assert workbook["Monthly Market Report"]["G2"].value == 2
    assert workbook["Monthly Market Report"]["I2"].value == 120.3
    assert workbook["Monthly Market Report"]["I2"].number_format == "$#,##0.00"
    assert workbook["Monthly Market Report"]["G1"].value == "Category A1 - Transfers"
    assert (
        workbook["Monthly Market Report"]["H1"].value == "Category A1 - Credit Volume"
    )
    assert (
        workbook["Monthly Market Report"]["I1"].value == "Category A1 - Average Price"
    )
    assert workbook["Monthly Market Report"]["J1"].value == "Category A - Transfers"
    assert workbook["Quarterly Market Report"]["B1"].value == "Transfers (#)"
    assert workbook["Quarterly Market Report"]["C1"].value == "Volume (Credits)"
    assert workbook["Quarterly Market Report"]["B2"].value == 3
    assert workbook["Quarterly Market Report"]["C2"].value == 75
    assert workbook["Quarterly Market Report"]["D2"].value == 160.28
    assert workbook["Quarterly Market Report"]["D2"].number_format == "$#,##0.00"
    assert workbook["Annual Report Data"]["B1"].value == "Transfers (#)"
    assert workbook["Annual Report Data"]["C1"].value == "Volume (Credits)"
    assert workbook["Annual Report Data"]["B2"].value == 2
    assert workbook["Annual Report Data"]["C2"].value == 100
    assert workbook["Annual Report Data"]["D2"].value == 152.98
    assert workbook["Annual Report Data"]["D2"].number_format == "$#,##0.00"
    assert workbook["Annual Report Data"]["E2"].value == 129299893.93
    assert workbook["Annual Report Data"]["E2"].number_format == "$#,##0.00"
    assert attachment["filename"].startswith("credit-market-report-")
    assert attachment["encoding"] == "base64"


def test_workbook_formats_annual_year_and_quarter_effective_dates():
    report = MetabaseDashboardReport(
        dashboard_name="Credit Market",
        dashboard_description=None,
        dashboard_url="https://metabase.example/dashboard/1",
        tables=[
            MetabaseTable(
                name="Annual Market Report",
                columns=["Calculated Effective date: Year", "Transfers (#)"],
                rows=[["2026-06-01T00:00:00-07:00", 89]],
            ),
            MetabaseTable(
                name="Quarterly Market Report",
                columns=["Calculated Effective date: Quarter", "Transfers (#)"],
                rows=[
                    ["2026-01-01T00:00:00-08:00", 64],
                    ["2026-04-01T00:00:00-07:00", 25],
                ],
            ),
        ],
    )

    attachment = CreditMarketReportBuilder().build_attachment(report)
    workbook = load_workbook(io.BytesIO(base64.b64decode(attachment["content"])))

    assert workbook["Annual Market Report"]["A2"].value == 2026
    assert workbook["Quarterly Market Report"]["A2"].value == "Q1, 2026"
    assert workbook["Quarterly Market Report"]["A3"].value == "Q2, 2026"


def test_builds_dashboard_email_context():
    report = MetabaseDashboardReport(
        dashboard_name="Credit Market",
        dashboard_description=None,
        dashboard_url="https://metabase.example/dashboard/1",
        tables=[
            MetabaseTable(
                name="Transfers (Number)",
                columns=["value", "month", "change"],
                rows=[[6, "Jun 2026", "50% vs. Jun 2025: 4"]],
                display="scalar",
            ),
            MetabaseTable(
                name="Monthly Price (CA$ per credit)",
                columns=[
                    "Month",
                    "Min of Price Per Unit (CA$)",
                    "Max of Price Per Unit (CA$)",
                    "Weighted average price (CA$)",
                ],
                rows=[["2026-06-01T00:00:00-07:00", 108.50, 169.92, 143.32]],
            ),
            MetabaseTable(
                name="All Time",
                columns=[
                    "Volume (Credits)",
                    "Transfers (#)",
                    "Min of Price Per Unit (CA$)",
                    "Sum of Transfer Value (CA$)",
                ],
                rows=[[7823273, 847, 20, 2625743719.68]],
            ),
            MetabaseTable(
                name="Credit Transfer Price Trend",
                columns=[
                    "Transfer Date",
                    "Min of Price Per Unit",
                    "Max of Price Per Unit",
                    "Weighted average",
                ],
                rows=[
                    ["2023", "CA$200.00", "CA$510.00", "CA$471.86"],
                    ["2024", "CA$205.00", "CA$519.19", "CA$430.00"],
                    ["2025", "CA$90.00", "CA$495.00", "CA$259.50"],
                ],
            ),
            MetabaseTable(
                name="Credit Trade Volume",
                columns=["Month/Year", "Sum of Credit Quantity"],
                rows=[
                    ["January 2025", "456.4k"],
                    ["February 2025", "122.8k"],
                    ["March 2025", "314.8k"],
                ],
            ),
        ],
    )

    context = CreditMarketReportBuilder().build_email_context(report)

    assert context["metric_cards"][0]["name"] == "Transfers (Number)"
    assert context["metric_cards"][0]["primary"] == "6"
    assert context["summary_cards"][0]["name"] == "Monthly Price (CA$ per credit)"
    assert context["summary_cards"][0]["columns"] == ["Metric", "Value"]
    assert context["summary_cards"][0]["rows"][0] == ["Month", "Jun 2026"]
    assert context["summary_cards"][0]["rows"][1] == [
        "Min of Price Per Unit (CA$)",
        "CA$108.50",
    ]
    assert context["summary_cards"][1]["rows"][2] == [
        "Min of Price Per Unit (CA$)",
        "CA$20.00",
    ]
    assert context["summary_cards"][1]["rows"][3] == [
        "Sum of Transfer Value (CA$)",
        "CA$2,625,743,719.68",
    ]
    assert context["chart_sections"][0]["name"] == "Credit Transfer Price Trend"
    assert context["chart_sections"][0]["image_data_uri"].startswith(
        "data:image/png;base64,"
    )
    assert context["chart_sections"][1]["name"] == "Credit Trade Volume"
    assert context["chart_sections"][1]["image_data_uri"].startswith(
        "data:image/png;base64,"
    )


def test_report_column_display_names_are_normalized():
    report = MetabaseDashboardReport(
        dashboard_name="Credit Market",
        dashboard_description=None,
        dashboard_url="https://metabase.example/dashboard/1",
        tables=[
            MetabaseTable(
                name="Credit Trade Volume",
                columns=["Month/Year", "Sum of Quantity", "Count"],
                rows=[
                    ["January 2026", 75491, 6],
                    ["February 2026", 177020, 12],
                ],
            ),
            MetabaseTable(
                name="Monthly Market Report",
                columns=["Month", "Sum of Quantity", "Count"],
                rows=[["January 2026", 75491, 6]],
            ),
        ],
    )

    builder = CreditMarketReportBuilder()
    context = builder.build_email_context(report)
    workbook = load_workbook(
        io.BytesIO(base64.b64decode(builder.build_attachment(report)["content"]))
    )

    assert context["chart_sections"][0]["columns"] == [
        "Month/Year",
        "Volume (Credits)",
        "Transfers",
    ]
    assert workbook["Monthly Market Report"]["B1"].value == "Transfers"
    assert workbook["Monthly Market Report"]["C1"].value == "Volume (Credits)"
    assert workbook["Monthly Market Report"]["B2"].value == 6
    assert workbook["Monthly Market Report"]["C2"].value == 75491


def test_dashboard_metric_cards_compare_current_month_to_same_month_last_year():
    report = MetabaseDashboardReport(
        dashboard_name="Credit Market",
        dashboard_description=None,
        dashboard_url="https://metabase.example/dashboard/1",
        tables=[
            MetabaseTable(
                name="Monthly Market Report",
                columns=[
                    "Month",
                    "Transfers (#)",
                    "Volume (Credits)",
                    "Weighted average price",
                    "Category A1 - Average Price",
                ],
                rows=[
                    ["2026-06-01T00:00:00-07:00", 6, 75491, "$143.32", "$120.30"],
                    ["May, 2026", 12, 177020, "$129.91", "$106.00"],
                    ["2025-06-01T00:00:00-07:00", 4, 122825, "$273.67", "$237.00"],
                ],
            ),
        ],
    )

    context = CreditMarketReportBuilder().build_email_context(report)

    assert context["metric_cards"] == [
        {
            "name": "Transfers (Number)",
            "primary": "6",
            "subtitle": "Jun 2026",
            "comparison": "50% vs. Jun 2025: 4",
            "direction": "up",
        },
        {
            "name": "Total Volume (Credits)",
            "primary": "75,491",
            "subtitle": "Jun 2026",
            "comparison": "38.54% vs. Jun 2025: 122,825",
            "direction": "down",
        },
        {
            "name": "Category A1 Transfers (weighted average)",
            "primary": "CA$120.30",
            "subtitle": "Jun 2026",
            "comparison": "49.24% vs. Jun 2025: CA$237.00",
            "direction": "down",
        },
    ]


def test_metabase_client_prefers_api_key_authentication():
    with patch("lcfs.services.metabase.client.settings") as mock_settings:
        mock_settings.metabase_base_url = "https://metabase.example"
        mock_settings.metabase_request_timeout_seconds = 30
        mock_settings.metabase_api_key = "test-key"
        mock_settings.metabase_session_token = ""
        mock_settings.metabase_username = ""
        mock_settings.metabase_password = ""

        client = MetabaseClient()
        client._authenticate()

    assert client.session.headers["X-API-Key"] == "test-key"


@pytest.mark.anyio
async def test_monthly_report_uses_notification_table_email_subscriptions(mock_app):
    report = MetabaseDashboardReport(
        dashboard_name="Credit Market",
        dashboard_description=None,
        dashboard_url="https://metabase.example/dashboard/1",
        tables=[
            MetabaseTable(name="Credits", columns=["Name"], rows=[["Org A"]]),
        ],
    )
    mock_session = AsyncMock()
    mock_app.state.db_session_factory.return_value.__aenter__.return_value = (
        mock_session
    )

    with patch("lcfs.services.jobs.jobs.MetabaseClient") as mock_client_class, patch(
        "lcfs.services.jobs.jobs.CHESEmailRepository"
    ) as mock_email_repo_class, patch(
        "lcfs.services.jobs.jobs.CHESEmailService"
    ) as mock_email_service_class:
        mock_client_class.return_value.fetch_credit_market_report.return_value = report
        mock_email_repo = MagicMock()
        mock_email_repo.get_subscribed_user_emails = AsyncMock(
            return_value=["one@example.com", "two@example.com"]
        )
        mock_email_repo_class.return_value = mock_email_repo
        mock_email_service = MagicMock()
        mock_email_service.send_credit_market_report_email = AsyncMock(
            return_value=True
        )
        mock_email_service_class.return_value = mock_email_service

        result = await send_monthly_credit_market_report(mock_app)

    assert result is True
    mock_email_repo.get_subscribed_user_emails.assert_awaited_once_with(
        "PUBLIC__CREDIT_MARKET_MONTHLY_REPORT"
    )
    mock_email_service.send_credit_market_report_email.assert_called_once()
    recipients = mock_email_service.send_credit_market_report_email.call_args.kwargs[
        "recipients"
    ]
    assert recipients == ["one@example.com", "two@example.com"]
    attachments = mock_email_service.send_credit_market_report_email.call_args.kwargs[
        "attachments"
    ]
    assert len(attachments) == 1


@pytest.mark.anyio
async def test_monthly_report_skips_when_no_email_subscriptions(mock_app):
    report = MetabaseDashboardReport(
        dashboard_name="Credit Market",
        dashboard_description=None,
        dashboard_url="https://metabase.example/dashboard/1",
        tables=[
            MetabaseTable(name="Credits", columns=["Name"], rows=[["Org A"]]),
        ],
    )
    mock_session = AsyncMock()
    mock_app.state.db_session_factory.return_value.__aenter__.return_value = (
        mock_session
    )

    with patch("lcfs.services.jobs.jobs.MetabaseClient") as mock_client_class, patch(
        "lcfs.services.jobs.jobs.CHESEmailRepository"
    ) as mock_email_repo_class, patch(
        "lcfs.services.jobs.jobs.CHESEmailService"
    ) as mock_email_service_class:
        mock_client_class.return_value.fetch_credit_market_report.return_value = report
        mock_email_repo = MagicMock()
        mock_email_repo.get_subscribed_user_emails = AsyncMock(return_value=[])
        mock_email_repo_class.return_value = mock_email_repo
        mock_email_service = MagicMock()
        mock_email_service.send_credit_market_report_email = AsyncMock(
            return_value=True
        )
        mock_email_service_class.return_value = mock_email_service

        result = await send_monthly_credit_market_report(mock_app)

    assert result is False
    mock_email_repo.get_subscribed_user_emails.assert_awaited_once_with(
        "PUBLIC__CREDIT_MARKET_MONTHLY_REPORT"
    )
    mock_email_service.send_credit_market_report_email.assert_not_called()
