"""
Comprehensive unit tests for ChargingEquipmentExporter (export.py).

Coverage areas:
  - Column definition contracts (CE_INDEX, CE_MANAGE, CE_EXPORT) — labels, order, count
  - _get_column_index / _get_column_letter helpers
  - _build_charging_site_formulas — keys and formula template correctness
  - load_charging_equipment_data — field mapping, null-safety
  - export_filtered — user scoping, row building, column selection, filename,
    pagination loop, empty results, null/missing ORM fields
  - _current_pacific_date — format contract
"""
import io
import re
from datetime import date, datetime
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest

from lcfs.db.models.compliance.ChargingEquipment import ChargingEquipment, PortsEnum
from lcfs.db.models.compliance.ChargingEquipmentStatus import ChargingEquipmentStatus
from lcfs.db.models.compliance.ChargingSite import ChargingSite
from lcfs.db.models.compliance.LevelOfEquipment import LevelOfEquipment
from lcfs.db.models.organization.Organization import Organization
from lcfs.db.models.user.UserProfile import UserProfile
from lcfs.utils.constants import FILE_MEDIA_TYPE
from lcfs.web.api.base import PaginationRequestSchema
from lcfs.web.api.charging_equipment.export import (
    CE_EXPORT_COLUMNS,
    CE_INDEX_EXPORT_COLUMNS,
    CE_MANAGE_EXPORT_COLUMNS,
    MANAGE_FSE_EXPORT_FILENAME,
    FSE_INDEX_EXPORT_FILENAME,
    FSE_FILTERED_EXPORT_SHEETNAME,
    ChargingEquipmentExporter,
)


# ---------------------------------------------------------------------------
# Test Factories
# ---------------------------------------------------------------------------

def _make_exporter(
    equipment_list: list = None,
    total_count: int = None,
) -> ChargingEquipmentExporter:
    """Return an exporter with a stubbed repo."""
    equipment_list = equipment_list or []
    total_count = total_count if total_count is not None else len(equipment_list)
    repo = AsyncMock()
    repo.get_charging_equipment_list.return_value = (equipment_list, total_count)
    repo.get_all_equipment_by_organization_id.return_value = equipment_list
    exporter = ChargingEquipmentExporter.__new__(ChargingEquipmentExporter)
    exporter.repo = repo
    return exporter


def _make_user(is_government: bool, org_id: int = 1) -> UserProfile:
    user = MagicMock(spec=UserProfile)
    user.is_government = is_government
    user.organization_id = None if is_government else org_id
    return user


def _make_site(
    org_name: str = "Supplier Co",
    allocating_org_name: str | None = "Allocating Org",
    site_code: str = "SITE1",
    site_name: str = "Test Site",
    latitude: float = 49.77,
    longitude: float = -123.42,
) -> ChargingSite:
    org = Organization(organization_id=1, name=org_name)
    site = ChargingSite(
        charging_site_id=1,
        organization_id=1,
        site_code=site_code,
        site_name=site_name,
        latitude=latitude,
        longitude=longitude,
    )
    site.organization = org
    site.allocating_organization_name = allocating_org_name
    return site


def _make_equipment(
    *,
    serial_number: str = "SN-001",
    manufacturer: str = "Tesla",
    model: str = "Supercharger V3",
    ports: PortsEnum = PortsEnum.DUAL_PORT,
    latitude: float = 49.77,
    longitude: float = -123.42,
    version: int = 1,
    equipment_number: str = "001",
    status_str: str = "Validated",
    level_name: str = "Level 2",
    intended_uses: list = None,
    intended_users: list = None,
    site: ChargingSite = None,
    create_date: datetime = datetime(2024, 1, 1),
    update_date: datetime = datetime(2024, 1, 2),
    notes: str = "",
) -> ChargingEquipment:
    status = ChargingEquipmentStatus(charging_equipment_status_id=1, status=status_str)
    level = LevelOfEquipment(level_of_equipment_id=1, name=level_name)

    eq = ChargingEquipment(
        charging_equipment_id=1,
        charging_site_id=1,
        status_id=1,
        equipment_number=equipment_number,
        serial_number=serial_number,
        manufacturer=manufacturer,
        model=model,
        level_of_equipment_id=1,
        ports=ports,
        latitude=latitude,
        longitude=longitude,
        version=version,
        notes=notes,
    )
    eq.charging_site = site or _make_site()
    eq.status = status
    eq.level_of_equipment = level
    eq.intended_uses = intended_uses or []
    eq.intended_users = intended_users or []
    eq.create_date = create_date
    eq.update_date = update_date
    return eq


def _labels(columns) -> list[str]:
    return [col.label for col in columns]


async def _read_xlsx(response) -> tuple[list[str], list[list]]:
    """Return (header_row, data_rows) from the first sheet of the streaming response.

    SpreadsheetBuilder pre-fills the sheet with up to 2000 empty rows, so we strip
    trailing rows where every cell is None to get only the meaningful data rows.
    """
    body = b""
    async for chunk in response.body_iterator:
        body += chunk
    wb = openpyxl.load_workbook(io.BytesIO(body))
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return [], []
    header = rows[0]
    data_rows = [r for r in rows[1:] if any(cell is not None for cell in r)]
    return header, data_rows


# ---------------------------------------------------------------------------
# 1. Column Definition Contracts
# ---------------------------------------------------------------------------

class TestColumnDefinitions:
    """The contract for each column list is locked — label names and ordering matter."""

    def test_index_columns_complete_set(self):
        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        expected = [
            "Status", "Site name", "Organization", "Allocating organization",
            "Registration #", "Version #", "Serial #", "Manufacturer", "Model",
            "Level of equipment", "Ports", "Intended use", "Intended users",
            "Latitude", "Longitude", "Created", "Last updated",
        ]
        assert labels == expected

    def test_manage_columns_complete_set(self):
        labels = _labels(CE_MANAGE_EXPORT_COLUMNS)
        expected = [
            "Status", "Site name", "Allocating organization",
            "Registration #", "Version #", "Serial #", "Manufacturer", "Model",
            "Level of equipment", "Ports", "Intended use", "Intended users",
            "Latitude", "Longitude", "Created", "Last updated",
        ]
        assert labels == expected

    def test_export_columns_complete_set(self):
        labels = _labels(CE_EXPORT_COLUMNS)
        expected = [
            "Charging Site", "Serial Number", "Manufacturer", "Model",
            "Level of Equipment", "Ports", "Intended Uses", "Intended Users",
            "Notes", "Latitude", "Longitude",
        ]
        assert labels == expected

    # ---- positional sanity checks for index columns ----

    def test_index_allocating_org_immediately_after_organization(self):
        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert labels.index("Allocating organization") == labels.index("Organization") + 1

    def test_index_allocating_org_before_registration(self):
        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert labels.index("Allocating organization") < labels.index("Registration #")

    # ---- positional sanity checks for manage columns ----

    def test_manage_allocating_org_immediately_after_site_name(self):
        labels = _labels(CE_MANAGE_EXPORT_COLUMNS)
        assert labels.index("Allocating organization") == labels.index("Site name") + 1

    def test_manage_no_organization_column(self):
        labels = _labels(CE_MANAGE_EXPORT_COLUMNS)
        assert "Organization" not in labels

    def test_index_has_organization_column(self):
        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert "Organization" in labels


# ---------------------------------------------------------------------------
# 2. _get_column_index / _get_column_letter
# ---------------------------------------------------------------------------

class TestColumnHelpers:
    def setup_method(self):
        self.exporter = _make_exporter()

    def test_get_column_index_known_label(self):
        # CE_EXPORT_COLUMNS: "Charging Site" is index 1
        assert self.exporter._get_column_index("Charging Site") == 1

    def test_get_column_index_last_label(self):
        # "Longitude" is the 11th column
        assert self.exporter._get_column_index("Longitude") == 11

    def test_get_column_index_unknown_raises(self):
        with pytest.raises(ValueError, match="Missing column label"):
            self.exporter._get_column_index("Nonexistent Column")

    def test_get_column_letter_a(self):
        assert self.exporter._get_column_letter("Charging Site") == "A"

    def test_get_column_letter_k(self):
        # Column 11 → "K"
        assert self.exporter._get_column_letter("Longitude") == "K"


# ---------------------------------------------------------------------------
# 3. _build_charging_site_formulas
# ---------------------------------------------------------------------------

class TestBuildChargingSiteFormulas:
    def setup_method(self):
        self.exporter = _make_exporter()

    def test_returns_two_entries(self):
        formulas = self.exporter._build_charging_site_formulas(5)
        assert len(formulas) == 2

    def test_keys_are_latitude_and_longitude_indices(self):
        formulas = self.exporter._build_charging_site_formulas(5)
        lat_idx = self.exporter._get_column_index("Latitude")
        lng_idx = self.exporter._get_column_index("Longitude")
        assert lat_idx in formulas
        assert lng_idx in formulas

    def test_formulas_contain_row_placeholder(self):
        formulas = self.exporter._build_charging_site_formulas(5)
        for formula in formulas.values():
            assert "{row}" in formula

    def test_lookup_range_uses_count_plus_one(self):
        formulas = self.exporter._build_charging_site_formulas(10)
        for formula in formulas.values():
            # lookup_end = 10 + 1 = 11
            assert "$G$11" in formula

    def test_latitude_formula_uses_column_6_lookup(self):
        formulas = self.exporter._build_charging_site_formulas(3)
        lat_idx = self.exporter._get_column_index("Latitude")
        assert ",6,FALSE" in formulas[lat_idx]

    def test_longitude_formula_uses_column_7_lookup(self):
        formulas = self.exporter._build_charging_site_formulas(3)
        lng_idx = self.exporter._get_column_index("Longitude")
        assert ",7,FALSE" in formulas[lng_idx]


# ---------------------------------------------------------------------------
# 4. load_charging_equipment_data
# ---------------------------------------------------------------------------

class TestLoadChargingEquipmentData:
    @pytest.mark.anyio
    async def test_maps_all_fields_correctly(self):
        use = MagicMock()
        use.type = "Commercial"
        user = MagicMock()
        user.type_name = "Fleet"

        eq = _make_equipment(
            serial_number="ABC123",
            manufacturer="ChargePoint",
            model="Express 250",
            ports=PortsEnum.SINGLE_PORT,
            latitude=48.5,
            longitude=-123.1,
            intended_uses=[use],
            intended_users=[user],
            notes="Some note",
        )
        eq.charging_site.site_name = "My Site"

        exporter = _make_exporter([eq])
        data = await exporter.load_charging_equipment_data(1)

        assert len(data) == 1
        row = data[0]
        assert row[0] == "My Site"        # Charging Site
        assert row[1] == "ABC123"         # Serial Number
        assert row[2] == "ChargePoint"    # Manufacturer
        assert row[3] == "Express 250"    # Model
        assert row[4] == "Level 2"        # Level of Equipment
        assert row[5] == "Single port"    # Ports
        assert row[6] == "Commercial"     # Intended Uses
        assert row[7] == "Fleet"          # Intended Users
        assert row[8] == "Some note"      # Notes
        assert row[9] == 48.5             # Latitude
        assert row[10] == -123.1          # Longitude

    @pytest.mark.anyio
    async def test_multiple_intended_uses_joined_with_comma(self):
        use1, use2 = MagicMock(), MagicMock()
        use1.type, use2.type = "Commercial", "Fleet"
        eq = _make_equipment(intended_uses=[use1, use2])

        exporter = _make_exporter([eq])
        data = await exporter.load_charging_equipment_data(1)
        assert data[0][6] == "Commercial, Fleet"

    @pytest.mark.anyio
    async def test_null_charging_site_gives_empty_site_name(self):
        eq = _make_equipment()
        eq.charging_site = None

        exporter = _make_exporter([eq])
        data = await exporter.load_charging_equipment_data(1)
        assert data[0][0] == ""

    @pytest.mark.anyio
    async def test_null_level_of_equipment_gives_empty_string(self):
        eq = _make_equipment()
        eq.level_of_equipment = None

        exporter = _make_exporter([eq])
        data = await exporter.load_charging_equipment_data(1)
        assert data[0][4] == ""

    @pytest.mark.anyio
    async def test_null_ports_gives_empty_string(self):
        eq = _make_equipment()
        eq.ports = None

        exporter = _make_exporter([eq])
        data = await exporter.load_charging_equipment_data(1)
        assert data[0][5] == ""

    @pytest.mark.anyio
    async def test_null_latitude_longitude_give_empty_string(self):
        eq = _make_equipment()
        eq.latitude = None
        eq.longitude = None

        exporter = _make_exporter([eq])
        data = await exporter.load_charging_equipment_data(1)
        assert data[0][9] == ""
        assert data[0][10] == ""

    @pytest.mark.anyio
    async def test_empty_intended_uses_and_users(self):
        eq = _make_equipment(intended_uses=[], intended_users=[])
        exporter = _make_exporter([eq])
        data = await exporter.load_charging_equipment_data(1)
        assert data[0][6] == ""
        assert data[0][7] == ""

    @pytest.mark.anyio
    async def test_returns_empty_list_when_no_equipment(self):
        exporter = _make_exporter([])
        data = await exporter.load_charging_equipment_data(1)
        assert data == []


# ---------------------------------------------------------------------------
# 5. export_filtered — core behaviour
# ---------------------------------------------------------------------------

class TestExportFilteredColumnSelection:
    @pytest.mark.anyio
    async def test_government_uses_index_columns(self):
        eq = _make_equipment()
        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        headers, _ = await _read_xlsx(response)

        assert headers == _labels(CE_INDEX_EXPORT_COLUMNS)

    @pytest.mark.anyio
    async def test_supplier_uses_manage_columns(self):
        eq = _make_equipment()
        exporter = _make_exporter([eq])
        user = _make_user(is_government=False)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        headers, _ = await _read_xlsx(response)

        assert headers == _labels(CE_MANAGE_EXPORT_COLUMNS)

    @pytest.mark.anyio
    async def test_government_filename_starts_with_fse_index(self):
        exporter = _make_exporter()
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        cd = response.headers.get("content-disposition", "")
        assert FSE_INDEX_EXPORT_FILENAME in cd
        assert ".xlsx" in cd

    @pytest.mark.anyio
    async def test_supplier_filename_starts_with_manage_fse(self):
        exporter = _make_exporter()
        user = _make_user(is_government=False)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        cd = response.headers.get("content-disposition", "")
        assert MANAGE_FSE_EXPORT_FILENAME in cd

    @pytest.mark.anyio
    async def test_filename_includes_date(self):
        exporter = _make_exporter()
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        with patch.object(
            ChargingEquipmentExporter, "_current_pacific_date", return_value="2025-06-15"
        ):
            response = await exporter.export_filtered(user=user, pagination=pagination)

        cd = response.headers.get("content-disposition", "")
        assert "2025-06-15" in cd

    @pytest.mark.anyio
    async def test_response_media_type_is_xlsx(self):
        exporter = _make_exporter()
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        assert response.media_type == FILE_MEDIA_TYPE["XLSX"].value

    @pytest.mark.anyio
    async def test_sheet_name_is_fse(self):
        eq = _make_equipment()
        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        body = b"".join([chunk async for chunk in response.body_iterator])
        wb = openpyxl.load_workbook(io.BytesIO(body))
        assert FSE_FILTERED_EXPORT_SHEETNAME in wb.sheetnames


class TestExportFilteredOrganizationScoping:
    @pytest.mark.anyio
    async def test_government_with_org_id_uses_provided_org_id(self):
        exporter = _make_exporter()
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        await exporter.export_filtered(user=user, pagination=pagination, organization_id=42)

        exporter.repo.get_charging_equipment_list.assert_called_once()
        call_org_id = exporter.repo.get_charging_equipment_list.call_args[0][0]
        assert call_org_id == 42

    @pytest.mark.anyio
    async def test_government_without_org_id_uses_none(self):
        exporter = _make_exporter()
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        await exporter.export_filtered(user=user, pagination=pagination, organization_id=None)

        call_org_id = exporter.repo.get_charging_equipment_list.call_args[0][0]
        assert call_org_id is None

    @pytest.mark.anyio
    async def test_supplier_always_uses_own_org_id(self):
        exporter = _make_exporter()
        user = _make_user(is_government=False, org_id=7)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        # Even if an org_id is passed, supplier's own org should be used
        await exporter.export_filtered(user=user, pagination=pagination, organization_id=99)

        call_org_id = exporter.repo.get_charging_equipment_list.call_args[0][0]
        assert call_org_id == 7

    @pytest.mark.anyio
    async def test_government_excludes_drafts(self):
        exporter = _make_exporter()
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        await exporter.export_filtered(user=user, pagination=pagination)

        call_kwargs = exporter.repo.get_charging_equipment_list.call_args.kwargs
        assert call_kwargs.get("exclude_draft") is True

    @pytest.mark.anyio
    async def test_supplier_does_not_exclude_drafts(self):
        exporter = _make_exporter()
        user = _make_user(is_government=False)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        await exporter.export_filtered(user=user, pagination=pagination)

        call_kwargs = exporter.repo.get_charging_equipment_list.call_args.kwargs
        assert call_kwargs.get("exclude_draft") is False


class TestExportFilteredRowBuilding:
    @pytest.mark.anyio
    async def test_government_row_maps_all_fields(self):
        use = MagicMock()
        use.type = "Commercial"
        end_user = MagicMock()
        end_user.type_name = "Fleet"
        eq = _make_equipment(
            status_str="Validated",
            manufacturer="Tesla",
            model="V3",
            serial_number="XY-99",
            level_name="Level 2",
            ports=PortsEnum.DUAL_PORT,
            latitude=49.77,
            longitude=-123.4,
            version=3,
            intended_uses=[use],
            intended_users=[end_user],
            create_date=datetime(2024, 3, 15),
            update_date=datetime(2024, 4, 20),
            site=_make_site(
                org_name="OrgA",
                allocating_org_name="OrgB",
                site_code="ABC01",
                site_name="Alpha Site",
            ),
        )

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)
        row = data_rows[0]

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert row[labels.index("Status")] == "Validated"
        assert row[labels.index("Site name")] == "Alpha Site"
        assert row[labels.index("Organization")] == "OrgA"
        assert row[labels.index("Allocating organization")] == "OrgB"
        assert row[labels.index("Serial #")] == "XY-99"
        assert row[labels.index("Manufacturer")] == "Tesla"
        assert row[labels.index("Model")] == "V3"
        assert row[labels.index("Level of equipment")] == "Level 2"
        assert row[labels.index("Ports")] == "Dual port"
        assert row[labels.index("Intended use")] == "Commercial"
        assert row[labels.index("Intended users")] == "Fleet"
        assert row[labels.index("Latitude")] == 49.77
        assert row[labels.index("Longitude")] == -123.4
        # openpyxl reads date cells back as datetime (midnight), not date
        created = row[labels.index("Created")]
        updated = row[labels.index("Last updated")]
        assert getattr(created, "date", lambda: created)() == date(2024, 3, 15)
        assert getattr(updated, "date", lambda: updated)() == date(2024, 4, 20)

    @pytest.mark.anyio
    async def test_supplier_row_does_not_include_organization_field(self):
        eq = _make_equipment()
        exporter = _make_exporter([eq])
        user = _make_user(is_government=False)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        headers, data_rows = await _read_xlsx(response)

        assert "Organization" not in headers
        assert len(data_rows[0]) == len(CE_MANAGE_EXPORT_COLUMNS)

    @pytest.mark.anyio
    async def test_supplier_row_allocating_org_after_site_name(self):
        eq = _make_equipment(
            site=_make_site(allocating_org_name="FortisBC", site_name="My Site")
        )
        exporter = _make_exporter([eq])
        user = _make_user(is_government=False)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_MANAGE_EXPORT_COLUMNS)
        row = data_rows[0]
        assert row[labels.index("Site name")] == "My Site"
        assert row[labels.index("Allocating organization")] == "FortisBC"

    @pytest.mark.anyio
    async def test_null_allocating_org_writes_empty(self):
        eq = _make_equipment(
            site=_make_site(allocating_org_name=None)
        )
        exporter = _make_exporter([eq])

        for is_gov in (True, False):
            user = _make_user(is_government=is_gov)
            pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])
            response = await exporter.export_filtered(user=user, pagination=pagination)
            _, data_rows = await _read_xlsx(response)

            cols = CE_INDEX_EXPORT_COLUMNS if is_gov else CE_MANAGE_EXPORT_COLUMNS
            alloc_idx = _labels(cols).index("Allocating organization")
            assert data_rows[0][alloc_idx] in (None, "")

    @pytest.mark.anyio
    async def test_null_charging_site_gives_empty_site_and_org_fields(self):
        eq = _make_equipment()
        eq.charging_site = None

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        row = data_rows[0]
        assert row[labels.index("Site name")] in (None, "")
        assert row[labels.index("Organization")] in (None, "")
        assert row[labels.index("Allocating organization")] in (None, "")

    @pytest.mark.anyio
    async def test_null_status_gives_empty_status_field(self):
        eq = _make_equipment()
        eq.status = None

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert data_rows[0][labels.index("Status")] in (None, "")

    @pytest.mark.anyio
    async def test_null_level_gives_empty_level_field(self):
        eq = _make_equipment()
        eq.level_of_equipment = None

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert data_rows[0][labels.index("Level of equipment")] in (None, "")

    @pytest.mark.anyio
    async def test_null_ports_gives_empty_ports_field(self):
        eq = _make_equipment()
        eq.ports = None

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert data_rows[0][labels.index("Ports")] in (None, "")

    @pytest.mark.anyio
    async def test_null_latitude_longitude_give_empty_values(self):
        eq = _make_equipment()
        eq.latitude = None
        eq.longitude = None

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert data_rows[0][labels.index("Latitude")] in (None, "")
        assert data_rows[0][labels.index("Longitude")] in (None, "")

    @pytest.mark.anyio
    async def test_null_dates_write_empty(self):
        eq = _make_equipment(create_date=None, update_date=None)
        eq.create_date = None
        eq.update_date = None

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert data_rows[0][labels.index("Created")] is None
        assert data_rows[0][labels.index("Last updated")] is None

    @pytest.mark.anyio
    async def test_multiple_intended_uses_joined_with_comma(self):
        u1, u2 = MagicMock(), MagicMock()
        u1.type, u2.type = "Commercial", "Fleet"
        eq = _make_equipment(intended_uses=[u1, u2])

        exporter = _make_exporter([eq])
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        labels = _labels(CE_INDEX_EXPORT_COLUMNS)
        assert data_rows[0][labels.index("Intended use")] == "Commercial, Fleet"

    @pytest.mark.anyio
    async def test_empty_result_set_writes_header_only(self):
        exporter = _make_exporter(equipment_list=[], total_count=0)
        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        headers, data_rows = await _read_xlsx(response)

        assert headers == _labels(CE_INDEX_EXPORT_COLUMNS)
        assert data_rows == []


class TestExportFilteredPagination:
    @pytest.mark.anyio
    async def test_fetches_all_pages_until_total_satisfied(self):
        """When total_count > first page size, the loop should page through all results."""
        page1 = [_make_equipment(serial_number="SN-1")]
        page2 = [_make_equipment(serial_number="SN-2")]

        repo = AsyncMock()
        repo.get_charging_equipment_list.side_effect = [
            (page1, 2),  # page 1: 1 row, total 2
            (page2, 2),  # page 2: 1 row, total 2 — all collected now
            ([], 2),     # safety: should not reach here
        ]

        exporter = ChargingEquipmentExporter.__new__(ChargingEquipmentExporter)
        exporter.repo = repo

        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        assert len(data_rows) == 2
        assert repo.get_charging_equipment_list.call_count == 2

    @pytest.mark.anyio
    async def test_stops_when_empty_page_returned(self):
        """If the repo returns an empty list before total is met, the loop must stop."""
        repo = AsyncMock()
        repo.get_charging_equipment_list.side_effect = [
            ([_make_equipment()], 999),  # large total, but next page is empty
            ([], 999),
        ]

        exporter = ChargingEquipmentExporter.__new__(ChargingEquipmentExporter)
        exporter.repo = repo

        user = _make_user(is_government=True)
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        response = await exporter.export_filtered(user=user, pagination=pagination)
        _, data_rows = await _read_xlsx(response)

        assert len(data_rows) == 1
        assert repo.get_charging_equipment_list.call_count == 2

    @pytest.mark.anyio
    async def test_pagination_uses_page_size_1000(self):
        exporter = _make_exporter()
        user = _make_user(is_government=True)
        # Pass a different size; exporter should override to 1000
        pagination = PaginationRequestSchema(page=1, size=25, sort_orders=[])

        await exporter.export_filtered(user=user, pagination=pagination)

        # Second positional arg to get_charging_equipment_list is the pagination object
        call_pagination = exporter.repo.get_charging_equipment_list.call_args.args[1]
        assert call_pagination.size == 1000


# ---------------------------------------------------------------------------
# 6. _current_pacific_date
# ---------------------------------------------------------------------------

class TestCurrentPacificDate:
    def test_returns_yyyy_mm_dd_format(self):
        result = ChargingEquipmentExporter._current_pacific_date()
        assert re.fullmatch(r"\d{4}-\d{2}-\d{2}", result), (
            f"Expected YYYY-MM-DD, got {result!r}"
        )

    def test_returns_string(self):
        assert isinstance(ChargingEquipmentExporter._current_pacific_date(), str)
