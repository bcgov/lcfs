"""
Unit tests for the FSE bulk-update importer.

Covers:
  - FSEReportingImporter.import_data  (job dispatch, MIME/size validation)
  - FSEReportingImporter.get_status   (Redis read including `skipped` field)
  - _parse_date                       (all recognised formats, sentinel types, errors)
  - _load_sheet                       (sheet name validation)
  - Row-processing logic:
      col layout: [Site name, Registration #, From, To, kWh, Notes]
      blank editable fields + existing record → deactivate called
      blank editable fields + no record      → skipped
      note only (no dates/kWh) + existing    → activate=True, kwh defaults to 0
      valid full row                         → updated=1, activate=True
      missing reg number                     → rejected
      invalid kWh                            → rejected
      negative kWh                           → rejected
      invalid date                           → rejected
      inverted date range                    → rejected
      registration not found in org          → rejected
"""

import asyncio
import concurrent.futures
import datetime
import io
import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import HTTPException
from openpyxl import Workbook
from redis.asyncio import Redis

from lcfs.services.clamav.client import ClamAVService
from lcfs.web.api.compliance_report.repo import ComplianceReportRepository
from lcfs.web.api.final_supply_equipment.fse_reporting_importer import (
    FSE_UPDATE_SHEETNAME,
    FSEReportingImporter,
    _parse_date,
    _load_sheet,
)
from lcfs.web.api.final_supply_equipment.repo import FinalSupplyEquipmentRepository


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def mock_fse_repo():
    repo = MagicMock(spec=FinalSupplyEquipmentRepository)
    repo.get_charging_equipment_by_registration_number = AsyncMock(return_value=None)
    repo.get_fse_reporting_record_for_group = AsyncMock(return_value=None)
    repo.bulk_update_fse_reporting_record = AsyncMock(return_value=None)
    return repo


@pytest.fixture
def mock_compliance_repo():
    repo = MagicMock(spec=ComplianceReportRepository)
    report = MagicMock()
    report.compliance_report_group_uuid = "group-uuid-123"
    report.organization_id = 42
    repo.get_compliance_report_by_id = AsyncMock(return_value=report)
    return repo


@pytest.fixture
def mock_redis():
    client = MagicMock(spec=Redis)
    client.set = AsyncMock()
    client.get = AsyncMock(return_value=None)
    return client


@pytest.fixture
def mock_executor():
    executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
    yield executor
    executor.shutdown(wait=False)


@pytest.fixture
def importer(mock_fse_repo, mock_compliance_repo, mock_redis, mock_executor):
    return FSEReportingImporter(
        repo=mock_fse_repo,
        compliance_report_repo=mock_compliance_repo,
        clamav_service=MagicMock(spec=ClamAVService),
        redis_client=mock_redis,
        executor=mock_executor,
    )


def _xlsx_file(filename: str = "update.xlsx"):
    """Return a minimal UploadFile mock carrying valid XLSX MIME."""
    mock = MagicMock()
    mock.filename = filename
    mock.content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    mock.read = AsyncMock(return_value=b"fake-excel-bytes")
    return mock


# ---------------------------------------------------------------------------
# FSEReportingImporter.import_data
# ---------------------------------------------------------------------------


@pytest.mark.anyio
async def test_import_data_returns_job_id(importer, mock_redis):
    """Successful dispatch → a non-empty UUID string is returned."""
    user = MagicMock()
    user.keycloak_username = "testuser"

    with patch(
        "lcfs.web.api.final_supply_equipment.fse_reporting_importer._import_async",
        new=AsyncMock(return_value=None),
    ):
        job_id = await importer.import_data(
            compliance_report_id=1,
            user=user,
            file=_xlsx_file(),
        )

    assert isinstance(job_id, str)
    assert len(job_id) == 36  # standard UUID format
    mock_redis.set.assert_called()


@pytest.mark.anyio
async def test_import_data_report_not_found_raises(importer, mock_compliance_repo):
    """DataNotFoundException when compliance report does not exist."""
    from lcfs.web.exception.exceptions import DataNotFoundException

    mock_compliance_repo.get_compliance_report_by_id = AsyncMock(return_value=None)

    with pytest.raises(DataNotFoundException):
        await importer.import_data(
            compliance_report_id=999,
            user=MagicMock(),
            file=_xlsx_file(),
        )


@pytest.mark.anyio
async def test_import_data_invalid_mime_raises(importer):
    """HTTPException 400 for disallowed MIME type."""
    bad_file = _xlsx_file()
    bad_file.content_type = "text/html"

    with patch(
        "lcfs.web.api.final_supply_equipment.fse_reporting_importer._import_async",
        new=AsyncMock(return_value=None),
    ):
        with pytest.raises(HTTPException) as exc:
            await importer.import_data(
                compliance_report_id=1,
                user=MagicMock(),
                file=bad_file,
            )

    assert exc.value.status_code == 400
    assert "not allowed" in str(exc.value.detail)


@pytest.mark.anyio
async def test_import_data_file_too_large_raises(importer):
    """HTTPException 400 when file exceeds the size limit."""
    big_file = _xlsx_file()
    big_file.read = AsyncMock(return_value=b"x" * (51 * 1024 * 1024))

    with patch(
        "lcfs.web.api.final_supply_equipment.fse_reporting_importer._import_async",
        new=AsyncMock(return_value=None),
    ):
        with pytest.raises(HTTPException) as exc:
            await importer.import_data(
                compliance_report_id=1,
                user=MagicMock(),
                file=big_file,
            )

    assert exc.value.status_code == 400
    assert "exceeds the maximum limit" in str(exc.value.detail)


# ---------------------------------------------------------------------------
# FSEReportingImporter.get_status
# ---------------------------------------------------------------------------


@pytest.mark.anyio
async def test_get_status_no_job(importer, mock_redis):
    mock_redis.get = AsyncMock(return_value=None)
    result = await importer.get_status("nonexistent")
    assert result["progress"] == 0
    assert "No job found" in result["status"]


@pytest.mark.anyio
async def test_get_status_invalid_json(importer, mock_redis):
    mock_redis.get = AsyncMock(return_value=b"not-json{{")
    result = await importer.get_status("bad-id")
    assert result["progress"] == 0
    assert "Invalid status data found." in result["status"]


@pytest.mark.anyio
async def test_get_status_success_includes_skipped(importer, mock_redis):
    """get_status must surface the `skipped` counter."""
    payload = {
        "progress": 100,
        "status": "Import process completed.",
        "created": 8,
        "rejected": 2,
        "skipped": 3,
        "errors": ["Row 5: kWh blank"],
    }
    mock_redis.get = AsyncMock(return_value=json.dumps(payload))

    result = await importer.get_status("job-abc")

    assert result["progress"] == 100
    assert result["created"] == 8
    assert result["rejected"] == 2
    assert result["skipped"] == 3
    assert len(result["errors"]) == 1


@pytest.mark.anyio
async def test_get_status_skipped_defaults_to_zero(importer, mock_redis):
    """Older payloads that lack `skipped` should default to 0."""
    payload = {
        "progress": 50,
        "status": "Running...",
        "created": 5,
        "rejected": 0,
        "errors": [],
    }
    mock_redis.get = AsyncMock(return_value=json.dumps(payload))

    result = await importer.get_status("job-old")
    assert result["skipped"] == 0


# ---------------------------------------------------------------------------
# _parse_date helper
# ---------------------------------------------------------------------------


def test_parse_date_from_datetime_object():
    dt = datetime.datetime(2024, 6, 15, 12, 0, 0)
    assert _parse_date(dt) == datetime.date(2024, 6, 15)


def test_parse_date_from_date_object():
    d = datetime.date(2024, 3, 1)
    assert _parse_date(d) == d


def test_parse_date_iso_string():
    assert _parse_date("2024-01-31") == datetime.date(2024, 1, 31)


def test_parse_date_us_slash_format():
    assert _parse_date("01/31/2024") == datetime.date(2024, 1, 31)


def test_parse_date_day_first_slash_format():
    assert _parse_date("31/01/2024") == datetime.date(2024, 1, 31)


def test_parse_date_none_raises():
    with pytest.raises((ValueError, TypeError)):
        _parse_date(None)


def test_parse_date_garbage_string_raises():
    with pytest.raises((ValueError, TypeError)):
        _parse_date("not-a-date")


def test_parse_date_partial_string_raises():
    with pytest.raises((ValueError, TypeError)):
        _parse_date("2024-99-99")


# ---------------------------------------------------------------------------
# _load_sheet helper
# ---------------------------------------------------------------------------


def _build_upload_file(sheetname: str) -> MagicMock:
    """Create an in-memory XLSX file with the given sheet name."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    # New column layout: Site name | Registration # | From | To | kWh | Notes
    ws.append(["Site name", "Registration #", "Dates of supply from",
               "Dates of supply to", "kWh usage", "Compliance notes"])
    ws.append(["Site A", "ORG-001-001", "2024-01-01", "2024-12-31", 500, "ok"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    mock_file = MagicMock()
    mock_file.file = buf
    return mock_file


def test_load_sheet_correct_name():
    file = _build_upload_file(FSE_UPDATE_SHEETNAME)
    sheet = _load_sheet(file)
    assert sheet.title == FSE_UPDATE_SHEETNAME
    assert sheet.max_row >= 2


def test_load_sheet_wrong_name_raises():
    file = _build_upload_file("WrongSheet")
    with pytest.raises(Exception, match=FSE_UPDATE_SHEETNAME):
        _load_sheet(file)


# ---------------------------------------------------------------------------
# Row-processing logic (tested via _import_async with mocked DB calls)
# Column layout: [Site name, Registration #, From, To, kWh, Notes]
# ---------------------------------------------------------------------------


async def _run_import(rows: list, fse_repo_override=None):
    """
    Build a minimal workbook, run _import_async against it, and return the
    final Redis progress payload.
    """
    from lcfs.web.api.final_supply_equipment.fse_reporting_importer import (
        _import_async,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = FSE_UPDATE_SHEETNAME
    # Header row must match new column layout
    ws.append(["Site name", "Registration #", "Dates of supply from",
               "Dates of supply to", "kWh usage", "Compliance notes"])
    for row in rows:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    upload_file = MagicMock()
    upload_file.file = buf

    progress_records = []

    async def fake_update(redis_client, job_id, progress, status_msg,
                          updated=0, skipped=0, rejected=0, errors=None):
        progress_records.append({
            "progress": progress,
            "status": status_msg,
            "updated": updated,
            "skipped": skipped,
            "rejected": rejected,
            "errors": errors or [],
        })

    if fse_repo_override is None:
        fse_repo = MagicMock(spec=FinalSupplyEquipmentRepository)
        fse_repo.get_charging_equipment_by_registration_number = AsyncMock(
            return_value=None
        )
        fse_repo.get_fse_reporting_record_for_group = AsyncMock(return_value=None)
        fse_repo.bulk_update_fse_reporting_record = AsyncMock(return_value=None)
    else:
        fse_repo = fse_repo_override

    engine = MagicMock()
    session = AsyncMock()
    session.__aenter__ = AsyncMock(return_value=session)
    session.__aexit__ = AsyncMock(return_value=False)
    session.begin = MagicMock(return_value=session)

    redis = MagicMock(spec=Redis)
    redis.set = AsyncMock()

    with (
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer.create_async_engine",
            return_value=engine,
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer.AsyncSession",
            return_value=session,
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer.Redis",
            return_value=redis,
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer.FuelSupplyRepository",
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer.FinalSupplyEquipmentRepository",
            return_value=fse_repo,
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer.ComplianceReportRepository",
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer.set_user_context",
            new=AsyncMock(),
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer._update_progress",
            new=fake_update,
        ),
        patch(
            "lcfs.web.api.final_supply_equipment.fse_reporting_importer._load_sheet",
            return_value=wb[FSE_UPDATE_SHEETNAME],
        ),
    ):
        engine.dispose = AsyncMock()
        user = MagicMock()
        user.keycloak_username = "testuser"

        await _import_async(
            compliance_report_id=1,
            user=user,
            compliance_report_group_uuid="group-uuid",
            report_organization_id=42,
            file=upload_file,
            job_id="test-job",
        )

    return progress_records


# col layout: [site_name, reg_num, serial#, from, to, kwh, notes]

@pytest.mark.anyio
async def test_row_missing_registration_is_rejected():
    """Empty registration number → rejected."""
    rows = [["Site A", None, "SN-1", "2024-01-01", "2024-12-31", 500, "note"]]
    records = await _run_import(rows)
    final = records[-1]
    assert final["rejected"] == 1
    assert final["updated"] == 0


@pytest.mark.anyio
async def test_row_invalid_kwh_is_rejected():
    """Non-numeric kWh value → rejected."""
    rows = [["Site A", "REG-001", "SN-1", "2024-01-01", "2024-12-31", "NOT_A_NUMBER", "note"]]
    records = await _run_import(rows)
    final = records[-1]
    assert final["rejected"] == 1
    assert final["updated"] == 0


@pytest.mark.anyio
async def test_row_negative_kwh_is_rejected():
    """Negative kWh value → rejected."""
    rows = [["Site A", "REG-001", "SN-1", "2024-01-01", "2024-12-31", -100, "note"]]
    records = await _run_import(rows)
    final = records[-1]
    assert final["rejected"] == 1


@pytest.mark.anyio
async def test_row_invalid_date_is_rejected():
    """Unparseable 'from' date → rejected."""
    rows = [["Site A", "REG-001", "SN-1", "not-a-date", "2024-12-31", 500, "note"]]
    records = await _run_import(rows)
    final = records[-1]
    assert final["rejected"] == 1


@pytest.mark.anyio
async def test_row_inverted_date_range_is_rejected():
    """supply_from > supply_to → rejected."""
    rows = [["Site A", "REG-001", "SN-1", "2024-12-31", "2024-01-01", 500, "note"]]
    records = await _run_import(rows)
    final = records[-1]
    assert final["rejected"] == 1


@pytest.mark.anyio
async def test_row_registration_not_found_is_rejected():
    """Registration number not found in org → rejected."""
    fse_repo = MagicMock(spec=FinalSupplyEquipmentRepository)
    fse_repo.get_charging_equipment_by_registration_number = AsyncMock(
        return_value=None
    )
    rows = [["Site A", "UNKNOWN-REG", "SN-1", "2024-01-01", "2024-12-31", 500, "note"]]
    records = await _run_import(rows, fse_repo_override=fse_repo)
    final = records[-1]
    assert final["rejected"] == 1


@pytest.mark.anyio
async def test_valid_row_existing_record_is_updated_and_activated():
    """Valid row with existing CRCE record → updated=1, activate=True called."""
    existing = MagicMock()
    existing.charging_equipment_compliance_id = 99

    equipment = MagicMock()
    equipment.charging_equipment_id = 1
    equipment.charging_equipment_version = 0

    fse_repo = MagicMock(spec=FinalSupplyEquipmentRepository)
    fse_repo.get_charging_equipment_by_registration_number = AsyncMock(
        return_value=equipment
    )
    fse_repo.get_fse_reporting_record_for_group = AsyncMock(return_value=existing)
    fse_repo.bulk_update_fse_reporting_record = AsyncMock(return_value=None)

    rows = [["Site A", "ORG-001-001", "SN-1", "2024-01-01", "2024-12-31", 500, "good note"]]
    records = await _run_import(rows, fse_repo_override=fse_repo)
    final = records[-1]

    assert final["updated"] == 1
    assert final["skipped"] == 0
    assert final["rejected"] == 0
    fse_repo.bulk_update_fse_reporting_record.assert_called_once_with(
        charging_equipment_compliance_id=99,
        supply_from_date=datetime.date(2024, 1, 1),
        supply_to_date=datetime.date(2024, 12, 31),
        kwh_usage=500.0,
        compliance_notes="good note",
        activate=True,
    )


@pytest.mark.anyio
async def test_row_all_editable_empty_with_existing_record_calls_deactivate():
    """
    Row with only site name + reg # (all editable fields blank) and an existing
    active record → deactivate=True must be called on bulk_update.
    """
    existing = MagicMock()
    existing.charging_equipment_compliance_id = 77
    existing.is_active = True

    equipment = MagicMock()
    equipment.charging_equipment_id = 5
    equipment.charging_equipment_version = 0

    fse_repo = MagicMock(spec=FinalSupplyEquipmentRepository)
    fse_repo.get_charging_equipment_by_registration_number = AsyncMock(
        return_value=equipment
    )
    fse_repo.get_fse_reporting_record_for_group = AsyncMock(return_value=existing)
    fse_repo.bulk_update_fse_reporting_record = AsyncMock(return_value=None)

    # All editable columns blank: no from, no to, no kWh, no notes
    rows = [["Site A", "ORG-001-001", "SN-1", None, None, None, None]]
    records = await _run_import(rows, fse_repo_override=fse_repo)

    fse_repo.bulk_update_fse_reporting_record.assert_called_once()
    call_kwargs = fse_repo.bulk_update_fse_reporting_record.call_args.kwargs
    assert call_kwargs.get("deactivate") is True
    assert call_kwargs.get("charging_equipment_compliance_id") == 77


@pytest.mark.anyio
async def test_row_all_editable_empty_no_existing_record_is_skipped():
    """
    Row with only site name + reg # (all editable fields blank) and NO existing
    record → skipped, no DB call.
    """
    equipment = MagicMock()
    equipment.charging_equipment_id = 5
    equipment.charging_equipment_version = 0

    fse_repo = MagicMock(spec=FinalSupplyEquipmentRepository)
    fse_repo.get_charging_equipment_by_registration_number = AsyncMock(
        return_value=equipment
    )
    fse_repo.get_fse_reporting_record_for_group = AsyncMock(return_value=None)
    fse_repo.bulk_update_fse_reporting_record = AsyncMock(return_value=None)

    rows = [["Site A", "ORG-001-001", "SN-1", None, None, None, None]]
    records = await _run_import(rows, fse_repo_override=fse_repo)
    final = records[-1]

    assert final["skipped"] == 1
    fse_repo.bulk_update_fse_reporting_record.assert_not_called()


@pytest.mark.anyio
async def test_row_note_only_activates_row_with_zero_kwh():
    """
    Row with only a compliance note (no dates, no kWh) → treated as has_any_data,
    row is activated and kwh defaults to 0.
    """
    existing = MagicMock()
    existing.charging_equipment_compliance_id = 55

    equipment = MagicMock()
    equipment.charging_equipment_id = 3
    equipment.charging_equipment_version = 0

    fse_repo = MagicMock(spec=FinalSupplyEquipmentRepository)
    fse_repo.get_charging_equipment_by_registration_number = AsyncMock(
        return_value=equipment
    )
    fse_repo.get_fse_reporting_record_for_group = AsyncMock(return_value=existing)
    fse_repo.bulk_update_fse_reporting_record = AsyncMock(return_value=None)

    rows = [["Site A", "ORG-001-001", "SN-1", None, None, None, "Just a note"]]
    records = await _run_import(rows, fse_repo_override=fse_repo)
    final = records[-1]

    assert final["updated"] == 1
    call_kwargs = fse_repo.bulk_update_fse_reporting_record.call_args.kwargs
    assert call_kwargs.get("activate") is True
    assert call_kwargs.get("kwh_usage") == 0
    assert call_kwargs.get("compliance_notes") == "Just a note"


@pytest.mark.anyio
async def test_fully_blank_row_is_silently_skipped():
    """Rows where every cell is None are silently ignored (not counted)."""
    rows = [[None, None, None, None, None, None, None]]
    records = await _run_import(rows)
    final = records[-1]
    assert final["updated"] == 0
    assert final["skipped"] == 0
    assert final["rejected"] == 0


@pytest.mark.anyio
async def test_mixed_rows_counters_are_accurate():
    """A mix of valid, deactivated, and rejected rows → correct counters."""
    equipment = MagicMock()
    equipment.charging_equipment_id = 1
    equipment.charging_equipment_version = 0
    existing = MagicMock()
    existing.charging_equipment_compliance_id = 10
    existing.is_active = True

    fse_repo = MagicMock(spec=FinalSupplyEquipmentRepository)
    fse_repo.get_charging_equipment_by_registration_number = AsyncMock(
        return_value=equipment
    )
    fse_repo.get_fse_reporting_record_for_group = AsyncMock(return_value=existing)
    fse_repo.bulk_update_fse_reporting_record = AsyncMock(return_value=None)

    rows = [
        ["Site A", "ORG-001", "SN-1", "2024-01-01", "2024-12-31", 300, "ok"],     # updated
        ["Site A", "ORG-001", "SN-1", None, None, None, None],                     # deactivated (also counted as updated)
        [None, None, None, "2024-01-01", "2024-12-31", 100, "no reg"],             # rejected
    ]
    records = await _run_import(rows, fse_repo_override=fse_repo)
    final = records[-1]

    # 2 calls to bulk_update (activate + deactivate), 1 rejected
    assert final["updated"] == 2
    assert final["rejected"] == 1
