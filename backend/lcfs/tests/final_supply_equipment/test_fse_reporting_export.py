"""
Unit tests for the FSE bulk-update Excel template exporter.

Covers:
  - FSEReportingExporter.export_update_template (success, report not found)
  - _build_workbook: sheet name, headers, column widths, sheet protection,
    locked/unlocked cells, data validators, date number format,
    empty rows for new entries, site-name column (col A) locked
  - _load_fse_data: site_name included, datetime → date normalisation,
    inactive rows (is_active=False) show only site name + reg #,
    rows from another report group show only site name + reg #,
    kwh_usage None → 0 for active rows
"""

import datetime
import io
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook
from starlette.responses import StreamingResponse

from lcfs.db.models import Organization
from lcfs.utils.constants import FILE_MEDIA_TYPE
from lcfs.web.api.final_supply_equipment.fse_reporting_export import (
    COLUMN_WIDTHS,
    FSE_UPDATE_EXPORT_FILENAME,
    FSE_UPDATE_SHEETNAME,
    FSEReportingExporter,
    HEADERS,
)
from lcfs.web.exception.exceptions import DataNotFoundException


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _compliance_period(year: int = 2024):
    return MagicMock(
        description=str(year),
        effective_date=datetime.datetime(year, 1, 1),
        expiration_date=datetime.datetime(year, 12, 31),
    )


def _make_exporter(fse_rows=None, report_found=True, group_uuid="group-abc"):
    repo = MagicMock()
    repo.get_fse_for_bulk_update_template = AsyncMock(return_value=fse_rows or [])

    cr_repo = MagicMock()
    if report_found:
        report = MagicMock(
            compliance_period=_compliance_period(),
            compliance_report_group_uuid=group_uuid,
        )
        cr_repo.get_compliance_report_by_id = AsyncMock(return_value=report)
    else:
        cr_repo.get_compliance_report_by_id = AsyncMock(return_value=None)

    exp = FSEReportingExporter(repo=repo, compliance_report_repo=cr_repo)
    return exp


def _fse_row(
    registration="ORG-AAAA1A-001",
    site_name="Charge Site 1",
    serial_number="SN-12345",
    supply_from=datetime.date(2024, 1, 1),
    supply_to=datetime.date(2024, 12, 31),
    kwh_usage=1500.0,
    notes="Test note",
    is_active=True,
    compliance_report_group_uuid="group-abc",
):
    row = MagicMock()
    row.registration_number = registration
    row.site_name = site_name
    row.serial_number = serial_number
    row.supply_from_date = supply_from
    row.supply_to_date = supply_to
    row.kwh_usage = kwh_usage
    row.compliance_notes = notes
    row.is_active = is_active
    row.compliance_report_group_uuid = compliance_report_group_uuid
    return row


# ---------------------------------------------------------------------------
# export_update_template
# ---------------------------------------------------------------------------


@pytest.mark.anyio
async def test_export_returns_streaming_response():
    exporter = _make_exporter(fse_rows=[_fse_row()])
    org = Organization(name="TestOrg")
    org.organization_id = 1

    response = await exporter.export_update_template(
        compliance_report_id=1,
        user=MagicMock(),
        organization=org,
    )

    assert isinstance(response, StreamingResponse)
    assert response.status_code == 200
    assert response.media_type == FILE_MEDIA_TYPE["XLSX"].value


@pytest.mark.anyio
async def test_export_filename_includes_org_and_period():
    exporter = _make_exporter(fse_rows=[])
    org = Organization(name="MyOrg")
    org.organization_id = 1

    response = await exporter.export_update_template(
        compliance_report_id=1,
        user=MagicMock(),
        organization=org,
    )

    disposition = response.headers["Content-Disposition"]
    assert FSE_UPDATE_EXPORT_FILENAME in disposition
    assert "MyOrg" in disposition
    assert "2024" in disposition


@pytest.mark.anyio
async def test_export_report_not_found_raises():
    exporter = _make_exporter(report_found=False)
    org = Organization(name="AnyOrg")
    org.organization_id = 1

    with pytest.raises(DataNotFoundException):
        await exporter.export_update_template(
            compliance_report_id=999,
            user=MagicMock(),
            organization=org,
        )


# ---------------------------------------------------------------------------
# Header / column spec
# ---------------------------------------------------------------------------


def test_headers_first_column_is_site_name():
    """Site name must be the first column."""
    assert HEADERS[0] == "Site name"


def test_headers_second_column_is_registration():
    assert HEADERS[1] == "Registration #"


def test_headers_third_column_is_serial():
    assert HEADERS[2] == "Serial #"


def test_header_labels():
    """Spot-check all header labels."""
    assert HEADERS[3] == "Dates of supply from"
    assert HEADERS[4] == "Dates of supply to"
    assert HEADERS[5] == "kWh usage"
    assert HEADERS[6] == "Compliance notes"


def test_column_count_matches_headers():
    assert len(COLUMN_WIDTHS) == len(HEADERS)


# ---------------------------------------------------------------------------
# _build_workbook content
# ---------------------------------------------------------------------------


def _parse_workbook(exporter, rows=None):
    """Call _build_workbook and return the reloaded worksheet for inspection."""
    period = _compliance_period()
    wb = exporter._build_workbook(rows or [], period)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return load_workbook(out)[FSE_UPDATE_SHEETNAME]


def _sample_row():
    return [
        "Charge Site 1",
        "ORG-001",
        "SN-12345",
        datetime.date(2024, 1, 1),
        datetime.date(2024, 12, 31),
        500,
        "note",
    ]


def test_sheet_name_is_fse():
    ws = _parse_workbook(_make_exporter())
    assert ws.title == FSE_UPDATE_SHEETNAME


def test_headers_match_spec():
    ws = _parse_workbook(_make_exporter())
    actual = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    assert actual == HEADERS


def test_site_name_cell_is_locked():
    """Col A (site name) must be locked for existing data rows."""
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([_sample_row()], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    assert ws.cell(row=2, column=1).protection.locked is True


def test_registration_cell_is_locked():
    """Col B (registration #) must be locked for existing data rows."""
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([_sample_row()], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    assert ws.cell(row=2, column=2).protection.locked is True


def test_serial_cell_is_locked():
    """Col C (serial #) must be locked for existing data rows."""
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([_sample_row()], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    assert ws.cell(row=2, column=3).protection.locked is True


def test_editable_columns_are_unlocked():
    """Cols D–G (dates, kWh, notes) must be unlocked for existing data rows."""
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([_sample_row()], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    for col in (4, 5, 6, 7):
        assert ws.cell(row=2, column=col).protection.locked is False, (
            f"Column {col} should be unlocked"
        )


def test_empty_rows_for_new_entries_are_unlocked():
    """The 500 empty rows appended after data rows must be unlocked (except Serial #)."""
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([_sample_row()], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    # Data rows start at row 2; one data row → first empty row is row 3
    first_empty = 3
    for col in range(1, 8):
        cell = ws.cell(row=first_empty, column=col)
        if col == 3:
            # Serial # column is always locked
            assert cell.protection.locked is True, (
                "Empty row Serial # col should be locked"
            )
        else:
            assert cell.protection.locked is False, (
                f"Empty row col {col} should be unlocked"
            )


def test_empty_rows_date_columns_have_date_format():
    """Empty rows' date columns (D, E) must carry the date number format."""
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    # No data rows → first empty row is row 2
    assert ws.cell(row=2, column=4).number_format == "yyyy-mm-dd"
    assert ws.cell(row=2, column=5).number_format == "yyyy-mm-dd"


def test_sheet_protection_is_enabled():
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    assert ws.protection.sheet is True


def test_date_columns_have_date_number_format():
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([_sample_row()], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    # Col D = 4, Col E = 5
    assert ws.cell(row=2, column=4).number_format == "yyyy-mm-dd"
    assert ws.cell(row=2, column=5).number_format == "yyyy-mm-dd"


def test_kwh_column_has_numeric_format():
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([_sample_row()], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    # Col F = 6
    assert ws.cell(row=2, column=6).number_format == "#,##0"


def test_data_validators_are_added():
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    assert len(ws.data_validations.dataValidation) == 2


def test_no_background_fill_on_headers():
    """Headers must NOT have a coloured fill."""
    exp = _make_exporter()
    period = _compliance_period()
    wb = exp._build_workbook([], period)
    ws = wb[FSE_UPDATE_SHEETNAME]
    for col in range(1, len(HEADERS) + 1):
        fill = ws.cell(row=1, column=col).fill
        pattern = getattr(fill, "patternType", None)
        assert pattern in (None, "none"), (
            f"Header col {col} has unexpected fill: {pattern}"
        )


# ---------------------------------------------------------------------------
# _load_fse_data
# ---------------------------------------------------------------------------


@pytest.mark.anyio
async def test_load_fse_data_site_name_is_first_column():
    """site_name must be col 0 in every row tuple."""
    row = _fse_row(site_name="My Site")
    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")
    assert rows[0][0] == "My Site"


@pytest.mark.anyio
async def test_load_fse_data_registration_is_second_column():
    row = _fse_row(registration="0000A-001")
    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")
    assert rows[0][1] == "0000A-001"


@pytest.mark.anyio
async def test_load_fse_data_serial_is_third_column():
    row = _fse_row(serial_number="SN-99")
    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")
    assert rows[0][2] == "SN-99"


@pytest.mark.anyio
async def test_load_fse_data_converts_datetime_to_date():
    """datetime objects from the DB must be converted to plain date."""
    dt_from = datetime.datetime(2024, 3, 15, 0, 0, 0)
    dt_to = datetime.datetime(2024, 9, 30, 23, 59, 59)
    row = _fse_row(supply_from=dt_from, supply_to=dt_to)

    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")

    assert rows[0][3] == datetime.date(2024, 3, 15)
    assert rows[0][4] == datetime.date(2024, 9, 30)


@pytest.mark.anyio
async def test_load_fse_data_kwh_none_becomes_zero_for_active_rows():
    """Active rows with kwh_usage=None must write 0, not None."""
    row = _fse_row(kwh_usage=None)
    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")
    assert rows[0][5] == 0


@pytest.mark.anyio
async def test_load_fse_data_inactive_row_shows_only_site_and_reg():
    """Inactive rows (is_active=False) must have None for all data columns."""
    row = _fse_row(is_active=False, kwh_usage=500, notes="should be hidden")
    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")

    assert rows[0][0] == "Charge Site 1"   # site name preserved
    assert rows[0][1] == "ORG-AAAA1A-001"  # reg # preserved
    assert rows[0][2] == "SN-12345"        # serial # preserved
    assert rows[0][3] is None              # from date blank
    assert rows[0][4] is None              # to date blank
    assert rows[0][5] is None              # kWh blank
    assert rows[0][6] is None              # notes blank


@pytest.mark.anyio
async def test_load_fse_data_other_report_group_shows_only_site_and_reg():
    """Rows belonging to a different compliance report group must also be blanked."""
    row = _fse_row(compliance_report_group_uuid="other-group", kwh_usage=999)
    exporter = _make_exporter(fse_rows=[row])
    # Current report group is "group-abc"; row belongs to "other-group"
    rows = await exporter._load_fse_data(1, "group-abc")

    assert rows[0][0] == "Charge Site 1"
    assert rows[0][1] == "ORG-AAAA1A-001"
    assert rows[0][2] == "SN-12345"        # serial # preserved
    assert rows[0][3] is None
    assert rows[0][5] is None


@pytest.mark.anyio
async def test_load_fse_data_empty_registration_becomes_empty_string():
    row = _fse_row(registration=None)
    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")
    assert rows[0][1] == ""


@pytest.mark.anyio
async def test_load_fse_data_empty_site_name_becomes_empty_string():
    row = _fse_row(site_name=None)
    exporter = _make_exporter(fse_rows=[row])
    rows = await exporter._load_fse_data(1, "group-abc")
    assert rows[0][0] == ""
